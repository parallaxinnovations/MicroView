# =========================================================================
#
# Copyright (c) 2000-2002 Enhanced Vision Systems
# Copyright (c) 2002-2008 GE Healthcare
# Copyright (c) 2011-2015 Parallax Innovations Inc.
#
# Use, modification and redistribution of the software, in source or
# binary forms, are permitted provided that the following terms and
# conditions are met:
#
# 1) Redistribution of the source code, in verbatim or modified
#   form, must retain the above copyright notice, this license,
#   the following disclaimer, and any notices that refer to this
#   license and/or the following disclaimer.
#
# 2) Redistribution in binary form must include the above copyright
#    notice, a copy of this license and the following disclaimer
#   in the documentation or with other materials provided with the
#   distribution.
#
# 3) Modified copies of the source code must be clearly marked as such,
#   and must not be misrepresented as verbatim copies of the source code.
#
# EXCEPT WHEN OTHERWISE STATED IN WRITING BY THE COPYRIGHT HOLDERS AND/OR
# OTHER PARTIES, THE COPYRIGHT HOLDERS AND/OR OTHER PARTIES PROVIDE THE
# SOFTWARE "AS IS" WITHOUT EXPRESSED OR IMPLIED WARRANTY INCLUDING, BUT
# NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS
# FOR A PARTICULAR PURPOSE.  IN NO EVENT UNLESS AGREED TO IN WRITING WILL
# ANY COPYRIGHT HOLDER OR OTHER PARTY WHO MAY MODIFY AND/OR REDISTRIBUTE
# THE SOFTWARE UNDER THE TERMS OF THIS LICENSE BE LIABLE FOR ANY DIRECT,
# INDIRECT, INCIDENTAL OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED
# TO, LOSS OF DATA OR DATA BECOMING INACCURATE OR LOSS OF PROFIT OR
# BUSINESS INTERRUPTION) ARISING IN ANY WAY OUT OF THE USE OR INABILITY TO
# USE THE SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGES.
#
# =========================================================================

#
# This file represents a derivative work by Parallax Innovations Inc.
#

"""
MicroView is the main class.  MicroView allows the visualization and quantification of image data.
"""

import logging
import logging.handlers
import encodings
import sys
import time


class MicroViewModuleLogger(object):

    """meta module loader class - useful for providing user feedback when loading modules"""

    def __init__(self):
        self._count = 0
        self._max_count = 2196
        self._time = time.time()
        self._enabled = True

    def enable(self):
        self._enabled = True

    def disable(self):
        self._enabled = False

    def find_module(self, name, path=None):
        self._count += 1
        t = time.time()
        if t - self._time > 0.5:
            self._time = t
            # emit a log message
            if self._enabled:
                progress = 100.0 * float(self._count) / self._max_count
                logging.getLogger("splash-progress").info("%0.1f" % progress)
        return None

microview_module_logger = MicroViewModuleLogger()
sys.meta_path.insert(0, microview_module_logger)

########################################
logging.getLogger("splash").info("Loading VTK...")
import vtk
from vtkAtamai import ClippingCubeFactory

########################################
logging.getLogger("splash").info("Loading App...")

import collections
import datetime
import json
import math
import numpy
from setuptools.command import easy_install
import re
import string
import subprocess
import platform
import urllib
import urllib2
import time
import os
from xml.etree.ElementTree import fromstring, tostring
import webbrowser

# enable zope events
# don't delete this line - events won't work otherwise -- why?
import zope.component.event
from zope import interface, component, event
# Needed for packaging on ubuntu?
import zc

import thread
from twisted.internet import reactor

from vtkEVS import EVSFileDialog, EVSSphereMarkerListFactory, TitleBar

from PI.egg.pkg_resources import working_set, Environment

from PI.visualization.MicroView.interfaces import IMicroViewMainFrame, IImageProcessor, \
    IStockIconProvider, IViewportManager, IApplicationPreferencesManager, \
    IPluginManager, IResultWindow, IViewerState, \
    IOrthoPlanes, ICurrentOrthoPlanes, \
    IOrthoView, ICurrentViewportManager, \
    ICurrentOrthoView, IImage, ICurrentImage, \
    IMicroViewInput, IMicroViewOutput, ICurrentMicroViewInput, ISpreadsheet
from PI.visualization.MicroView.events import StatusbarCursorXYPosition, StatusbarCursorSlicePosition, \
    StatusbarVoxelValue, ViewportMouseDown, \
    PlotWindowCursorPosition, HelpEvent, \
    HistogramSelectionModifiedEvent, BinSizeModifiedEvent, \
    GlobalHistogramChangedEvent, \
    MotionMonitoringDisableRequestedEvent, \
    MotionMonitoringEnableRequestedEvent, \
    WindowLevelTextChangeEvent, ScalesVisibilityChangeEvent, ColourBarVisibilityChangeEvent, \
    CurrentImageChangeEvent, ROIKeyEvent, \
    ROIEnabledEvent, ROIDisabledEvent, HistogramClosedEvent, \
    AutoThresholdCommandEvent, GetROIStencilEvent, \
    OrthoPlanesRemoveInputEvent, DisableMotionMonitoring,  \
    EnableMotionMonitoring, XSliceValueChangeEvent, YSliceValueChangeEvent, ZSliceValueChangeEvent,\
    NoImagesLoadedEvent, ChangeImageLUT, ImageLUTChanged, \
    HistogramIsActiveROIEvent, ROIModifiedEvent, ShowWindowLevelDialogEvent, \
    NotebookPageChangingEvent, NotebookPageChangedEvent, \
    CurrentImageClosingEvent, WindowLevelPresetEvent, SynchronizeEvent, \
    TrackedSliceIndexChangeEvent

from PI.visualization.common.events import ProgressEvent, ConfigModifiedEvent, MicroViewSettingsModifiedEvent
from PI.visualization.common import MicroViewHistory, MicroViewSettings, MicroViewObjectShelve, wxWindowLevelControls
# this is needed by plugins
from PI.visualization.common.PluginHelper import install_package
from PI.visualization import MicroView  # pull in package version
from PI.wx import AutoResizeListCtrl

from PI.visualization.vtkMultiIO import vtkImageReaderBase, MVImage

# Adjust our import directory list to include a user-writable area for
# pip/easy_install interactions
import appdirs
import glob
data_dir = appdirs.user_data_dir("MicroView", "Parallax Innovations")
egg_dir = os.path.join(data_dir, 'eggs')
plugins_dir = os.path.join(data_dir, 'Plugins')
for p in (egg_dir, plugins_dir):
    if not os.path.exists(p):
        logging.info("creating directory {}".format(p))
        os.makedirs(p)

# Add user-writeable eggs to our import path
# Use egg mechanisms directly to handle versioning
environment = Environment([egg_dir])
distributions, errors = working_set.find_plugins(environment)
for d in distributions:
    if d.location not in sys.path:
        sys.path.append(d.location)

# MicroView classes import
try:
    import UpstreamMessageDialogGUI
except:
    pass
import MicroViewIO
import StockItems
import OrthoView
import HelpDictionary
import ROIObject
import MicroViewModePalette
import MicroViewImageProcessor
import MicroViewStatusBar
import SpectrumPlotWindow
import VTKPlotWindow
import ViewportManager
import SliceViewportManager
import StatisticsGUIC
import LogWindowGUIC
import ResultsWindowGUIC
import UpdateAvailableGUIC
import SpreadsheetGUIC
import AnisotropicSmoothingDialogC
import AdaptiveOtsuThresholdDialogC
import ApplicationPreferencesDialogC
import ChooseFilterRadiusDialogC
import shutil
import data
import LUTSelectionDialogC
import SynchronizeImageDialogC
import OpenCloseMorphologyDialogC
import ErodeDilateMorphologyDialogC

from PluginManager import *
import ROIStatistics

from PI.visualization.MicroView import _MicroView
o = _MicroView.vtkStderrOutputWindow()
o.SetInstance(o)
from PI.wx import wxLogHandler

import wx
import wx.lib.agw.aui
import wx.lib.agw.persist as PM
import wx.lib.agw.shortcuteditor as SE
from wx.lib.agw.aui.auibook import TabFrame, AuiTabCtrl
from wx.lib.agw.aui import framemanager
import wx.lib.wordwrap
import gc
import locale

# these allow us to read csv and xls files (why use one package when we
# can use 4?)
import xlrd
import xlwt
import openpyxl
import openpyxl.xml.functions
import unicodecsv


# plugins require these
import decorator
import struct
import pyfits as fits  # we almost don't need this (2015-01-06)
import uuid
import wx.aui
import wx.lib.combotreebox
import wx.lib.intctrl
import wx.lib.mixins.listctrl
import wx.py.crust
import netdicom
import yaml
import PIL.ExifTags
from PIL import Image
import scipy.io
import scipy.ndimage
import scipy.special._ufuncs_cxx
import skimage
import skimage.measure
import skimage.morphology
import skimage._shared.geometry

import skimage.color.colorlabel  # docker builds on Ubuntu 14.04 need this
import skimage.filters.rank.core_cy
from PI.cache import cache
from PI.IO import vff
from PI.IO import CncrdPETFile
from PI.dicom import convert
from PI.twisted import Zeroconf

import twisted.web.xmlrpc

try:
    import scipy.sparse.csgraph
except:
    logging.warning("Unable to import scipy.sparse.csgraph")
from vtk.util import vtkImageImportFromArray, vtkImageExportToArray

import gdcm
import vtkgdcm

from PI.visualization.MicroView import PACKAGE_VERSION
from PI.visualization.MicroView import PACKAGE_SHA1


##########################################################################
def setDropTargets(parent, dtclass):
    if hasattr(parent, "SetDropTarget"):
        dt = dtclass(parent)
        parent.SetDropTarget(dt)
    if hasattr(parent, "Children"):
        for child in parent.Children:
            setDropTargets(child, dtclass)


class FileDrop(wx.FileDropTarget):

    def __init__(self, window):
        wx.FileDropTarget.__init__(self)
        self.window = window

    def OnDropFiles(self, x, y, filenames):
        pass


class XLSReader(object):

    """Hides some of the difference between xlrd and openpyxl"""

    def __init__(self, filename):
        self._filename = filename
        self._book = None
        self.handle_format = True
        self.font_list = []
        self.ncols = 0
        self.nrows = 0
        self.open_book()
        self._sheet = self.get_active_sheet()
        self.update_dimensions()

    def get_sheet(self):
        return self._sheet

    def get_cell(self, r, c):

        cell = self._sheet.cell(r, c)

        val = cell.value
        fg = bg = font = None
        xf_index = cell.xf_index

        # check and convert data type
        if cell.ctype == xlrd.XL_CELL_DATE:
            # assume a windows date - this may cause problems
            val = datetime.datetime(
                *xlrd.xldate_as_tuple(cell.value, 0)).isoformat()

        if xf_index is not None:
            _format = self._book.xf_list[xf_index]
            font_index = _format.font_index
            if font_index:
                font = self._book.font_list[font_index]
            fg = self._book.colour_map[
                self._book.font_list[font_index].colour_index]
            bg = self._book.colour_map[_format.background.pattern_colour_index]

        return val, font, fg, bg

    def open_book(self):
        try:
            self._book = xlrd.open_workbook(
                self._filename, formatting_info=True)
        except NotImplementedError:
            self.handle_format = False
            self._book = xlrd.open_workbook(filename)

        # create font list if we support such things
        if self.handle_format:
            self.font_list = []
            for f in self._book.font_list:
                point_size = f.height / 20
                family = wx.FONTFAMILY_DEFAULT
                style = wx.FONTSTYLE_NORMAL
                weight = wx.FONTWEIGHT_NORMAL
                underlined = False
                facename = ''
                if f.italic:
                    style = wx.FONTSTYLE_ITALIC
                if f.bold:
                    weight = wx.FONTWEIGHT_BOLD
                if f.underlined:
                    underlined = True
                if f.name:
                    facename = f.name
                wx_font = wx.Font(
                    point_size, family, style, weight, underlined, facename)
                self.font_list.append(wx_font)

    def get_active_sheet(self):
        sheet_names = self._book.sheet_names()
        sheet = self._book.sheet_by_name(sheet_names[0])
        return sheet

    def update_dimensions(self):
        sheet = self._sheet
        self.ncols = sheet.ncols
        self.nrows = sheet.nrows


class XLSXReader(XLSReader):

    def open_book(self):
        self._book = openpyxl.load_workbook(self._filename)

        self.colors = []

        # parse theme
        if self._book.loaded_theme:
            xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
            root = openpyxl.xml.functions.fromstring(self._book.loaded_theme)
            themeEl = root.find(
                openpyxl.xml.functions.QName(xlmns, 'themeElements').text)
            colorSchemes = themeEl.findall(
                openpyxl.xml.functions.QName(xlmns, 'clrScheme').text)
            firstColorScheme = colorSchemes[0]

            for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
                accent = firstColorScheme.find(
                    openpyxl.xml.functions.QName(xlmns, c).text)

                if 'window' in accent.getchildren()[0].attrib['val']:
                    self.colors.append(
                        accent.getchildren()[0].attrib['lastClr'])
                else:
                    self.colors.append(accent.getchildren()[0].attrib['val'])

    def get_active_sheet(self):
        return self._book.active

    def hex_to_rgb(self, value):
        hexcolor = int(value, 16)
        r = (hexcolor >> 16) & 0xFF
        g = (hexcolor >> 8) & 0xFF
        b = hexcolor & 0xFF
        return r, g, b

    def get_cell(self, r, c):

        cell = self._sheet.cell(row=r + 1, column=c + 1)
        val = cell.value

        # check and convert data type
        if cell.is_date:
            val = val.isoformat()

        font = fg = bg = None

        if self.colors and cell.style.font.color.type == 'theme':
            fg = self.hex_to_rgb(self.colors[cell.style.font.color.value])
        else:
            fg = self.hex_to_rgb(cell.font.color.value[2:])

        if cell.fill.patternType is not None:
            if cell.fill.bgColor.type == 'theme':
                bg = self.hex_to_rgb(
                    self.colors[cell.style.fill.bgColor.value])
            elif cell.fill.bgColor.type == 'rgb':
                bg = self.hex_to_rgb(cell.fill.bgColor.value[2:])

        return val, font, fg, bg

    def update_dimensions(self):
        sheet = self._sheet
        ncols = 0
        for col in sheet.rows[0]:
            if col.value is None:
                break
            else:
                ncols += 1

        self.nrows = len(sheet.rows)
        self.ncols = ncols


class MicroViewMainFrame(wx.Frame, wx.FileDropTarget):

    interface.implements(IMicroViewMainFrame)

    def __del__(self):
        logging.debug('MicroViewMainFrame deleted!')
        # unregister our zope handlers
        gsm = component.getGlobalSiteManager()
        gsm.unregisterHandler(self.onVoxelValueTextChange)
        gsm.unregisterHandler(self.onCursorPositionTextChange)
        gsm.unregisterHandler(self.onCursorSliceTextChange)
        gsm.unregisterHandler(self.onPlotWindowCursorPositionChange)
        gsm.unregisterHandler(self.ProgressMethod)
        gsm.unregisterHandler(self.OnUserHelp)
        gsm.unregisterHandler(self.onHistogramSelectionModified)
        gsm.unregisterHandler(self.onHistogramIsActiveROIEvent)
        gsm.unregisterHandler(self.OnConfigModified)
        gsm.unregisterHandler(self.onHistogramBinSizeChangeEvent)
        gsm.unregisterHandler(self.onMotionMonitoringDisableRequestedEvent)
        gsm.unregisterHandler(self.onMotionMonitoringEnableRequestedEvent)
        gsm.unregisterHandler(self.onWindowLevelTextChange)
        gsm.unregisterHandler(self.onROIEnabledEvent)
        gsm.unregisterHandler(self.onROIDisabledEvent)
        gsm.unregisterHandler(self.onROIModifiedEvent)
        gsm.unregisterHandler(self.onHistogramClosedEvent)
        gsm.unregisterHandler(self.onAutoThresholdCommandEvent)
        gsm.unregisterHandler(self.onGetROIStencilEvent)
        gsm.unregisterHandler(self.onOrthoPlanesRemoveInputEvent)
        gsm.unregisterHandler(self.onEnableMotionMonitoring)
        gsm.unregisterHandler(self.onDisableMotionMonitoring)
        gsm.unregisterHandler(self.onXSliceValueChanged)
        gsm.unregisterHandler(self.onYSliceValueChanged)
        gsm.unregisterHandler(self.onZSliceValueChanged)
        gsm.unregisterHandler(self.onChangeImageLUT)
        gsm.unregisterHandler(self.onShowWindowLevelDialogEvent)
        gsm.unregisterHandler(self.onViewportMouseDown)
        gsm.unregisterHandler(self.onTrackedSliceIndexChangeEvent)

    def convert_version_string(self, s):
        """
        converts a git tag string to a comparable number

        handle A.B.C-D and A.B.C
        """

        try:
            vals = re.compile(
                '([0-9]*)\.([0-9]*)\.([0-9]*)(-(.*))*').match(s).groups()
            if vals[-1] is None:
                vals = vals[:-2]
                vals = map(int, vals)
                val = vals[0] * 10000 ** 3 + vals[
                    1] * 10000 ** 2 + vals[2] * 10000 ** 1
            else:
                vals = map(int, vals)
                val = vals[0] * 10000 ** 3 + vals[
                    1] * 10000 ** 2 + vals[2] * 10000 ** 1 + vals[-1]
        except:
            logging.error("Unable to parse version string '%s'" % s)
            val = -1

        return val

    def handle_error_response(self, response):
        logging.exception(response)

    def __init__(self, splash, **kw):

        # get a configuration object
        self._config = MicroViewSettings.MicroViewSettings.getObject()

        # workout size
        w, h = wx.DisplaySize()
        w -= 100
        h -= 100

        # init wxFrame
        wx.Frame.__init__(self, None, -1, '', size=(w, h),
                          style=wx.DEFAULT_FRAME_STYLE | wx.NO_FULL_REPAINT_ON_RESIZE)

        # init drop target
        wx.FileDropTarget.__init__(self)
        self.SetDropTarget(self)

        self.splash = splash
        self.pane3D = None
        self.pane2d_1 = None
        self.pane2d_2 = None
        self.pane2d_3 = None

        # plot deferred drawing stuff
        self._delayed_plot_update_cb = None
        self._last_delayed_plot_redraw = 0

        # check if the /3GB system start option is on using the registry key
        if sys.platform == 'win32' and (sys.maxsize == 2 ** 31 - 1):
            self._the_3GB_switch = False
            try:
                import _winreg
                key = _winreg.OpenKey(
                    _winreg.HKEY_LOCAL_MACHINE, "SYSTEM\CurrentControlSet\Control")
                val = _winreg.QueryValueEx(key, "SystemStartOptions")
                if '3GB' in val[0].split():
                    self._the_3GB_switch = True
            except:
                pass

        # Line profile tool
        self.LineSegment = vtk.vtkLineSource()

        # create a probe filter
        self.probe = vtk.vtkProbeFilter()

        # connect probe and line segment
        self.probe.SetInputConnection(self.LineSegment.GetOutputPort())

        # create the subvolume selection actor for doing histograms
        self._ROIStats = ROIStatistics.ROIStatistics()

        self.install_dir = os.path.abspath(sys.path[0])

        # keep track of how many images are loaded
        self._numberOfImages = 0
        self._numberOfCurrentImagesDisplayed = 0
        self._current_image = 0
        self._current_title = ''

        self._numberOfSpreadsheets = 0
        self._numberOfCurrentSpreadsheetsDisplayed = 0
        self._current_spreadsheet = 0

        # stock item manager
        self._stockicons = StockItems.StockIconFactory()

        self.pluginManager = None

        if hasattr(sys, 'frozen'):
            self.html_dir = os.path.join(self.install_dir, 'doc')
        else:
            self.html_dir = os.path.join(
                self.install_dir, '..', '..', '..', 'docs', 'sphinx_html')

        if sys.platform != 'darwin':
            if not os.access(self.html_dir, os.R_OK):  # just in case
                self.html_dir = self.install_dir
        else:
            pass

        self._helpDict = HelpDictionary.HelpDictionary()
        self._helpDict.SetDictionaryFile(os.path.join(
            self.html_dir, 'HelpDictionary'))

        self._nImageIdx = 0
        self._maskValue = None
        self.gaussian_radius = 3
        self.median_filter_kernel_size = 3
        self._OpenCloseMorphologyDialog = None
        self._threshold_dlg = None

        self._plotwindow = self._histowindow = self._spectrumwindow = self._aboutDialog = None
        self._plotwindow_title = "Plot Line Profile"
        self._histowindow_title = "Plot Line Profile"

        # mean/std dev text window
        self._statswindow = None
        self._statswindow_labels = {}

        # create default image transform
        self.title = TitleBar.TitleBar()
        self.title.SetTemplate("MicroView - %s")

        self.SetTitle(self.title.GetTitle())
        self.SetRepresentedFilename(self.title.GetTitle() or '')

        # Handle window closes
        wx.EVT_CLOSE(self, self.Quit)

        self.mviewOut = MicroViewIO.MicroViewOutput()
        component.getGlobalSiteManager().registerUtility(
            self.mviewOut, IMicroViewOutput)

        # ------------------------------------------------------------------------------------
        # Some objects have a large scoping requirement: use zope's component model to keep
        # track of these objects
        # ---------------------------------------------------------------------

        # start by registering ourselves with the object registry
        component.getGlobalSiteManager().registerUtility(
            self, IMicroViewMainFrame)

        # Register a MicroView Image Processing object with zope
        obj = MicroViewImageProcessor.MicroViewImageProcessor()
        component.getGlobalSiteManager().registerUtility(
            obj, IImageProcessor)

        # Add stock icon object to registry
        component.getGlobalSiteManager().registerUtility(
            self._stockicons, IStockIconProvider)

        # ------------------------------------------------------------------------------------
        # Set up menu bar for the program.
        # ---------------------------------------------------------------------
        self.menubar = self.CreateMenuBar()

        # ------------------------------------------------------------------------------------
        # Set up status bar for the program.
        # ------------------------------------------------------------------------------------

        self.statusbar = MicroViewStatusBar.MicroViewStatusBar(self)
        self.SetStatusBar(self.statusbar)
        self.statusbar.SetName('StatusBar1')

        # create a AUI paned widget manager
        self._mgr = wx.lib.agw.aui.AuiManager()

        # set flags
        agwFlags = self._mgr.GetAGWFlags()
        agwFlags ^= wx.lib.agw.aui.AUI_MGR_AERO_DOCKING_GUIDES
        agwFlags &= ~wx.lib.agw.aui.AUI_MGR_WHIDBEY_DOCKING_GUIDES
        agwFlags ^= wx.lib.agw.aui.AUI_MGR_ALLOW_ACTIVE_PANE

        self._mgr.SetAGWFlags(agwFlags)
        self._mgr.SetManagedWindow(self)

        # activate persistence manager
        self._persistMgr = PM.PersistenceManager.Get()

        # determine default filename
        _dir = appdirs.user_data_dir("MicroView", "Parallax Innovations")
        self._persistMgr.SetPersistenceFile(
            os.path.join(_dir, 'Persistence_Options'))

        self.leftframe = wx.Panel(self)

        sizer = wx.BoxSizer(wx.VERTICAL)
        self.leftframe.SetSizer(sizer)

        self._mgr.AddPane(self.leftframe, wx.lib.agw.aui.AuiPaneInfo().Left().Name("AppPane").Caption(
            "Applications & Tools").MinimizeButton(True).CloseButton(True).BestSize(
            (350, 600)).FloatingSize((350, 600)))

        # create a panel to hold views and window/level/slice controls
        main_panel = wx.Panel(self)
        main_panel.SetName("MainPanel1")
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        main_panel.SetSizer(main_sizer)

        # create a notebook to hold viewport managers
        self._view_notebook = wx.lib.agw.aui.AuiNotebook(main_panel)
        main_sizer.Add(self._view_notebook, 1, wx.EXPAND)

        style = wx.lib.agw.aui.AUI_NB_DEFAULT_STYLE | wx.lib.agw.aui.AUI_NB_TAB_MOVE |\
            wx.lib.agw.aui.AUI_NB_TAB_EXTERNAL_MOVE | wx.lib.agw.aui.AUI_NB_TAB_SPLIT |\
            wx.lib.agw.aui.AUI_NB_SUB_NOTEBOOK | wx.lib.agw.aui.AUI_NB_CLOSE_ON_ALL_TABS

        style ^= aui.AUI_NB_WINDOWLIST_BUTTON
        self._view_notebook.SetAGWWindowStyleFlag(style)

        self._view_notebook.Bind(
            aui.EVT_AUINOTEBOOK_PAGE_CLOSE, self.OnNotebookPageClose)
        self._view_notebook.Bind(
            aui.EVT_AUINOTEBOOK_PAGE_CHANGING, self.OnNotebookPageChanging)
        self._view_notebook.Bind(
            aui.EVT_AUINOTEBOOK_PAGE_CHANGED, self.OnNotebookPageChanged)
        self._view_notebook.Bind(
            aui.EVT_AUINOTEBOOK_TAB_RIGHT_DOWN, self.OnNotebookRightClick)
        self._mgr.AddPane(main_panel, wx.lib.agw.aui.AuiPaneInfo().CenterPane().Name(
            "Viewers").Caption("Viewers").CaptionVisible(True).MaximizeButton(True).CloseButton(False))

        # create a bottom notebook for logging, results, plotting, histogram
        # etc.
        self._bottom_notebook = wx.lib.agw.aui.AuiNotebook(self)

        style = wx.lib.agw.aui.AUI_NB_DEFAULT_STYLE | wx.lib.agw.aui.AUI_NB_TAB_EXTERNAL_MOVE | \
            wx.lib.agw.aui.AUI_NB_TAB_MOVE | wx.lib.agw.aui.AUI_NB_BOTTOM

        # disable close buttons
        style &= ~(wx.lib.agw.aui.AUI_NB_CLOSE_BUTTON |
                   wx.lib.agw.aui.AUI_NB_CLOSE_ON_ACTIVE_TAB | wx.lib.agw.aui.AUI_NB_CLOSE_ON_ALL_TABS)

        self._bottom_notebook.SetAGWWindowStyleFlag(style)

        self._mgr.AddPane(self._bottom_notebook, wx.lib.agw.aui.AuiPaneInfo().Bottom().Name("Log").Caption("Logs & Analysis").
                          CaptionVisible(True).MinimizeButton(True).MaximizeButton(True).CloseButton(True).Hide().
                          BestSize((800, 200)).FloatingSize((800, 600)))

        # listen to changes to the logging notebook displayed page
        self._bottom_notebook.Bind(
            aui.EVT_AUINOTEBOOK_PAGE_CHANGED, self.OnBottomNotebookPageChanged)

        # create a window for logging
        self.logPane = LogWindowGUIC.LogWindowGUIC(self)
        self.logPane.SetName("LogPane1")

        self.logger = wx.LogTextCtrl(self.logPane.GetTextControl())
        self._bottom_notebook.AddPage(self.logPane, "Log")
        wx.Log.SetActiveTarget(self.logger)

        # create a window for results
        self.resultsPane = ResultsWindowGUIC.ResultsWindowGUIC(self)
        self._bottom_notebook.AddPage(self.resultsPane, "Results")

        # create application preferences page
        dlg = ApplicationPreferencesDialogC.ApplicationPreferencesDialogC(self)
        component.getGlobalSiteManager().registerUtility(
            dlg, IApplicationPreferencesManager)

        # register object
        component.getGlobalSiteManager().registerUtility(
            self.resultsPane, IResultWindow)

        # setup a handler to capture all standard logging info and set it to
        # the 'log' window
        logr = logging.getLogger('')
        logr.setLevel(logging.DEBUG)
        hdlr = wxLogHandler.wxLogHandler(self.logPane.GetTextControl())

        hdlr.setFormatter(logging.Formatter(
            fmt='%(asctime)s [%(levelname)s]:  %(message)s', datefmt='[%Y-%m-%dT%H:%M:%S]'))
        logr.addHandler(hdlr)

        # setup a handler to capture results logging info and send this to the
        # 'results' window
        logr = logging.getLogger('results')
        logr.setLevel(logging.DEBUG)
        hdlr = wxLogHandler.wxLogHandler(self.resultsPane.GetTextControl())
        hdlr.setFormatter(logging.Formatter(
            fmt='%(asctime)s:  %(message)s', datefmt='[%Y-%m-%dT%H:%M:%S]'))
        logr.addHandler(hdlr)

        # setup a rotating file logger too
        logr = logging.getLogger('')
        _dir = os.path.join(
            appdirs.user_data_dir("MicroView", "Parallax Innovations"), "Logs")
        if not os.path.exists(_dir):
            os.makedirs(_dir)
        log_filename = os.path.join(_dir, "MicroView.log")
        hdlr = logging.handlers.TimedRotatingFileHandler(
            log_filename, when='D', interval=1, backupCount=2)
        hdlr.setFormatter(logging.Formatter(
            fmt='%(asctime)s [%(levelname)s] [%(module)s.%(funcName)s]:  %(message)s', datefmt='[%Y-%m-%dT%H:%M:%S]'))
        logr.addHandler(hdlr)

        # log software version here
        logging.info("MicroView Version: %s" % (PACKAGE_VERSION))
        logging.info("MicroView SHA1: %s" % (PACKAGE_SHA1))

        # Add a mode palette aui toolbar for switching mouse interaction mode
        self.modePalettePane = MicroViewModePalette.MicroViewModePalette(self)
        self.modePalettePane.SetName("modePalettePane1")
        self._mgr.AddPane(self.modePalettePane, wx.lib.agw.aui.AuiPaneInfo().Name("modePalettePane").Caption("Mouse Mode Palette").
                          ToolbarPane().Bottom().Row(1).Position(1))

        # Add a window/level slider control aui toolbar for controlling image
        # window/level values
        self.WindowLevelToolbar = wxWindowLevelControls.wxWindowLevelControls(
            main_panel)
        self.WindowLevelToolbar.SetName("WindowLevelToolbar1")
        main_sizer.Add(self.WindowLevelToolbar, 0, wx.EXPAND)

        self.WindowLevelToolbar.GetSliceSlider().Bind(
            wx.EVT_SLIDER, self.onSliceChanged)

        # self.WindowLevelToolbar.GetZSliceLabel()

        # ------------------------------------------------------------------------------------
        # Set up some zope event handlers
        # ---------------------------------------------------------------------
        component.provideHandler(self.onVoxelValueTextChange)
        component.provideHandler(self.onCursorPositionTextChange)
        component.provideHandler(self.onCursorSliceTextChange)
        component.provideHandler(self.onPlotWindowCursorPositionChange)
        component.provideHandler(self.ProgressMethod)
        component.provideHandler(self.OnUserHelp)
        component.provideHandler(self.onHistogramSelectionModified)
        component.provideHandler(self.onHistogramIsActiveROIEvent)
        component.provideHandler(self.OnConfigModified)
        component.provideHandler(self.onHistogramBinSizeChangeEvent)
        component.provideHandler(self.onMotionMonitoringDisableRequestedEvent)
        component.provideHandler(self.onMotionMonitoringEnableRequestedEvent)
        component.provideHandler(self.onWindowLevelTextChange)
        component.provideHandler(self.onROIEnabledEvent)
        component.provideHandler(self.onROIDisabledEvent)
        component.provideHandler(self.onROIModifiedEvent)
        component.provideHandler(self.onHistogramClosedEvent)
        component.provideHandler(self.onAutoThresholdCommandEvent)
        component.provideHandler(self.onGetROIStencilEvent)
        component.provideHandler(self.onOrthoPlanesRemoveInputEvent)
        component.provideHandler(self.onEnableMotionMonitoring)
        component.provideHandler(self.onDisableMotionMonitoring)
        component.provideHandler(self.onXSliceValueChanged)
        component.provideHandler(self.onYSliceValueChanged)
        component.provideHandler(self.onZSliceValueChanged)
        component.provideHandler(self.onChangeImageLUT)
        component.provideHandler(self.onShowWindowLevelDialogEvent)
        component.provideHandler(self.onViewportMouseDown)
        component.provideHandler(self.onTrackedSliceIndexChangeEvent)

        # create first image page
        self.CreateImageTab()
        self._mgr.Update()

        # show busy cursor
        with wx.BusyCursor():

            self.pluginManager = PluginManager(self.leftframe, width=200)

            # keep track of plugin manager
            component.getGlobalSiteManager().registerUtility(
                self.pluginManager, IPluginManager)

            self.leftframe.GetSizer().Add(self.pluginManager, 1, wx.EXPAND)
            self.pluginManager.SetMenu(self.menubar)
            self.pluginManager.LoadPlugins()

            # we always want the ROIManager plugin running, so start it up from
            # the beginning
            # self.pluginManager.GetPluginReferenceByName('ROIManager')

        # magic VTK sauce to avoid polygon coincidence flickering
        vtk.vtkMapper.SetResolveCoincidentTopologyToPolygonOffset()
        vtk.vtkMapper.SetResolveCoincidentTopologyPolygonOffsetParameters(1, 1)

        self.DisplayView("All")

        # register widgets with persistence manager
        reactor.callWhenRunning(self.RegisterControls)

    def RegisterControls(self):
        self.Freeze()
        self.SetFocus()  # otherwise an aui bug occurs
        self.Register()
        self.Thaw()

        # replay button event here (act as if user just pressed mode palette button)
        # this makes new images behave correctly
        self.modePalettePane.activateMode(self.modePalettePane.GetMode())

    def Register(self, children=None):

        if children is None:
            self._persistMgr.RegisterAndRestore(self)
            self._persistMgr.RegisterAndRestore(self.GetMenuBar())
            children = self.GetChildren()

        for child in children:

            name = child.GetName()
#            if name not in PM.BAD_DEFAULT_NAMES and "widget" not in name and \
#                "StatusBar" not in name and \
#               "wxSpinButton" not in name:
            if name not in PM.BAD_DEFAULT_NAMES and "widget" not in name and \
               "wxSpinButton" not in name and "auiFloating" not in name and \
               "AuiTabCtrl" not in name and "AuiNotebook" not in name and "floatslider" not in name and \
               "StatusBar" not in name:
                try:
                    self._persistMgr.RegisterAndRestore(child)
                except Exception, e:
                    print 'unable to restore:', child.GetName()
                    print e

            if child.GetChildren():
                try:
                    self.Register(child.GetChildren())
                except Exception, e:
                    print 'unable to restore:', child.GetName()
                    print e

    def CreateImageTab(self, image=None, mviewIn=None, filename=None, title=None):

        # Increase image counter
        self._numberOfImages += 1
        self._numberOfCurrentImagesDisplayed += 1

        # create a default image and register it with the system
        if image is None:
            mviewIn = MicroViewIO.MicroViewInput()
            image = MicroViewIO.CreateDefaultImage()

        component.getGlobalSiteManager().registerUtility(
            image, IImage, name='Image-%d' % self._numberOfImages)
        component.getGlobalSiteManager().registerUtility(
            mviewIn, IMicroViewInput, name='MicroViewInput-%d' % self._numberOfImages)

        self.bIsMicroView = True

        if self.bIsMicroView:

            # create a new viewport manager
            viewportManager = ViewportManager.ViewportManager(
                self, index=self._numberOfImages)

        else:
            # SliceView
            viewportManager = SliceViewportManager.SliceViewportManager(
                self, index=self._numberOfImages)

        # freeze it for now
        viewportManager.Freeze2()

        # Add this viewport object to registry; keep track of the number
        component.getGlobalSiteManager().registerUtility(
            viewportManager, IViewportManager, name='ViewportManager-%d' % self._numberOfImages)

        if self.bIsMicroView:

            # create an orthoview object for this viewport
            orthoView = OrthoView.OrthoView(
                viewportManager, index=self._numberOfImages)

            with wx.WindowDisabler():
                with wx.BusyCursor():
                    image.Update()
                    self.SetOrthoInput(
                        image, orthoView=orthoView, mviewIn=mviewIn)
            # TODO: fix this VTK-6 line here
            #  orthoView.SetInputData(image, mviewIn)

        if self.bIsMicroView:
            # set default lookup table
            self.SetLookupTable(
                image, viewportManager.GetPageState(), orthoView)

            # Add this object to the registry too; keep track of the number
            component.getGlobalSiteManager().registerUtility(
                orthoView, IOrthoView, name='OrthoView-%d' % self._numberOfImages)

            # register it's contained orthoplane collection
            component.getGlobalSiteManager().registerUtility(
                orthoView.GetOrthoPlanes(), IOrthoPlanes, name='OrthoPlanes-%d' % self._numberOfImages)

        # create an image information object - this binds all the objects created above together
        # along with common state variables

        PageState = viewportManager.GetPageState()

        PageState.type = 'image'
        if filename:
            PageState.SetFileName(filename)

        if self.bIsMicroView:
            # hook up events
            LineSegmentSelection = orthoView.lineSegment
            LineSegmentSelection.AddObserver(
                'ModifiedEvent', self.LineSegmentObserver)
            viewportManager.BindEvent(
                "<KeyPress-s>", lambda v: self.SaveCropRegionCoordinates())
            viewportManager.BindEvent(
                "<KeyPress-v>", lambda v: self.SaveCropRegion())
            viewportManager.BindEvent(
                "<KeyPress-m>", self.ShowMeanAndStdDeviation)
            viewportManager.BindEvent("<KeyPress-u>", self.SaveAreaAsImage)
            viewportManager.BindEvent("<KeyPress-y>", lambda v, n=self._numberOfImages: self.ClearLineSegment(
                None, orthoName='OrthoView-%d' % n))
            viewportManager.BindEvent(
                "<KeyPress-h>", lambda v: self.showHelp())
            viewportManager.BindEvent(
                "<KeyPress-F1>", lambda v: self.showHelp())
            viewportManager.BindEvent(
                "<Alt-KeyPress-F4>", lambda v: self.Quit(0))
            viewportManager.BindEvent(
                "<KeyPress-0>", lambda e, s=LineSegmentSelection: s.SavePointsToFile())
            viewportManager.BindEvent(
                "<KeyPress-a>", lambda e, s=LineSegmentSelection: s.SaveLineGrayScaleValues(1))
            viewportManager.BindEvent(
                "<KeyPress-A>", lambda e, s=LineSegmentSelection: s.SaveLineGrayScaleValues(0))
            viewportManager.BindEvent(
                "<Control-KeyPress-a>", self.SaveLineLengthToFile)
            viewportManager.BindEvent(
                "<KeyPress-p>", lambda e, s=self: s.PlotGrayScaleValues(e, force=True))
            viewportManager.BindEvent("<KeyPress-g>", self.PlotHistogramValues)
            viewportManager.BindEvent("<KeyPress-c>", self.ClearROI)

            # Some OS X specific keyboard bindings
            viewportManager.BindEvent("<Control-KeyPress-w>", self.onCloseTab)

            for key in ('KeyPress-3', 'KeyPress-7', 'Control-KeyPress-7', 'KeyPress-8', 'Control-KeyPress-8'):
                viewportManager.BindEvent("<%s>" % key, self.HandleROIKeyEvent)

            # limit the scope of <Delete> key events to just the 3D pane - other event handlers
            # are dynamically used elsewhere to override this type of behaviour and we don't want
            # the two to collide (Bug fix for Milton 2014-11-22)

            obj = orthoView.renderPanes[0]
            obj.BindEvent(
                '<Delete>', lambda e: self.BlankROI(e, invert=False))
            obj.BindEvent(
                '<Shift-Delete>', lambda e: self.BlankROI(e, invert=True))

            orthoView.renderPanes[0].GetRenderWindow().AddObserver(
                "AbortCheckEvent", self.CheckAbort)

            image.GetPointData().GetScalars().AddObserver(
                'ModifiedEvent', lambda e, o, s=self: s.PropagateImageLevels())

            for pane in orthoView.renderPanes:
                # observe mouse events over the pane
                pane.AddObserver('MouseMoveEvent', lambda o, e,
                                 num=self._numberOfImages: self.OnMouseMove(o, e, num))
                # attach the pane to the mode palette
                self.modePalettePane.AddPane(pane)

            # replay button event here (act as if user just pressed mode palette button)
            # this makes new images behave correctly
            self.modePalettePane.activateMode(self.modePalettePane.GetMode())

        # finally, add this tab
        if filename:
            if title is None:
                title = os.path.split(filename)[-1]
            tabname = '%d: %s' % (self._numberOfImages, title)
        else:
            tabname = '%d: Default View' % self._numberOfImages

        PageState.SetTitle(tabname)

        # create some custom buttons for the viewport manager
        if self._numberOfCurrentImagesDisplayed == 1:
            self._view_notebook.AddTabAreaButton(
                aui.AUI_BUTTON_CUSTOM1, wx.RIGHT, self._stockicons.getMenuBitmap(
                    "stock_palette"),
                "Select Palette")
            self._view_notebook.AddTabAreaButton(
                aui.AUI_BUTTON_CUSTOM2, wx.RIGHT, self._stockicons.getMenuBitmap(
                    "stock_synchronize"),
                "Synchronize Images")
            self._view_notebook.Bind(
                aui.EVT_AUINOTEBOOK_BUTTON, self.OnTabAreaButton)

        self._view_notebook.AddPage(viewportManager, tabname, select=True)

        # unfreeze viewport
        viewportManager.Thaw2()
        viewportManager.Refresh()
        viewportManager.Update()

    def get_current_viewport_manager(self, evt):
        # This next bit should be re-examined - surely
        # there's a better way to determine which widget generated
        # the event
        pages = evt.GetEventObject().GetPages()
        last_time = -1
        psel = None
        for p in pages:
            if p.active:
                psel = p
                break
            else:
                t = time.mktime(p.access_time.timetuple())
                if t > last_time:
                    last_time = t
                    psel = p

        return psel.window

    def OnTabAreaButton(self, evt):

        if evt.GetInt() == aui.AUI_BUTTON_CLOSE:
            evt.Skip()
            return

        if evt.GetInt() == aui.AUI_BUTTON_CUSTOM1:

            viewportManager = self.get_current_viewport_manager(evt)
            page_state = viewportManager.GetPageState()

            # is this an image viewer?
            if not IViewerState.providedBy(page_state):
                logging.debug(
                    "OnTabAreaButton:  page type isn't an image viewer.  Ignoring...")
                return

            dlg = LUTSelectionDialogC.LUTSelectionDialogC(self)
            dlg.SetLUTIndex(page_state.lut_index)

            if dlg.ShowModal() == wx.ID_OK:
                num = viewportManager.GetImageIndex()
                orthoView = component.getUtility(
                    IOrthoView, 'OrthoView-%d' % num)
                image = component.getUtility(IImage, 'Image-%d' % num)
                event.notify(ChangeImageLUT(
                    orthoView, image, page_state, dlg.GetLUTIndex()))

            dlg.Destroy()

        elif evt.GetInt() == aui.AUI_BUTTON_CUSTOM2:

            viewportManager = self.get_current_viewport_manager(evt)
            page_state = viewportManager.GetPageState()

            # is this an image viewer?
            if not IViewerState.providedBy(page_state):
                logging.debug(
                    "OnTabAreaButton:  page type isn't an image viewer.  Ignoring...")
                return

            dlg = SynchronizeImageDialogC.SynchronizeImageDialogC(self)

            # determine list of current images
            my_idx = viewportManager.GetImageIndex()
            values = []

            current_sync_list = viewportManager.getSyncList()

            # sanity check
            current_image_indices = [i[0] for i in self.getImageList()]
            for i in current_sync_list[:]:
                if i not in current_image_indices:
                    # image must have been deleted!
                    logging.warning(
                        "Image {0} has been deleted - correct synchronization code!".format(i))
                    current_sync_list.remove(i)

            for idx, desc in self.getImageList():
                if idx != my_idx:
                    values.append((idx, desc))
            dlg.SetCurrentSyncList(my_idx, values, current_sync_list)

            if dlg.ShowModal() == wx.ID_OK:
                new_sync_list = []
                for idx in dlg.GetCurrentSyncList():
                    new_sync_list.append(values[idx][0])

                # setup synchronization between this image and whatever the
                # user has selected
                if my_idx not in current_sync_list:
                    current_sync_list.append(my_idx)
                if my_idx not in new_sync_list:
                    new_sync_list.append(my_idx)

                event.notify(
                    SynchronizeEvent(current_sync_list, new_sync_list))

            dlg.Destroy()

    def getImageList(self):
        """Gets the full list of loaded images

        Returns:
            list: A list of tuples.  Each tuple is a (int, str) pair

        """

        # determine list of current images
        image_info = []
        n_vals = []

        for val in component.getUtilitiesFor(IViewportManager):

            viewport = val[1]
            n = viewport.GetImageIndex()

            if n not in n_vals:

                n_vals.append(n)
                page_state = viewport.GetPageState()
                image_info.append((n, page_state.GetTitle()))

        return image_info

    def getSpreadsheetList(self):
        """Gets the full list of loaded spreadsheets

        Returns:
            list: A list of tuples.  Each tuple is a (int, str) pair

        """

        # determine list of current spreadsheets
        sheet_info = []
        n_vals = []

        for val in component.getUtilitiesFor(ISpreadsheet):

            viewport = val[1]
            n = viewport.GetSpreadsheetIndex()

            if n not in n_vals:

                n_vals.append(n)
                page_state = viewport.GetPageState()
                sheet_info.append((n, page_state.GetTitle()))

        return sheet_info

    @component.adapter(ChangeImageLUT)
    def onChangeImageLUT(self, evt):
        self.SetOrthoViewLUTIndex(
            evt.orthoView, evt.image, evt.pageState, evt.lut_index)

    def SetOrthoViewLUTIndex(self, orthoView, image, pageState, lut_index):
        """Set up the lookup table for an orthoview"""

        LD = data.LutData()
        rgb = LD.get_rgb(lut_index)
        rgb2 = numpy.zeros([len(rgb[0]), 3], dtype='uint8')
        for i in range(3):
            rgb2[:, i] = numpy.fromstring(rgb[i], dtype='uint8')
        rgb = rgb2

        # replace existing lookup table to avoid excessive modification events
        table = vtk.vtkLookupTable()
        table.SetNumberOfColors(len(rgb))
        for i in range(len(rgb)):
            r, g, b = rgb[i]
            table.SetTableValue(i, r / 255.0, g / 255.0, b / 255.0, 0.0)

        orthoView.SetLUTIndex(lut_index)
        orthoView.SetLookupTable(table)
        self.WindowLevelToolbar.SetLookupTable(orthoView.GetLookupTable())

        orthoView.autoscale(image)

        pageState.lut_index = lut_index

        # notify listeners that an image lut has changed
        event.notify(ImageLUTChanged())

        orthoView.Render()

    def onCloseTab(self, evt):
        """Respond to Ctrl+W -- close notebook page"""
        page_idx = self._view_notebook.GetSelection()
        if page_idx != -1:
            self.Freeze()
            self.OnNotebookPageClose(None)
            self._view_notebook.DeletePage(page_idx)

            self.Thaw()

    def onCloseOtherTabs(self, evt):
        """Close all notebook pages except the current one"""

        # TODO: what happens when a window gets deleted?

        current_page_idx = self._view_notebook.GetSelection()
        if current_page_idx == -1:
            return
        current_window = self._view_notebook.GetPage(current_page_idx)

        # first get list of windows
        windows = []
        for page_idx in range(self._view_notebook.GetPageCount()):
            windows.append(self._view_notebook.GetPage(page_idx))

        # now iterate
        self.Freeze()
        for window in windows:
            if window != current_window:
                page_idx = self._view_notebook.GetPageIndex(window)
                self.OnNotebookPageClose(None, page_index=page_idx)
                self._view_notebook.DeletePage(page_idx)

        self.Thaw()

    def onSliceChanged(self, evt):

        # move slice position
        try:
            orthoView = component.getUtility(ICurrentOrthoView)
            op = orthoView.GetOrthoPlanes()
        except:
            # no image loaded
            return

        orthoView = component.getUtility(ICurrentOrthoView)

        slice_id = orthoView.GetTrackedSliceIndex()
        old_index = op.GetPlane(2 - slice_id).GetSliceIndex()
        new_index = evt.GetInt()

        if new_index != old_index:
            op.GetPlane(2 - slice_id).SetSliceIndex(new_index)
            component.getUtility(ICurrentViewportManager).Render()

    @component.adapter(ROIEnabledEvent)
    def onROIEnabledEvent(self, evt):
        """Respond to an ROIEnabled event

        We modify our menus and cause a redraw here"""

        self.ConfigMenuItemsForROI(True)
        component.getUtility(ICurrentOrthoPlanes).Modified()
        self.pane3D.Render()

    @component.adapter(ROIModifiedEvent)
    def onROIModifiedEvent(self, evt):

        self.ConfigMenuItemsForROI(True)
        # is this needed?
        # component.getUtility(ICurrentOrthoPlanes).Modified()
        self.pane3D.Render()

        # we act as a proxy for histogram tool here
        if evt.GetPluginName() != "histogram":
            # we might need to disable histogram highlighting here
            if self._histowindow:
                image = component.getUtility(
                    IImage, "Image-%d" % (evt.GetImageIndex() + 1))
                if image.GetStencilDataOwner() == "histogram":
                    self._histowindow.HideHighlight(evt.GetImageIndex())

    @component.adapter(ROIDisabledEvent)
    def onROIDisabledEvent(self, evt):
        self.ConfigMenuItemsForROI(False)
        component.getUtility(ICurrentOrthoPlanes).Modified()
        self.pane3D.Render()

    @component.adapter(StatusbarVoxelValue)
    def onVoxelValueTextChange(self, evt):
        self.SetStatusText(evt.GetText(), 2)

    @component.adapter(StatusbarCursorXYPosition)
    def onCursorPositionTextChange(self, evt):
        self.SetStatusText(evt.GetText(), 3)

    @component.adapter(StatusbarCursorSlicePosition)
    def onCursorSliceTextChange(self, evt):
        self.SetStatusText(evt.GetText(), 4)

    @component.adapter(WindowLevelTextChangeEvent)
    def onWindowLevelTextChange(self, evt):
        self.GetStatusBar().SetWindowLevelText(evt.GetText())

    @component.adapter(PlotWindowCursorPosition)
    def onPlotWindowCursorPositionChange(self, evt):

        x1 = evt._x
        y1 = evt._y
        label1 = ''

        if x1 is not None:
            label1 = "Pos(X,Y): %0.3f, %0.3f %s" % (
                x1, y1, self._config.GetUnitLabel())

        event.notify(StatusbarCursorXYPosition(label1))
        event.notify(StatusbarCursorSlicePosition())

    def OnNotebookPageClose(self, evt, page_index=None):

        if page_index is None:
            if evt:
                # use event itself to determine which tab to close
                page_index = evt.GetSelection()
            else:
                # we didn't get here via an event - close current page
                page_index = self._view_notebook.GetSelection()

        page = self._view_notebook.GetPage(page_index)
        state = page.GetPageState()
        page_type = state.GetPageType()
        page_title = state.GetTitle()

        # the page is closing
        state.setClosing(True)

        if page_type == 'spreadsheet':

            num = page.GetSpreadsheetIndex()

            # unregister objects from the zope global site manager
            component.getGlobalSiteManager().unregisterUtility(
                page, ISpreadsheet, name='Spreadsheet-%d' % num)

            self._numberOfCurrentSpreadsheetsDisplayed -= 1

        elif page_type == 'image':

            # before doing anything, let anyone who's interested know that the
            # current image is about to go away
            event.notify(CurrentImageClosingEvent(self.GetCurrentImageIndex(),
                                                  self.GetNumberOfImagesCurrentlyLoaded(),
                                                  page_title))

            # disconnect viewport synchronization on this window
            my_idx = page.GetImageIndex()
            current_sync_list = page.getSyncList()
            if current_sync_list:
                event.notify(SynchronizeEvent(current_sync_list, [my_idx]))

            self._numberOfCurrentImagesDisplayed -= 1
            viewportManager = page
            num = viewportManager.GetImageIndex()
            image = component.getUtility(IImage, 'Image-%d' % num)
            orthoView = component.getUtility(IOrthoView, 'OrthoView-%d' % num)
            viewerState = viewportManager.GetPageState()
            mviewIn = component.getUtility(
                IMicroViewInput, name='MicroViewInput-%d' % num)

            # unregister objects from the zope global site manager
            component.getGlobalSiteManager().unregisterUtility(
                image, IImage, name='Image-%d' % num)
            component.getGlobalSiteManager().unregisterUtility(
                orthoView, IOrthoView, name='OrthoView-%d' % num)
            component.getGlobalSiteManager().unregisterUtility(
                orthoView.GetOrthoPlanes(), IOrthoPlanes, name='OrthoPlanes-%d' % num)
            component.getGlobalSiteManager().unregisterUtility(
                viewportManager, IViewportManager, name='ViewportManager-%d' % num)
            component.getGlobalSiteManager().unregisterUtility(
                mviewIn, IMicroViewInput, name='MicroViewInput-%d' % num)

            # is this viewer the current viewer? If so, we've got more tear
            # down work
            current_viewportManager = component.getUtility(
                ICurrentViewportManager)
            if (current_viewportManager is viewportManager):
                del(current_viewportManager)
                component.getGlobalSiteManager().unregisterUtility(
                    viewportManager, ICurrentViewportManager)
                component.getGlobalSiteManager().unregisterUtility(
                    orthoView, ICurrentOrthoView)
                component.getGlobalSiteManager().unregisterUtility(
                    orthoView.GetOrthoPlanes(), ICurrentOrthoPlanes)
                component.getGlobalSiteManager().unregisterUtility(
                    image, ICurrentImage)
                component.getGlobalSiteManager().unregisterUtility(
                    mviewIn, ICurrentMicroViewInput)

                # The current image is being deleted - make sure certain
                # filters are made aware
                self.probe.SetSourceConnection(None)

            # tear down VTK pipeline
            image.DecrementReferenceCount()

            # disconnect this viewer from the toolbar control
            for pane in orthoView.renderPanes:
                self.modePalettePane.RemovePane(pane)

            viewportManager.tearDown()
            orthoView.tearDown()
            mviewIn.tearDown()

            self.WindowLevelToolbar.RemoveRenderWindow(viewportManager)

            del(orthoView)
            del(viewportManager)
            del(viewerState)
            del(image)
            del(mviewIn)

            del(self.pane3D)
            del(self.pane2d_1)
            del(self.pane2d_2)
            del(self.pane2d_3)

            if self._numberOfCurrentImagesDisplayed == 0:
                # we've close all windows - post a message
                event.notify(NoImagesLoadedEvent())
                # Reset titlebar
                self.SetWindowTitle(None, None)

            if evt:
                evt.Skip()

        # update menus
        self.UpdateMenus()

    def GetCurrentImage(self):
        """Protect ourselves a bit here"""
        try:
            image = component.getUtility(ICurrentImage)
        except:
            dlg = wx.MessageDialog(
                self, "No image selected.  Please select an image before proceeding.", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            image = None
        return image

    def GetCurrentImageIndex(self):
        return self._current_image

    def SetCurrentImageIndex(self, idx):
        self._current_image = idx

    def GetCurrentImageTitle(self):
        return self._current_title

    def SetCurrentImageTitle(self, title):
        self._current_title = title

    def GetNumberOfImagesCurrentlyLoaded(self):
        return self._numberOfCurrentImagesDisplayed

    def OnBottomNotebookPageChanged(self, evt):

        # capture page changes in the lower logging/plot/spectrum  viewer
        # notebook
        current_page = self._bottom_notebook.GetCurrentPage()

        if (current_page == self._plotwindow):
            # update the plot widget
            self.PlotGrayScaleValues(force=True)

    def OnNotebookPageChanging(self, evt):
        """
        Process event that occurs when the current notebook tab selection is about to be changed
        """

        evt.Skip()

        # capture page changes in the main viewer notebook
        page_index = self._view_notebook.GetSelection()

        # determine current objects
        page = self._view_notebook.GetPage(page_index)
        page_state = page.GetPageState()
        page_title = page_state.GetTitle()

        # Notify everyone that the current notebook tab is changing
        event.notify(NotebookPageChangingEvent(
            self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), page_title))

    def OnNotebookPageChanged(self, evt):

        evt.Skip()

        # capture page changes in the main viewer notebook
        page_index = self._view_notebook.GetSelection()

        # get current viewer state
        try:
            currentViewerState = component.getUtility(
                ICurrentViewportManager).GetPageState()
            currentOrthoView = component.getUtility(ICurrentOrthoView)
        except:
            currentViewerState = None
            currentOrthoView = None

        # determine current objects
        page = self._view_notebook.GetPage(page_index)
        page_state = page.GetPageState()
        page_type = page_state.GetPageType()
        page_title = page_state.GetTitle()

        # we can register this page as the current page regardless of what type
        # it is
        component.getGlobalSiteManager().registerUtility(
            page, ICurrentViewportManager)

        # is this an image page, or something else
        if page_type == 'spreadsheet':
            # disable scrollbars
            self.WindowLevelToolbar.EnableZSlider(False)
            self.WindowLevelToolbar.EnableLevelSlider(False)
            self.WindowLevelToolbar.EnableWindowSlider(False)
            # current page isn't an image
            try:
                image = component.getUtility(ICurrentImage)
                # unregister this image
                component.getGlobalSiteManager().unregisterUtility(
                    image, ICurrentImage)
            except:
                logging.warning("No current image to deregister")

            num = page.GetSpreadsheetIndex()

            # spreadsheets have no ROI
            self.ConfigMenuItemsForROI(False)

        elif page_type == 'image':
            num = page.GetImageIndex()
            self.SetCurrentImageIndex(num - 1)
            if self.bIsMicroView:
                orthoView = component.getUtility(
                    IOrthoView, name='OrthoView-%d' % num)
            mviewIn = component.getUtility(
                IMicroViewInput, name='MicroViewInput-%d' % num)

            # assign objectsset these objects as the current set of objects
            if self.bIsMicroView:
                component.getGlobalSiteManager().registerUtility(
                    orthoView, ICurrentOrthoView)
                component.getGlobalSiteManager().registerUtility(
                    orthoView.GetOrthoPlanes(), ICurrentOrthoPlanes)
            component.getGlobalSiteManager().registerUtility(
                mviewIn, ICurrentMicroViewInput)

            # save current window/level values
            if currentViewerState:
                currentViewerState.table_range = currentOrthoView.GetLookupTable(
                ).GetTableRange()

            # connect window/level toolbar to this viewport
            if self.bIsMicroView:
                self.WindowLevelToolbar.SetLookupTable(
                    orthoView.GetLookupTable())
            self.WindowLevelToolbar.SetRenderWindow(page)
            if self.bIsMicroView:
                self.modePalettePane.SetLookupTable(orthoView.GetLookupTable())

            if self.bIsMicroView:
                # force a redraw
                orthoView.GetLookupTable().Modified()

                # backward compatibility code
                self.pane3D = orthoView.renderPanes[0]
                self.pane2d_1 = orthoView.renderPanes[1]
                self.pane2d_2 = orthoView.renderPanes[2]
                self.pane2d_3 = orthoView.renderPanes[3]

            image = component.getUtility(IImage, name='Image-%d' % num)
            component.getGlobalSiteManager().registerUtility(
                image, ICurrentImage)

            with wx.WindowDisabler():
                with wx.BusyCursor():
                    image.Update()

            self.UpdateDimensionInfo()

            # set limit/value of slice slider
            slice_id = orthoView.GetTrackedSliceIndex()
            _range = image.GetDimensions()[slice_id]

            # TODO: refactor this
            _slice = 0
            if self.bIsMicroView:
                plane = orthoView.GetOrthoPlanes().GetPlane(2 - slice_id)
                if plane:
                    _slice = plane.GetSliceIndex()

            self.SetSliderValue(slice_id, _slice, _range)

            # VTK-6
            if vtk.vtkVersion().GetVTKMajorVersion() > 5:
                self.probe.SetSourceData(image.GetRealImage())
            else:
                self.probe.SetSource(image.GetRealImage())

            # update MicroViewIO
            # self.mviewOut.SetupWriterFilter(image)

            self.UpdateDimensionInfo()

            # Update window/level info
            self.PropagateImageLevels()

            # adjust scrollbar enable/disable
            numc = image.GetNumberOfScalarComponents()
            self.SetWindowLevelState(numc, _range)

            # Notify everyone that the current input image has changed
            event.notify(CurrentImageChangeEvent(
                self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), page_title))

        # set page title
        self.SetWindowTitle(page_state.GetFileName(), page_state.GetTitle())
        self.SetCurrentImageTitle(page_title)

        # update menus (do this here because not all platforms capture events
        # when menus are opened)
        self.UpdateMenus()

        # Notify everyone that the current notebook tab has changed
        event.notify(NotebookPageChangedEvent(
            self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), page_title))

    def SetSliderValue(self, slice_id, _slice, _range):
        label = {0: 'X-Slice', 1: 'Y-Slice', 2: 'Z-Slice'}[slice_id]
        self.WindowLevelToolbar.GetSliceLabel().SetLabel(label)
        self.WindowLevelToolbar.GetSliceSlider().SetRange(0, _range - 1)
        self.WindowLevelToolbar.GetSliceSlider().SetValue(_slice)
        self.WindowLevelToolbar.GetSliceSlider().SetPageSize(
            float(_range) / 20.0)

    def OnNotebookRightClick(self, evt):

        # display a pop-up menu
        menu = wx.Menu()

        _id = wx.NewId()
        menu.Append(_id, "&New &Tab\tCtrl+T", "Open a new tab")
        self.Bind(wx.EVT_MENU, self.OnNewTab, id=_id)
        menu.AppendSeparator()
        _id = wx.NewId()
        menu.Append(_id, "&Reload", "Reload file from disk")
        self.Bind(
            wx.EVT_MENU, lambda e, page=evt.page: self.OnReload(page), id=_id)

        # disable?
        if not evt.page.GetPageState().GetFileName():
            menu.Enable(_id, False)

        _id = wx.NewId()
        menu.Append(_id, "Re&name", "Rename image")
        self.Bind(
            wx.EVT_MENU, lambda e, page=evt.page: self.OnRename(page), id=_id)

        # disable?
        if not IViewerState.providedBy(evt.page.GetPageState()):
            menu.Enable(_id, False)

        _id = wx.NewId()
        menu.Append(_id, "&Duplicate", "Clone image into a new window")
        self.Bind(
            wx.EVT_MENU, lambda e, page=evt.page: self.OnDuplicateImage(page), id=_id)

        # disable?
        if not IViewerState.providedBy(evt.page.GetPageState()):
            menu.Enable(_id, False)

        menu.AppendSeparator()
        _id = wx.NewId()
        menu.Append(_id, "&Close\tCtrl+W", "Close the current tab")
        self.Bind(wx.EVT_MENU, self.onCloseTab, id=_id)
        _id = wx.NewId()
        menu.Append(_id, "&Close other tabs", "Close all other tabs")
        self.Bind(wx.EVT_MENU, self.onCloseOtherTabs, id=_id)

        self._view_notebook.PopupMenu(menu)
        menu.Destroy()

    def setModePaletteToolState(self, image):
        self.modePalettePane.EnableRotate(image.GetDimensions()[-1] > 1)
        self.modePalettePane.EnableSlice(image.GetDimensions()[-1] > 1)
        self.modePalettePane.EnableWindowLevel(
            image.GetNumberOfScalarComponents() < 3)

    def SetWindowLevelState(self, numc, zrange):

        # enable/disable this scrollbar component
        self.WindowLevelToolbar.EnableZSlider(zrange > 1)

        # Does W/L adjustment make sense for this image?
        self.WindowLevelToolbar.EnableLevelSlider(numc == 1)
        self.WindowLevelToolbar.EnableWindowSlider(numc == 1)

    def OnMouseMove(self, obj, evt, num):

        # TODO: For now, we don't handle image transforms
        # TODO: handle scaling

        image = component.getUtility(IImage, 'Image-%d' % num)

        numC = image.GetNumberOfScalarComponents()
        _type = image.GetScalarType()
        label = {1: 'Gray Scale Value', 2:
                 'Values', 3: 'RGB', 4: 'RGBA'}[numC]

        # Update cursor position - here are the default labels
        xypos = "Pos(X,Y): "

        line1 = '%s: ' % label

        # get a global object that can manages reading images from disk
        dimension = image.GetDimensionInformation()[-1]
        name = dimension.GetTypeName()

        # Build up label in two stages... - this is stage 1
        if name == 'Distance':
            zpos_label = 'Pos(Z): '
        elif name == 'Wavelength':
            zpos_label = 'Wavelength: '
        elif name == 'Time (ms)':
            zpos_label = 'Time-point: '
        else:
            zpos_label = 'Z-Pos: '

        if obj.position:

            v = []
            x, y, z = obj.position
            dims = image.GetDimensions()
            x0, y0, z0 = image.GetOrigin()
            xspacing, yspacing, zspacing = image.GetSpacing()

            if name == 'Distance':
                # this is a standard z-slice
                if self._config.bMeasurementUnitIsMM:
                    zpos_label += '%0.3f %s' % (
                        obj.position[2], self._config.GetUnitLabel())
                else:
                    zpos_label += '%0.3f pixels' % (
                        (obj.position[2] - z0) / zspacing)
            elif name == 'Wavelength':
                slice_number = int(round(component.getUtility(
                    ICurrentOrthoView).GetOrthoPlanes().GetAxialPlane().GetSliceIndex()))
                z_slice_positions = image.GetZSlicePositions()
                if slice_number <= len(z_slice_positions):
                    zpos_label += '%0.3f nm' % z_slice_positions[slice_number]
            elif name == 'Time (ms)':
                zpos_label += '%0.2f ms' % obj.position[2]

            # optionally apply transform to world coordinate point
            t = component.getUtility(ICurrentOrthoView).GetImageTransform()
            if t is not None:
                x, y, z = t.GetInverse().TransformPoint(x, y, z)

            # convert from world coordinate to pixel
            xpos = (x - x0) / xspacing + 0.5  # 2014-03-05 - this is required
            ypos = (y - y0) / yspacing + 0.5
            zpos = (z - z0) / zspacing + 0.5

            if dims[2] == 1:
                zpos = 0

            if xpos >= 0 and xpos < dims[0] and ypos >= 0 and ypos < dims[1] and zpos >= 0 and zpos < dims[2]:
                for i in range(numC):
                    v.append(image.GetScalarComponentAsDouble(
                        int(xpos), int(ypos), int(zpos), i))
                if _type in (vtk.VTK_FLOAT, vtk.VTK_DOUBLE):
                    line1 = line1 + (numC - 1) * '%1.3f, ' + '%1.3f\n'
                else:
                    line1 = line1 + (numC - 1) * '%1.0f, ' + '%1.0f\n'
                line1 = line1 % tuple(v)
            else:
                v = [None for i in range(numC)]

            if self._config.bMeasurementUnitIsMM:
                xypos = "Pos(X,Y): %0.3f, %0.3f %s" % (
                    obj.position[0], obj.position[1], image.GetMeasurementUnit())
            else:
                xypos = "Pos(X,Y): %0.1f, %0.1f pixels" % (
                    (obj.position[0] - x0) / xspacing, (obj.position[1] - y0) / yspacing)

        # fire off three events
        event.notify(StatusbarVoxelValue(line1))
        event.notify(StatusbarCursorXYPosition(xypos))
        event.notify(StatusbarCursorSlicePosition(zpos_label))

    def OnDropFiles(self, x, y, filenames):

        # determine total amount user has requested
        total = 0
        try:
            for _file in filenames:
                total += os.path.getsize(_file)
        except:
            logging.error("Unable to process file '%s'" % _file)

        # if the user has requested more than 15 images, assume this is not
        # what they want to do
        if len(filenames) > 15:
            dlg = wx.MessageDialog(
                None, u"%d files will be opened into separate tabs - are you sure this is what you want to do?\nNote that File->Image Import... can be used to import a stack of 2D slices" % len(
                    filenames), 'Are you sure?',
                wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION)
            bAreYouSure = (dlg.ShowModal() == wx.ID_YES)
            dlg.Destroy()
            if not bAreYouSure:
                return False

        # if the user has requested a lot of memory (2 GB), assume this is not
        # what they want to do
        if (total > 2 * 1024 * 1024 * 1024):
            total_in_mb = (total / 1024. / 1024.)
            dlg = wx.MessageDialog(
                None, "This action will require %0.2f MiB of memory - are you sure this is what you want to do?" % total_in_mb, 'Are you sure?',
                wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION)
            bAreYouSure = (dlg.ShowModal() == wx.ID_YES)
            dlg.Destroy()
            if not bAreYouSure:
                return False

        reactor.callLater(0.1, self.OnDropFiles2, filenames)

        return True

    def OnDropFiles2(self, filenames):

        # give any providers a chance at this file drop
        for name, obj in component.getUtilitiesFor(IFileDropHandler):
            if obj.OnDropFilesEvent(filenames):
                # this plugin accepted files
                return

        # nobody has accepted the file drop - open images in new tabs
        for _file in filenames:
            wx.SafeYield(onlyIfNeeded=True)
            self.OnFileOpen(_file, create_tab=True)

    def createStatisticsWindow(self):

        if self._statswindow is None:
            self._statswindow = StatisticsGUIC.StatisticsGUIC(self)
            self._bottom_notebook.AddPage(
                self._statswindow, "Image Statistics")
            self._statswindow.m_gridStatistics.AutoSize()

    def createHistogramWindow(self):

        if self._histowindow is None:
            self._histowindow = VTKPlotWindow.VTKPlotWindow(self,
                                                            title='Histogram of Voxel Values',
                                                            xlabel='Voxel Value',
                                                            ylabel='Frequency',
                                                            bar=1, center=1, subdivisions=10)
            self._bottom_notebook.AddPage(self._histowindow, 'Histogram')

        return self._histowindow

    def createPlotWindow(self):

        if self._plotwindow is None:
            self._plotwindow = VTKPlotWindow.VTKPlotWindow(
                self, title='Line Profile', ylabel='Grayscale Value', center=1, subdivisions=10, scale=1.0,
                units=MicroViewSettings.MicroViewSettings.getObject().GetUnitLabel())
            self._bottom_notebook.AddPage(self._plotwindow, 'Line Profile')

            # VTK-6
            if vtk.vtkVersion().GetVTKMajorVersion() > 5:
                self._plotwindow.SetInputConnection(self.probe.GetOutputPort())
            else:
                self._plotwindow.SetInput(self.probe.GetOutput())

        return self._plotwindow

    def createSpectrumWindow(self):
        if not self._spectrumwindow:
            self._spectrumwindow = SpectrumPlotWindow.SpectrumPlotWindow(
                self, title='Spectrum', xlabel='Wavelength (nm)', ylabel='Grayscale Value', center=1,
                subdivisions=10, scale=1.0, units='nm')
            self._spectrumwindow.SetFileName(self.GetFileName())
        self._bottom_notebook.AddPage(self._spectrumwindow, 'Spectrum')

        return self._spectrumwindow

    def CreateMenuBar(self):

        menubar = wx.MenuBar()
        menubar.SetName("MenuBar1")

        #----------------------------------------------------------------------
        # File Menu
        #----------------------------------------------------------------------
        menu1 = wx.Menu()

        self._save_file_ids = []
        self._transform_ids = []

        _ids = [wx.NewId() for i in range(18)]

        _id_new_tab = wx.NewId()
        item = wx.MenuItem(
            menu1, _id_new_tab, "&New &Tab\tCtrl+T", "Open a new tab")
        menu1.AppendItem(item)

        item = wx.MenuItem(menu1, _ids[
                           0], "&Open...\tCtrl+O", "Open a file into current tab")
        item.SetBitmap(
            self._stockicons.getMenuBitmap("glyphicons_144_folder_open"))
        menu1.AppendItem(item)

        _ids_open_new_tab = wx.NewId()
        item = wx.MenuItem(
            menu1, _ids_open_new_tab, "&Open (New Tab)...\tCtrl+Shift+O", "Open a file into a new tab")
        menu1.AppendItem(item)

        menu1.AppendSeparator()
        menu1.Append(_ids[1], 'Read Transform...', "Read and apply a transform")
        item = menu1.Append(_ids[2], 'Reset Transform...', "Restores default transform")
        item.Enable(False)
        item = menu1.Append(_ids[3], 'Save Transformed Image...', 'Save transformed image to file')
        item.Enable(False)
        self._save_file_ids.append(_ids[3])

        # keep track of transform-related IDs
        for _id in (_ids[1], _ids[2], _ids[3]):
            self._transform_ids.append(_ids[1])

        menu1.AppendSeparator()

        self._id_close_tab = wx.NewId()
        item = wx.MenuItem(
            menu1, self._id_close_tab, "&Close\tCtrl+W", "Close the current tab")
        menu1.AppendItem(item)

        menu1.AppendSeparator()
        item = wx.MenuItem(menu1, _ids[
                           4], 'Save As...', 'Save current data to a file')
        item.SetBitmap(self._stockicons.getMenuBitmap(wx.ART_FILE_SAVE_AS))
        menu1.AppendItem(item)
        self._save_file_ids.append(_ids[4])
        menu1.Append(_ids[
                     5], 'Save Reoriented Image...', 'Save a reformatted copy of the image')
        item = menu1.Append(_ids[
                            6], 'Save Crop region...', 'Save crop-region to file')
        item.Enable(False)
        self._save_file_ids.append(_ids[5])
        self._save_file_ids.append(_ids[6])
        item = menu1.Append(_ids[
                            7], 'Save Crop coordinates', 'Save crop-region coords')
        item.Enable(False)
        self._save_file_ids.append(_ids[7])
        item = wx.MenuItem(menu1, _ids[
                           9], 'Save Snapshot...', 'Save snapshot of window')
        item.SetBitmap(self._stockicons.getMenuBitmap('camera'))
        menu1.AppendItem(item)
        self._save_file_ids.append(_ids[9])

        # item = wx.MenuItem(menu1, _ids[
        #                   10], 'Save Scene...', 'Save 3D scene to a file')
        # menu1.AppendItem(item)
        # self._save_file_ids.append(_ids[10])

        # menu1.AppendSeparator()
        ##menu1.Append(_ids[11], 'Save Session...', 'Save session')
        ##menu1.Append(_ids[12], 'Load Session...', 'Load session')

        menu1.AppendSeparator()
        menu1.Append(_ids[
                     13], 'Load Crop Coordinates...', 'Load crop coordinates')

        # handle mac slightly differently
        if sys.platform != 'darwin':
            menu1.AppendSeparator()
            key_binding = 'Ctrl+P'
        else:
            key_binding = 'Ctrl+,'

        item = menu1.Append(
            wx.NewId(), "Customize shortcuts...", 'Customize shortcuts')
        self.Bind(wx.EVT_MENU, self.OnKeyboardShortcuts, item)

        #item = menu1.Append(wx.ID_PREFERENCES, "&Preferences...\t%s" %
        #                    key_binding, 'Edit Preferences')
        #item.SetBitmap(self._stockicons.getMenuBitmap('cogs'))
        #self.Bind(wx.EVT_MENU, self.OnPreferences, item)

        menu1.AppendSeparator()
        item = wx.MenuItem(menu1, _ids[
                           14], 'Import Image...', 'Import a stack of 2D images')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_import'))
        menu1.AppendItem(item)
        item = wx.MenuItem(menu1, _ids[
                           15], 'Export Image...', 'Export to a stack of 2D images')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_export'))
        menu1.AppendItem(item)
        self._save_file_ids.append(_ids[15])

        item = wx.MenuItem(
            menu1, _ids[16], 'Export DICOMDIR...', 'Export image to DICOMDIR folder')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_export'))
        menu1.AppendItem(item)
        self._save_file_ids.append(_ids[16])

        item = wx.MenuItem(
            menu1, _ids[17], 'Regenerate DICOMDIR index...', 'Generate a DICOMDIR index')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_export'))
        menu1.AppendItem(item)
        # self._save_file_ids.append(_ids[17])

        menu1.AppendSeparator()
        item = menu1.Append(
            wx.ID_EXIT, 'Q&uit MicroView\tAlt+F4', 'Quit MicroView')

        # File history
        self._history = MicroViewHistory.MicroViewHistory.getObject()
        self._history.UseMenu(menu1)
        self._history.load()

        # Events
        self.Bind(wx.EVT_MENU_RANGE, self.OnFileHistory,
                  id=wx.ID_FILE1, id2=wx.ID_FILE5)
        self.Bind(wx.EVT_MENU, lambda e: self.OnFileOpen(), id=_ids[0])
        self.Bind(wx.EVT_MENU, self.OnFileOpenInNewTab, id=_ids_open_new_tab)
        self.Bind(wx.EVT_MENU, self.OnNewTab, id=_id_new_tab)
        self.Bind(wx.EVT_MENU, self.ReadTransformFile, id=_ids[1])
        self.Bind(
            wx.EVT_MENU, lambda e: self.DeleteImageTransform(), id=_ids[2])
        self.Bind(wx.EVT_MENU, lambda e: self.SaveTransformedImageAs(), id=_ids[3])
        self.Bind(wx.EVT_MENU, lambda e: self.SaveAs(), id=_ids[4])
        self.Bind(wx.EVT_MENU, self.SaveReorientedImageAs, id=_ids[5])
        self.Bind(wx.EVT_MENU, lambda e: self.SaveCropRegion(), id=_ids[6])
        self.Bind(wx.EVT_MENU, lambda e:
                  self.SaveCropRegionCoordinates(), id=_ids[7])
        self.Bind(wx.EVT_MENU, self.showTKeyHint, id=_ids[9])
        self.Bind(wx.EVT_MENU, self.saveScene, id=_ids[10])
        #self.Bind(wx.EVT_MENU, self.SerializeApp, id=_ids[11])
        #self.Bind(wx.EVT_MENU, self.DeSerializeApp, id=_ids[12])
        self.Bind(wx.EVT_MENU, self.OnLoadCropCoords, id=_ids[13])
        self.Bind(wx.EVT_MENU, self.OnFileImport, id=_ids[14])
        self.Bind(wx.EVT_MENU, self.OnFileExport, id=_ids[15])
        self.Bind(wx.EVT_MENU, self.OnDICOMDIRExport, id=_ids[16])
        self.Bind(wx.EVT_MENU, self.OnDICOMDIRRegenerate, id=_ids[17])
        self.Bind(wx.EVT_MENU, self.onCloseTab, id=self._id_close_tab)
        self.Bind(wx.EVT_MENU, self.Quit, id=wx.ID_EXIT)

        #----------------------------------------------------------------------
        # Edit
        #----------------------------------------------------------------------
        menu2 = wx.Menu()

        _ids = [wx.NewId() for i in range(6)]

        item = menu2.Append(_ids[
                            0], 'Show Line', 'Display measurement tool')

        menu2.AppendSeparator()
        item = menu2.Append(_ids[
                            1], 'Clear Line\tCtrl+Y', 'Remove line profiling tool')
        item = menu2.Append(_ids[
                            2], 'Clear ROI\tCtrl+C', 'Remove ROI from screen')

        menu2.AppendSeparator()
        item = menu2.Append(_ids[
                            3], 'Edit Background Properties\tCtrl+L', 'Edit background properties')
        item = menu2.Append(_ids[
                            4], 'Edit Window/Level Properties', 'Edit Window/Level properties')

        menu2.AppendSeparator()
        item = wx.MenuItem(menu2, _ids[
                           5], '&Reset Display\tCtrl+R', 'Resets all views')
        item.SetBitmap(self._stockicons.getMenuBitmap(wx.ART_REDO))
        menu2.AppendItem(item)

        # Events
        self.Bind(wx.EVT_MENU, self.ShowLineSegment, id=_ids[0])
        self.Bind(wx.EVT_MENU, self.ClearLineSegment, id=_ids[1])
        self.Bind(wx.EVT_MENU, self.ClearROI, id=_ids[2])
        self.Bind(wx.EVT_MENU, lambda evt: component.getUtility(
            ICurrentOrthoView).onKeyPressL(True), id=_ids[3])
        self.Bind(wx.EVT_MENU, self.onShowWindowLevelDialog, id=_ids[4])
        self.Bind(wx.EVT_MENU, self.Reset, id=_ids[5])

        #----------------------------------------------------------------------
        # Tools
        #----------------------------------------------------------------------
        menu3 = wx.Menu()

        _ids = [wx.NewId() for i in range(4)]

        item = menu3.Append(
            _ids[0], 'Plot...\tCtrl+P', 'Plot points along a line')
        item.Enable(False)

        item = wx.MenuItem(
            menu3, _ids[1], 'Histogram...\tCtrl+G', 'Plot Voxel Histogram')
        item.SetBitmap(
            self._stockicons.getMenuBitmap('select_histogram_region'))
        menu3.AppendItem(item)

        item.Enable(False)

        # Events
        self.Bind(wx.EVT_MENU, lambda e: self.PlotGrayScaleValues(
            force=True), id=_ids[0])
        self.Bind(wx.EVT_MENU, self.PlotHistogramValues, id=_ids[1])

        #----------------------------------------------------------------------
        # Process
        #----------------------------------------------------------------------
        menu4 = wx.Menu()

        _ids = [wx.NewId() for i in range(5)]

        self.downsample_menu = wx.Menu()
        menu4.AppendMenu(wx.NewId(), 'Downsample Menu', self.downsample_menu)

        item = self.downsample_menu.Append(
            _ids[0], 'Downsample to 8-bit unsigned Int', 'Downsample to an 8-bit Image')
        item.Enable(False)
        item = self.downsample_menu.Append(
            _ids[1], 'Downsample to 16-bit', 'Downsample to a 16-bit Image')
        item.Enable(False)
        item = self.downsample_menu.Append(_ids[2], 'Convert to Grayscale',
                                           'Convert a multichannel or colour image to Grayscale')
        item.Enable(False)

        item = wx.MenuItem(menu4, _ids[
                           3], 'Image Invert', 'Generate the negative version of the Image')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_invert'))
        menu4.AppendItem(item)

        item = wx.MenuItem(menu4, _ids[
                           4], 'Image Reorient', 'Reformat image along current axes')
        menu4.AppendItem(item)

        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.DownsampleImage(8), id=_ids[0])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.DownsampleImage(16), id=_ids[1])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.ConvertToGrayScale(), id=_ids[2])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.ImageInvert(), id=_ids[3])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.CreateNewImageFromReorientedImage(), id=_ids[4])

        # Filter Menu
        self.filter_menu = wx.Menu()
        menu4.AppendMenu(wx.NewId(), 'Filter', self.filter_menu)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Gaussian Smooth...', 'Apply a Gaussian smoothing filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s._gaussSmooth(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Minimum Filter...', 'Apply a minimum filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s.minimumFilter(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Maximum Filter...', 'Apply a maximum filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s.maximumFilter(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Median Filter...', 'Apply a median filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s._medianFilter(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Uniform Filter...', 'Apply a uniform smoothing filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s.uniformFilter(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Anisotropic Smooth...', 'Apply an edge-preserving smoothing filter')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s._anisotropicFilter(), id=_new_id)

        self.filter_menu.AppendSeparator()

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Gradient Magnitude Filter', 'Compute the image gradient magnitude')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s._gradientMagnitudeFilter(), id=_new_id)

        _new_id = wx.NewId()
        item = self.filter_menu.Append(
            _new_id, 'Laplacian Filter', 'Compute the image Laplacian')
        self.Bind(
            wx.EVT_MENU, lambda evt, s=self: s._laplacianFilter(), id=_new_id)

        # Threshold Menu
        self.threshold_menu = wx.Menu()
        menu4.AppendMenu(wx.NewId(), 'Threshold', self.threshold_menu)

        _new_id = wx.NewId()
        self.threshold_menu.Append(_new_id, 'Manual Threshold...',
                                   'Binarize an image based on a user-defined threshold')
        self.Bind(wx.EVT_MENU, self.OnThresholdImage, id=_new_id)

        _new_id = wx.NewId()
        self.threshold_menu.Append(
            _new_id, 'Adaptive Threshold...', 'Binarize an image based on an adaptive threshold')
        self.Bind(wx.EVT_MENU, self.OnAdaptiveThresholdImage, id=_new_id)

        _new_id = wx.NewId()
        self.threshold_menu.Append(
            _new_id, 'Otsu Threshold...', 'Binarize an image based on Otsu threshold')
        self.Bind(wx.EVT_MENU, self.OnOtsuThresholdImage, id=_new_id)

        _new_id = wx.NewId()
        self.threshold_menu.Append(_new_id, 'Adaptive Otsu Threshold...',
                                   'Binarize an image based on adaptive Otsu thresholding')
        self.Bind(wx.EVT_MENU, self.OnAdaptiveOtsuThresholdImage, id=_new_id)

        # Image Flip
        _ids = [wx.NewId() for i in range(3)]
        menu4_2 = wx.Menu()
        menu4.AppendMenu(wx.NewId(), 'Image Flip',  menu4_2)

        item = wx.MenuItem(menu4_2, _ids[
                           0], 'x-axis', 'Flip image about x-axis')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_arrow_x'))
        menu4_2.AppendItem(item)

        item = wx.MenuItem(menu4_2, _ids[
                           1], 'y-axis', 'Flip image about y-axis')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_arrow_y'))
        menu4_2.AppendItem(item)

        item = wx.MenuItem(menu4_2, _ids[
                           2], 'z-axis', 'Flip image about z-axis')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_arrow_z'))
        menu4_2.AppendItem(item)

        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.ImageFlip(0), id=_ids[0])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.ImageFlip(1), id=_ids[1])
        self.Bind(wx.EVT_MENU, lambda evt,
                  s=self: s.ImageFlip(2), id=_ids[2])

        # Binary/Morphological filter Menu
        self.binary_menu = menu4_3 = wx.Menu()
        menu4.AppendMenu(wx.NewId(), 'Morphological Filters', menu4_3)

        _new_id = wx.NewId()
        self.binary_menu.Append(
            _new_id, 'Label', 'Label regions in a black-and-white image')
        self.Bind(wx.EVT_MENU, self.OnLabel, id=_new_id)

        _new_id = wx.NewId()
        self.binary_menu.Append(
            _new_id, 'Region Properties', 'Calculate region properties from a labelled image')
        self.Bind(wx.EVT_MENU, self.OnRegionProps, id=_new_id)

        menu4_3.AppendSeparator()

        _new_id = wx.NewId()
        self.binary_menu.Append(_new_id, 'Erode', 'Erode the image')
        self.Bind(wx.EVT_MENU, self.OnErodeMorphology, id=_new_id)

        _new_id = wx.NewId()
        self.binary_menu.Append(_new_id, 'Dilate', 'Dilate the image')
        self.Bind(wx.EVT_MENU, self.OnDilateMorphology, id=_new_id)

        _new_id = wx.NewId()
        self.binary_menu.Append(
            _new_id, 'Open/Close...', 'Open or close a binary or greylevel image')
        self.Bind(wx.EVT_MENU, self.onOpenCloseMorphology, id=_new_id)

        menu4_3.AppendSeparator()

        _new_id = wx.NewId()
        self.binary_menu.Append(
            _new_id, 'Distance Map', 'Compute Euclidean distance map')
        self.Bind(wx.EVT_MENU, self.onComputeDistanceMap, id=_new_id)

        menu4.AppendSeparator()

        _ids = [wx.NewId() for i in range(3)]

        menu4_4 = wx.Menu()
        # TODO: figure out how to wire up missing icon --
        # self._stockicons.getMenuIcon('scissors9')
        menu4.AppendMenu(wx.NewId(), 'ROI Blanking', menu4_4)

        item = wx.MenuItem(menu4_4, _ids[
                           0], 'Blank Inside', 'Set voxels within the active ROI to a user-defined value')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_delete_inside'))
        menu4_4.AppendItem(item)

        item = wx.MenuItem(menu4_4, _ids[
                           1], 'Blank Outside', 'Set voxels outside of the active ROI to a user-defined value')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_delete_outside'))
        menu4_4.AppendItem(item)
        item = menu4_4.Append(_ids[
                              2], 'Set Background Value...', 'Set the ROI blanking user-defined value')

        menu4.AppendSeparator()

        # Events
        self.Bind(wx.EVT_MENU, lambda evt, s=self: s.BlankROI(
            evt, invert=False), id=_ids[0])
        self.Bind(wx.EVT_MENU, lambda evt, s=self: s.BlankROI(
            evt, invert=True), id=_ids[1])
        self.Bind(wx.EVT_MENU, lambda evt, s=self:
                  s.OnSetBlankROIBackgroundValue(), id=_ids[2])

        # Visualize
        menu5 = wx.Menu()

        # Analyze
        menu6 = wx.Menu()

        # Plugins
        menu7 = wx.Menu()

        # Window
        menu8 = wx.Menu()

        _ids = [wx.NewId() for i in range(13)]

        self.show_left_panel = menu8.Append(_ids[0], 'Show Applications && Tools Panel', 'Show Applications && Tools Panel',
                                            kind=wx.ITEM_CHECK)
        self.show_left_panel.Check()

        self.show_mode_palette_panel = menu8.Append(_ids[1], 'Show Mouse Mode Palette', 'Show MouseMode Palette',
                                                    kind=wx.ITEM_CHECK)
        self.show_mode_palette_panel.Check()

        self.show_windowlevel_panel = menu8.Append(_ids[2], 'Show Window/Level Widget', 'Show Window/Level Widget',
                                                   kind=wx.ITEM_CHECK)
        self.show_windowlevel_panel.Check()

        self.show_scales = menu8.Append(_ids[3], 'Show Scales', 'Show Scales along side of 2D viewports',
                                        kind=wx.ITEM_CHECK)

        self.show_colour_bar = menu8.Append(_ids[4], 'Show Colorbar', 'Show Colorbar along side of 3D viewports',
                                            kind=wx.ITEM_CHECK)
        self.show_windowlevel_panel.Check(False)

        self.Bind(wx.EVT_MENU, self.menuTogglePanelSize, id=_ids[0])
        self.Bind(wx.EVT_MENU, self.menuToggleShowModePalette, id=_ids[1])
        self.Bind(wx.EVT_MENU, self.menuToggleShowWindowLevelPanel, id=_ids[2])
        self.Bind(wx.EVT_MENU, self.menuToggleShowScales, id=_ids[3])
        self.Bind(wx.EVT_MENU, self.menuToggleShowColourBar, id=_ids[4])

        #####################################################################
        menu8.AppendSeparator()

        self.show_results_menu = wx.Menu()
        menu8.AppendMenu(wx.NewId(), 'Bottom Panel', self.show_results_menu)

        _ids = [wx.NewId() for i in range(5)]

        self.hide_bottom_panel = self.show_results_menu.Append(_ids[
                                                               0], 'Hide all Results', 'Hide the bottom window',
                                                               kind=wx.ITEM_RADIO)
        self.show_log_panel = self.show_results_menu.Append(_ids[
                                                            1], 'Show Log Window', 'Show logging  window',
                                                            kind=wx.ITEM_RADIO)
        self.show_results_panel = self.show_results_menu.Append(_ids[
                                                                2], 'Show Results Window', 'Show results window',
                                                                kind=wx.ITEM_RADIO)
        self.show_plot_panel = self.show_results_menu.Append(_ids[
                                                             3], 'Show Plot Window', 'Show plot Window',
                                                             kind=wx.ITEM_RADIO)
        self.show_histogram_panel = self.show_results_menu.Append(_ids[
            4], 'Show Histogram Window', 'Show histogram Window',
            kind=wx.ITEM_RADIO)

        self.Bind(wx.EVT_MENU, self.menuHideBottomWindow, id=_ids[0])
        self.Bind(wx.EVT_MENU, self.menuShowLog, id=_ids[1])
        self.Bind(wx.EVT_MENU, self.menuShowResults, id=_ids[2])
        self.Bind(wx.EVT_MENU, self.menuShowPlot, id=_ids[3])
        self.Bind(wx.EVT_MENU, self.menuShowHistogram, id=_ids[4])

        #####################################################################
        menu8.AppendSeparator()

        self.layout_menu = wx.Menu()
        menu8.AppendMenu(wx.NewId(), 'Layout', self.layout_menu)

        _ids = [wx.NewId() for i in range(6)]
        item = self.layout_menu.Append(_ids[
                                       0], 'Set Layout to 1-by-3', 'Set Layout to 1-by-3', wx.ITEM_RADIO)
        item = self.layout_menu.Append(_ids[
                                       1], 'Set Layout to 2-by-2', 'Set Layout to 2-by-2', wx.ITEM_RADIO)
        item = self.layout_menu.Append(_ids[
                                       2], 'Set Layout to the 3D View', 'Set Layout to the 3D View', wx.ITEM_RADIO)
        self.layout_menu_z_only = self.layout_menu.Append(_ids[
                                                          3], 'Set Layout to the Z-Slice View',
                                                          'Set Layout to the Z-Slice View', wx.ITEM_RADIO)
        item = self.layout_menu.Append(_ids[
                                       4], 'Set Layout to the Y-Slice View', 'Set Layout to the Y-Slice View',
                                       wx.ITEM_RADIO)
        item = self.layout_menu.Append(_ids[
                                       5], 'Set Layout to the X-Slice View', 'Set Layout to the X-Slice View',
                                       wx.ITEM_RADIO)

        # keep track of layout window menu items
        self.__layout_menu_items = [_ids[
            0], _ids[1], _ids[2], _ids[4], _ids[5]]

        self._layoutToId = {'All': _ids[0],
                            'All_2x2': _ids[1],
                            'pane3D': _ids[2],
                            'pane2d_1': _ids[3],
                            'pane2d_2': _ids[4],
                            'pane2d_3': _ids[5]}

        self.Bind(wx.EVT_MENU, lambda e, mode='All': self.SetLayout(
            mode), id=_ids[0])
        self.Bind(wx.EVT_MENU, lambda e, mode='All_2x2': self.SetLayout(
            mode), id=_ids[1])
        self.Bind(wx.EVT_MENU, lambda e, mode='pane3D': self.SetLayout(
            mode), id=_ids[2])
        self.Bind(wx.EVT_MENU, lambda e, mode='pane2d_1': self.SetLayout(
            mode), id=_ids[3])
        self.Bind(wx.EVT_MENU, lambda e, mode='pane2d_2': self.SetLayout(
            mode), id=_ids[4])
        self.Bind(wx.EVT_MENU, lambda e, mode='pane2d_3': self.SetLayout(
            mode), id=_ids[5])

        menu8.AppendSeparator()
        _id = wx.NewId()
        item = menu8.Append(
            _id, 'Tile Vertically', 'Tile all Windows Vertically')
        self.Bind(wx.EVT_MENU, self.onTileLayoutVertical, id=_id)

        _id = wx.NewId()
        item = menu8.Append(
            _id, 'Tile Horizontally', 'Tile all Windows Horizontally')
        self.Bind(wx.EVT_MENU, self.onTileLayoutHorizontal, id=_id)

        _id = wx.NewId()
        item = menu8.Append(_id, 'Reset Layout', 'Collect all images')
        self.Bind(wx.EVT_MENU, self.onResetTabbedLayout, id=_id)

        # we must update certain menu entries when this menu is opened
        self.Bind(wx.EVT_MENU_OPEN, self.UpdateMenus)

        # Help
        menu9 = wx.Menu()

        _ids = [wx.NewId() for i in range(5)]

        item = wx.MenuItem(menu9, _ids[
                           0], 'MicroView Help...', 'Get online Help')
        item.SetBitmap(self._stockicons.getMenuBitmap("stock_help"))

        menu9.AppendItem(item)
        item = menu9.Append(wx.ID_ABOUT, 'About...', 'Display About Box')
        item = wx.MenuItem(menu9, _ids[
                           2], 'Network Info...', 'Display network information')
        item.SetBitmap(self._stockicons.getMenuBitmap('stock_info'))
        menu9.AppendItem(item)
        item = menu9.Append(
            _ids[4], 'Open Plugin directory', 'Examine user plugins')

        self.Bind(wx.EVT_MENU, lambda e: self.showHelp(), id=_ids[0])
        self.Bind(wx.EVT_MENU, self.showAbout, id=wx.ID_ABOUT)
        self.Bind(wx.EVT_MENU, self.showMacAddresses, id=_ids[2])
        self.Bind(wx.EVT_MENU, self.openPluginDirectory, id=_ids[4])

        menubar.Append(menu1, '&File')
        menubar.Append(menu2, "Edit")
        menubar.Append(menu3, "&Tools")
        menubar.Append(menu4, "&Process")
        menubar.Append(menu5, "&Visualize")
        menubar.Append(menu6, "&Analyze")
        menubar.Append(menu7, "&Plugins")
        menubar.Append(menu8, "&Window")
        menubar.Append(menu9, "&Help")

        self.SetMenuBar(menubar)

        return menubar

    def DisplayWebPage(self, _url):
        """Display a popup dialog - load a url into it"""

        try:
            dlg = UpstreamMessageDialogGUI.UpstreamMessageDialogGUI(self)
            dlg.m_wxHtmlWindow.LoadURL(_url)
        except:
            # platform probably doesn't support webkit - spawn a browser
            # instead
            webbrowser.open(_url)
            return

        if sys.platform == 'darwin':
            # see http://trac.wxwidgets.org/ticket/13863
            dlg.Show()
        else:
            try:
                dlg.ShowModal()
            finally:
                dlg.Destroy()

    def onOpenCloseMorphology(self, evt):
        """Perform binary/greylevel open or close morphological operation"""
        res = None

        if self._OpenCloseMorphologyDialog is None:
            self._OpenCloseMorphologyDialog = OpenCloseMorphologyDialogC.OpenCloseMorphologyDialogC(
                self)

        try:
            if self._OpenCloseMorphologyDialog.ShowModal() == wx.ID_OK:
                res = self._OpenCloseMorphologyDialog.GetResults()
        finally:
            self._OpenCloseMorphologyDialog.Close()

        if res is None:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.OpenCloseImageFilter(
                component.getUtility(ICurrentImage), res)
            self.SetInput(image)
            event.notify(WindowLevelPresetEvent('Auto-adjust (99%)'))

    def OnFileHistory(self, evt):

        shift_state = wx.GetKeyState(wx.WXK_SHIFT)
        control_state = wx.GetKeyState(wx.WXK_CONTROL)

        if shift_state and control_state:
            create_tab = True
        else:
            create_tab = False

        # get the file based on the menu ID
        fileNum = evt.GetId() - wx.ID_FILE1
        path = self._history.GetHistoryFile(fileNum)

        if os.path.exists(path):
            self.OnFileOpen(path, create_tab=create_tab)
            # add it back to the history so it will be moved up the list
            self._history.AddFileToHistory(path)
        else:
            dlg = wx.MessageDialog(
                self, "File not found: '%s'" % path, 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            self._history.RemoveFileFromHistory(fileNum)

    def disable_all_menus(self):
        """Disable all menu entries except critical ones like exit, about and close"""

        return

        menubar = self.GetTopLevelParent().GetMenuBar()
        for menu_id in range(menubar.GetMenuCount()):
            menu = menubar.GetMenu(menu_id)
            for item in menu.GetMenuItems():
                if item.GetId() not in (wx.ID_EXIT, wx.ID_ABOUT, self._id_close_tab):
                    item.Enable(False)

    def UpdateMenus(self, evt=None):

        menu = None
        title = ''
        process_all = False
        page_type = None
        rank = 0

        # get page info +++

        page_index = self._view_notebook.GetSelection()

        try:
            orthoView = component.getUtility(ICurrentOrthoView)
        except interface.interfaces.ComponentLookupError:
            orthoView = None

        try:
            image = component.getUtility(ICurrentImage)
        except interface.interfaces.ComponentLookupError:
            image = None

        if page_index == -1:
            bCanSave = False
        else:
            # page exists -- but can it be saved?
            page = self._view_notebook.GetPage(page_index)
            page_type = page.GetPageState().GetPageType()
            closing = page.GetPageState().getClosing()
            if page_type == 'image':
                bCanSave = True and not closing
            elif page_type == 'spreadsheet':
                bCanSave = False and not closing  # for now, disable saving
            else:
                # we shouldn't get here
                bCanSave = False

        # get page info ---

        if evt:
            try:
                menu = evt.GetMenu()
                title = menu.GetTitle()
            except:
                # catch system menu click
                menu = None
                title = ''
        else:
            process_all = True

        if process_all or title == '&Process':

            if evt is None:
                menu = self.GetTopLevelParent().GetMenuBar().GetMenu(3)

            # non-images don't belong here
            self.GetTopLevelParent().GetMenuBar().EnableTop(
                3, page_type == 'image')

            enabled = True

            # disable image-based menus if no image is loaded
            if self._numberOfCurrentImagesDisplayed == 0:
                enabled = False

            # disable image-based menus if this page doesn't hold an image
            if page_type != 'image':
                enabled = False

            # enable/disable some menu entries
#            menu.Enable(menu.FindItem('Downsample Menu'), enabled)
#            menu.Enable(menu.FindItem('Image Invert'), enabled)
#            menu.Enable(menu.FindItem('Image Reorient'), enabled)
#            menu.Enable(menu.FindItem('Filter'), enabled)
#            menu.Enable(menu.FindItem('Threshold'), enabled)
#            menu.Enable(menu.FindItem('Image Flip'), enabled)
#            menu.Enable(menu.FindItem('ROI Blanking'), enabled)
#            menu.Enable(menu.FindItem('Morphological Filters'), enabled)

#            menu.Enable(menu.FindItem('Gaussian Smooth...'), enabled)
#            menu.Enable(menu.FindItem('Minimum Filter...'), enabled)
#            menu.Enable(menu.FindItem('Maximum Filter...'), enabled)
#            menu.Enable(menu.FindItem('Median Filter...'), enabled)
#            menu.Enable(menu.FindItem('Uniform Filter...'), enabled)
#            menu.Enable(menu.FindItem('Anisotropic Smooth...'), enabled)
#            menu.Enable(menu.FindItem('Gradient Magnitude Filter'), enabled)
#            menu.Enable(menu.FindItem('Laplacian Filter...'), enabled)

            # update downsample options based on number of bits in image
            bits = 0
            nc = 0
            if enabled and image:
                bits = image.GetScalarSize() * 8
                nc = image.GetNumberOfScalarComponents()

            self.downsample_menu.Enable(self.downsample_menu.FindItem(
                'Downsample to 8-bit unsigned Int'), (bits != 8 and enabled))
            self.downsample_menu.Enable(self.downsample_menu.FindItem(
                'Downsample to 16-bit'), (bits > 16 and enabled))
            self.downsample_menu.Enable(self.downsample_menu.FindItem('Convert to Grayscale'), (
                nc > 1 and enabled))

        if process_all or title == '&File':

            if evt is None:
                menu = self.GetTopLevelParent().GetMenuBar().GetMenu(0)

            # adjust entries for transform manipulation
            transform_applied = False
            if orthoView:
                transform_applied = (orthoView.GetImageTransform() is not None)
            menu.Enable(self._transform_ids[1], transform_applied)
            menu.Enable(self._transform_ids[2], transform_applied)

            # disable/enable Save as... type buttons if the current notebook page either
            # doesn't exist, or can't be saved

            for _id in self._save_file_ids:
                menu.Enable(_id, bCanSave)

            # adjust transform-related entries
            if page_type == 'image' and orthoView:
                t = orthoView.GetImageTransform()
            else:
                t = None

            menu.Enable(menu.FindItem('Reset Transform...'), t is not None)
            menu.Enable(menu.FindItem('Read Transform...'),
                        bCanSave)
            menu.Enable(menu.FindItem('Save Transformed Image...'),
                        t is not None and bCanSave )
            menu.Enable(menu.FindItem('Load Crop Coordinates...'),
                        bCanSave)

        if process_all or title == 'Edit':

            if evt is None:
                menu = self.GetTopLevelParent().GetMenuBar().GetMenu(1)

            if image:
                enabled = True
            else:
                enabled = False

            # enable/disable some menu entries
            menu.Enable(menu.FindItem('Show Line'), enabled)
            menu.Enable(menu.FindItem('Clear Line'), enabled)
            menu.Enable(menu.FindItem('Clear ROI'), enabled)
            menu.Enable(menu.FindItem('Reset Display'), enabled)
            menu.Enable(menu.FindItem('Edit Background Properties'), enabled)
            menu.Enable(menu.FindItem('Edit Window/Level Properties'), enabled)

        if process_all or title == '&Tools':
            if evt is None:
                menu = self.GetTopLevelParent().GetMenuBar().GetMenu(2)

            item = self.menubar.FindItemById(
                self.menubar.FindMenuItem('Tools', 'Plot...'))
            enabled = False
            if self._numberOfCurrentImagesDisplayed > 0:
                if orthoView and orthoView.lineSegment.GetStatus():
                    enabled = True
            item.Enable(enabled)

        if process_all or title == '&Window':

            if evt is None:
                menu = self.GetTopLevelParent().GetMenuBar().GetMenu(7)

            # update checkboxes
            panel = self._mgr.GetPane(self.leftframe)
            self.show_left_panel.Check(panel.IsShown())
            panel = self._mgr.GetPane(self.modePalettePane)
            self.show_mode_palette_panel.Check(panel.IsShown())
            panel = self._mgr.GetPane(self.WindowLevelToolbar)
            self.show_windowlevel_panel.Check(panel.IsShown())

            plot_shown = False
            if self._plotwindow:
                if self._mgr.GetPane(self._plotwindow).IsShown() is True:
                    index = self._bottom_notebook.GetPageIndex(
                        self._plotwindow)
                    index2 = self._bottom_notebook.GetSelection()
                    plot_shown = (index == index2)

            self.show_plot_panel.Check(plot_shown)

            scales_shown = False
            if page_type == 'image' and orthoView:
                scales_shown = orthoView.GetScalebarVisibility()
                self.show_scales.Enable(True)
            else:
                self.show_scales.Enable(False)

            self.show_colour_bar.Check(scales_shown)

            colourbar_shown = False
            if page_type == 'image' and orthoView:
                colourbar_shown = orthoView.GetColourbarVisibility()
                self.show_colour_bar.Enable(True)
            else:
                self.show_colour_bar.Enable(False)

            self.show_colour_bar.Check(colourbar_shown)

        if process_all or menu == self.show_results_menu:
            # Update menu buttons
            panel = self._mgr.GetPane(self._bottom_notebook)

            # is bottom window shown?
            if not panel.IsShown():
                self.hide_bottom_panel.Check(True)
            else:
                idx = self._bottom_notebook.GetSelection()
                if (idx == self._bottom_notebook.GetPageIndex(self.logPane)):
                    self.show_log_panel.Check(True)
                elif (idx == self._bottom_notebook.GetPageIndex(self.resultsPane)):
                    self.show_results_panel.Check(True)
                elif (idx == self._bottom_notebook.GetPageIndex(self._plotwindow)):
                    self.show_plot_panel.Check(True)

        if process_all or menu == self.layout_menu:

            bImageIsDisplayed = True

            # is there a current viewport?
            try:
                viewport = component.getUtility(ICurrentViewportManager)
            except interface.interfaces.ComponentLookupError:
                viewport = None
                bImageIsDisplayed = False

            # is the viewport an image?
            if viewport and not IViewportManager.providedBy(viewport):
                bImageIsDisplayed = False

            # if we're not dealing with an image, disable/clear everything
            if not bImageIsDisplayed:
                # there is no current viewport
                for _id in self.__layout_menu_items:
                    menu.Enable(_id, False)
                    menu.Check(_id, False)
                menu.Enable(self.layout_menu_z_only.GetId(), False)
                menu.Check(self.layout_menu_z_only.GetId(), False)
                return

            # disable window layouts that don't make sense if a single 2D image is loaded
            # also, disable window layouts that don't make sense for image stacks that
            # don't have a z dimension that is a distance
            try:
                image = component.getUtility(ICurrentImage)
                dim_info = image.GetDimensionInformation()
                rank = image.GetRank()
                bzAxisIsSpatial = isinstance(dim_info[2], MVImage.Distance)
            except:
                bzAxisIsSpatial = False

            for _id in self.__layout_menu_items:
                menu.Enable(_id, (rank > 2) and bzAxisIsSpatial)

            if (not bzAxisIsSpatial) or (rank == 2):
                self.layout_menu_z_only.Check(True)
                menu.Enable(self.layout_menu_z_only.GetId(), True)

            # determine menu id that should be current
            _id = self._layoutToId[viewport.GetCurrentViewMode()]
            menu.Check(_id, True)

    def onShowWindowLevelDialog(self, evt):
        # invoke a Zope event to tell app to display a dialog
        event.notify(ShowWindowLevelDialogEvent())

    @component.adapter(ShowWindowLevelDialogEvent)
    def onShowWindowLevelDialogEvent(self, evt):

        try:
            orthoview = component.getUtility(ICurrentOrthoView)
            orthoview.WindowLevelPopup()
        except component.ComponentLookupError:
            pass

    @component.adapter(ViewportMouseDown)
    def onViewportMouseDown(self, evt):

        if not self._view_notebook.GetPage(self._view_notebook.GetSelection()) is evt.GetViewport():
            # mouse click came from a different viewport - switch focus to it
            self._view_notebook.SetSelectionToWindow(evt.GetViewport())

    @component.adapter(TrackedSliceIndexChangeEvent)
    def onTrackedSliceIndexChangeEvent(self, evt):
        """change label """
        slice_id = evt.GetTrackedSliceIndex()
        image = component.getUtility(ICurrentImage)
        orthoView = component.getUtility(ICurrentOrthoView)
        _range = image.GetDimensions()[slice_id]
        plane = orthoView.GetOrthoPlanes().GetPlane(2 - slice_id)
        if plane:  # on Centos 6.5 we seem to get plane=None sometimes.  Why?
            _slice = plane.GetSliceIndex()  # Note the index is flipped here
            self.SetSliderValue(slice_id, _slice, _range)

    def SerializeApp(self, filename=None):
        """
        SerializeApp - save the state of MicroView to disk.

        This routine creates a xml pickleable class, packs it with objects derived from other important
        classes in MicroView.  Each class in turn must define 'GetObjectState()', which returns a pickleable
        object.
        """

        curr = self.GetCurrentDirectory()
        if not filename:
            filename = EVSFileDialog.asksaveasfilename(
                message='Save Session', filetypes=collections.OrderedDict([("MicroView session file", ["*.xml"])]),
                defaultdir=curr, defaultextension='*.xml')

        if filename is None or filename == '':
            return

        # construct a class
        class MicroViewData(object):

            def __init__(self):
                pass
        obj = MicroViewData()

        obj.orthoView = component.getUtility(
            ICurrentOrthoView).GetObjectState()

        fp = file(filename, 'wt')
        # fp.write(xml_pickle.dumps(obj))
        fp.close()

    def CheckAbort(self, obj, evt):
        if obj.GetEventPending() != 0:
            obj.SetAbortRender(1)

    # Plot a graph showing the gray scale values along the line selected
    def PlotGrayScaleValues(self, e=None, force=False):

        # bail out immediately if an image transform has been loaded
        t = component.getUtility(ICurrentOrthoView).GetImageTransform()

        # Plot a graph only if a line has been selected
        if component.getUtility(ICurrentOrthoView).lineSegment.GetStatus() == 1:
            (x1, y1, z1) = component.getUtility(
                ICurrentOrthoView).lineSegment.GetFirstPoint()
            (x2, y2, z2) = component.getUtility(
                ICurrentOrthoView).lineSegment.GetSecondPoint()

            # apply transform to points
            if t is not None:
                t2 = t.GetInverse()
                x1, y1, z1 = t2.TransformPoint(x1, y1, z1)
                x2, y2, z2 = t2.TransformPoint(x2, y2, z2)

            # ensure round-off doesn't occur for 2D images
            if self.rank == 2:
                z1 = z2 = component.getUtility(ICurrentImage).GetOrigin()[2]

            # calculate the length of the line in pixels
            spacing = component.getUtility(ICurrentImage).GetSpacing()
            LineLength_pixels = math.sqrt(math.pow((x1 - x2) / spacing[0], 2) +
                                          math.pow((y1 - y2) / spacing[1], 2) +
                                          math.pow((z1 - z2) / spacing[2], 2))
            # Calculate the length of the line in mm
            LineLength_mm = math.sqrt(math.pow(x1 - x2, 2) +
                                      math.pow(y1 - y2, 2) +
                                      math.pow(z1 - z2, 2))

            # VTK-6
            if vtk.vtkVersion().GetVTKMajorVersion() > 5:
                self.probe.SetSourceData(
                    component.getUtility(ICurrentImage).GetRealImage())
            else:
                self.probe.SetSource(
                    component.getUtility(ICurrentImage).GetRealImage())

            self.LineSegment.SetPoint1(x1, y1, z1)
            self.LineSegment.SetPoint2(x2, y2, z2)
            if LineLength_pixels < 100:
                NumSegments = 100
            else:
                NumSegments = int(math.ceil(LineLength_pixels))
            self.LineSegment.SetResolution(NumSegments)
            self.LineSegment.Update()

            if self._plotwindow is None:
                force = True
                self.createPlotWindow()

            self._plotwindow.SetUnitScalings({'mm': 1.0, 'pixels': float(
                LineLength_pixels) / LineLength_mm})
            if force:
                self.ShowPlotWindow()

            if self.IsPlotWindowVisible():
                self._plotwindow.SetFileName(self.GetFileName())
                self._plotwindow.PlotData()

        return 0

    @component.adapter(MotionMonitoringDisableRequestedEvent)
    def onMotionMonitoringDisableRequestedEvent(self, evt):
        self.DisableIntensityAnnotation()

    @component.adapter(MotionMonitoringEnableRequestedEvent)
    def onMotionMonitoringEnableRequestedEvent(self, evt):
        self.EnableIntensityAnnotation()

    @component.adapter(BinSizeModifiedEvent)
    def onHistogramBinSizeChangeEvent(self, evt):

        bs = evt.GetBinSize()
        self._ROIStats.SetBinSize(bs)
        x, y = self._ROIStats.GetHistogram()
        self._histowindow.SetHistogramData(x, y)
        self._histowindow.PlotData()
        self._histowindow.SetFileName(self.GetFileName())

    @component.adapter(HistogramSelectionModifiedEvent)
    def onHistogramSelectionModified(self, evt):
        self.UpdateHistogramSelectionHighlight()

    @component.adapter(HistogramIsActiveROIEvent)
    def onHistogramIsActiveROIEvent(self, evt):

        # let everyone know that the active ROI for this image has changed - we act as a proxy for the
        # histogram tool here.  This should be done _first_ before setting our own ROI visual to make
        # sure other ROI tools don't post ROIChange events
        image_index = self.GetCurrentImageIndex()
        self.DisableROIPlugin(image_index)

        # change highlight from red to yellow
        table = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetLookupTable("histogram")
        table.SetTableValue(1, 1.0, 1.0, 0.0, 1.0)
        table.Build()
        component.getUtility(ICurrentViewportManager).Render()

        # we've got the ball
        event.notify(ROIEnabledEvent("histogram", image_index))

        # now convert image to a stencil and pass it in as ROI for image
        stencil_data = self.HistogramROIToStencil()
        image = component.getUtility(ICurrentImage)
        image.SetStencilData(stencil_data, owner="histogram")

    @component.adapter(HistogramClosedEvent)
    def onHistogramClosedEvent(self, evt):

        image_index = self.GetCurrentImageIndex()
        logging.warning("TODO: implement the rest of onHistogramClosedEvent")

    @component.adapter(AutoThresholdCommandEvent)
    def onAutoThresholdCommandEvent(self, evt):
        """Respond to a request to calculate an image ROI Otsu threshold"""
        with wx.BusyCursor():
            table = component.getUtility(ICurrentOrthoView).GetLookupTable()
            _min, _max = table.GetTableRange()
            l = self._ROIStats.GetThreshold()
            w = (_max - _min)
            _min = l - w / 2.0
            _max = l + w / 2.0

            table.SetTableRange(_min, _max)
            component.getUtility(ICurrentViewportManager).Render()

        if self._histowindow:
            self._histowindow.showThresholdMarker(l)

    @component.adapter(GetROIStencilEvent)
    def onGetROIStencilEvent(self, evt):
        self.HistogramROIToStencil()

    @component.adapter(OrthoPlanesRemoveInputEvent)
    def onOrthoPlanesRemoveInputEvent(self, evt):
        self.OrthoPlanesRemoveInput(evt)

    @component.adapter(EnableMotionMonitoring)
    def onEnableMotionMonitoring(self, evt):
        logging.info('TODO: Implement onEnableMotionMonitoring!!')

    @component.adapter(DisableMotionMonitoring)
    def onDisableMotionMonitoring(self, evt):
        logging.info('TODO: Implement onEnableMotionMonitoring!!')

    @component.adapter(XSliceValueChangeEvent)
    def onXSliceValueChanged(self, evt):
        return self.onSliceValueChanged(evt)

    @component.adapter(YSliceValueChangeEvent)
    def onYSliceValueChanged(self, evt):
        return self.onSliceValueChanged(evt)

    @component.adapter(ZSliceValueChangeEvent)
    def onZSliceValueChanged(self, evt):
        return self.onSliceValueChanged(evt)

    def onSliceValueChanged(self, evt):

        try:
            component.getUtility(ICurrentImage)
        except:
            # no image loaded so exit early
            return

        # get current z-slice index
        idx = int(round(evt.GetSliceValue()))
        old_idx = self.WindowLevelToolbar.GetSliceSlider().GetValue()

        # update our own GUI
        if idx != old_idx:
            self.WindowLevelToolbar.GetSliceSlider().SetValue(idx)

    def HistogramROIToStencil(self):
        logging.debug('generating ROI stencil from histogram in MicroViewMain')
        stencil_generator_filter = vtk.vtkImageToImageStencil()
        stencil_generator_filter.ThresholdByUpper(1)
        grow = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetInput("histogram")
        stencil_generator_filter.SetInput(grow)
        # TODO: VTK-6 is this needed?
        # self._histowindow.SetROIStencilData(s.GetOutput())
        return stencil_generator_filter.GetOutput()

    def GetClippedImage(self):

        image = component.getUtility(ICurrentImage)

        origin = image.GetOrigin()
        extent = self.GetROIExtent()
        clip = None
        if extent:
            stencil = vtk.vtkImageStencil()
            stencil.SetInput(image.GetRealImage())
            stencil.SetBackgroundValue(image.GetScalarRange()[0])
            stencil.SetStencil(self.GetROIStencilData())
            clip = vtk.vtkImageClip()
            clip.SetClipData(1)
            clip.GetOutput().SetOrigin(origin)
            clip.SetOutputWholeExtent(extent)
            clip.SetInput(stencil.GetOutput())
            clip.GetOutput().UpdateInformation()
        return clip

    def UpdateHistogramSelectionHighlight(self):

        # create histogram window if it doesn't already exist
        self.createHistogramWindow()

        highlightvar = self._histowindow.GetHighlightVisible()
        image = component.getUtility(ICurrentImage)
        histogram_input = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetInput("histogram")

        selectedrange = self._histowindow.GetSelectionRange()

        # no selected region?  abort early.
        if (selectedrange == (None, None)):
            return

        if (highlightvar is False):
            # set the look up table to be the invisible plane
            if histogram_input:
                component.getUtility(ICurrentOrthoView).GetOrthoPlanes().SetLookupTable(
                    self._histowindow.GetwlTableInvisible(), "histogram")
            component.getUtility(ICurrentViewportManager).Render()

            # highlighting disabled??  abort early.
            if highlightvar:
                self._histowindow.SetHighlightVisible(False)
            return

        thresholdfilter = vtk.vtkImageThreshold()
        thresholdfilter.ReplaceInOn()
        thresholdfilter.ReplaceOutOn()
        thresholdfilter.SetInValue(1)
        thresholdfilter.SetOutValue(0)
        thresholdfilter.ThresholdBetween(selectedrange[0], selectedrange[1])
        thresholdfilter.SetOutputScalarType(3)
        clip = self.GetClippedImage()
        if clip is None:
            thresholdfilter.SetInput(image.GetRealImage())
        else:
            thresholdfilter.SetInput(clip.GetOutput())
        thresholdfilter.Update()

        # existing histogram overlay?  delete it.
        if histogram_input:
            component.getUtility(
                ICurrentOrthoView).GetOrthoPlanes().RemoveInput("histogram")

        component.getUtility(ICurrentOrthoView).GetOrthoPlanes().SetInputData(
            thresholdfilter.GetOutput(), "histogram")

        component.getUtility(ICurrentOrthoView).GetOrthoPlanes().SetLookupTable(
            self._histowindow.CreateHighlightLookupTable(), "histogram")

        component.getUtility(ICurrentViewportManager).Render()

    def OrthoPlanesRemoveInput(self, name):

        input = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetInput(name)
        if input:
            component.getUtility(
                ICurrentOrthoView).GetOrthoPlanes().RemoveInput(name)

        if name == 'histogram':
            if self._histowindow:
                input_name = self._histowindow.GetInputName()
                if input_name:
                    # remove input
                    self._histowindow.SetInputName("")
                    self._histowindow.SetHighlightVisible(False)

    def onTileLayoutVertical(self, evt):
        self.TileLayout(1)

    def onTileLayoutHorizontal(self, evt):
        self.TileLayout(2)

    def TileLayout(self, orientation):

        nb = self._view_notebook

        count = nb.GetPageCount()

        # We need complete control over the layout
        manager = nb.GetAuiManager()

        # Taken from UnFreeze()
        self.Freeze()

        # remember the tab now selected
        nowSelected = nb.GetPage(nb.GetSelection())

        windows = []
        titles = []
        bitmaps = []
        infos = []
        new_tabs = []

        # iterate over all pages and dismantle the current arrangement
        for idx in xrange(count):
            # get win reference
            windows.append(nb.GetPage(idx))
            # get page info
            infos.append(nb.GetPageInfo(idx))
            # get tab title
            titles.append(nb.GetPageText(idx))
            # get page bitmap
            bitmaps.append(nb.GetPageBitmap(idx))

            # remove page from existing tab
            tab, idx = nb.FindTab(windows[idx])
            infos[-1].active = False
            tab.RemovePage(infos[-1].window)

        # now reconstruct the way we want it
        split_size = nb.GetClientSize()

        if orientation == 1:
            split_size.x /= count
        elif orientation == 2:
            split_size.y /= count

        x = 0
        y = 0

        # this contrivance makes sure the images are sorted in increasing order
        indices = collections.deque(range(count))
        indices.rotate(1)

        for idx in indices:
            # create a new tab frame
            new_tab = TabFrame(nb)
            new_tab._rect = wx.RectPS(wx.Point(x, y), split_size)

            # last tab eats up remaining space
            if idx == (count - 1):
                if orientation == 1:
                    new_tab._rect[2] = nb.GetClientSize()[0] - new_tab._rect[0]
                elif orientation == 2:
                    new_tab._rect[3] = nb.GetClientSize()[1] - new_tab._rect[1]

            if orientation == 1:
                x += split_size.x
            elif orientation == 2:
                y += split_size.y

            new_tab.SetTabCtrlHeight(nb._tab_ctrl_height)
            nb._tab_id_counter += 1
            new_tab._tabs = AuiTabCtrl(nb, nb._tab_id_counter)
            dest_tabs = new_tab._tabs

            new_tab._tabs.SetArtProvider(nb._tabs.GetArtProvider().Clone())
            new_tab._tabs.SetAGWFlags(nb._agwFlags)

            page_info = infos[idx]
            if page_info.control:
                nb.ReparentControl(page_info.control, dest_tabs)

            cloned_buttons = nb.CloneTabAreaButtons()
            for clone in cloned_buttons:
                dest_tabs.AddButton(
                    clone.id, clone.location, clone.bitmap, clone.dis_bitmap)
            # create a pane info structure with the information
            # about where the pane should be added
            pane_info = framemanager.AuiPaneInfo().Bottom().CaptionVisible(
                False)

            if orientation == 1:
                pane_info.Left()
                mouse_pt = wx.Point(0, split_size.y / 2)
            else:
                pane_info.Top()
                mouse_pt = wx.Point(split_size.x / 2, 0)

            manager.AddPane(new_tab, pane_info, mouse_pt)

            # add the page to the destination tabs
            dest_tabs.InsertPage(page_info.window, page_info, 0)

            new_tabs.append(dest_tabs)

        # remove empty tab frames
        nb.RemoveEmptyTabFrames()

        nb.DoSizing()

        for tab in new_tabs:
            tab.DoShowHide()
            tab.Refresh()

        # force the set selection function reset the selection
        nb._curpage = -1

        # reset/redraw each viewport
        for viewport in windows:
            nb.SetSelectionToWindow(viewport)
            viewport.ResetView()

        # set the active page to the one we just split off
        nb.SetSelectionToWindow(nowSelected)

        nb.UpdateHintWindowSize()

        self.Thaw()

    def onResetTabbedLayout(self, evt):

        self._view_notebook.UnSplit()

    def SetLayout(self, mode, image_num=None):
        """Set Layout to mode: All or All_2x2.

        mode: The view layout mode.
        image_num (optional): viewport number to adjust - defaults to current image
        """

        if image_num is None:
            viewport_manager = component.getUtility(ICurrentViewportManager)
        else:
            viewport_manager = component.getUtility(
                IViewportManager, name='ViewportManager-%d' % image_num)

        viewerState = viewport_manager.GetPageState()
        viewerState.layout = mode
        viewerState.DisplayAllViews = True
        self.DisplayView(image_num=image_num)
        viewport_manager.Render()

    def GetCurrentDirectory(self):
        # determine working directory
        curr_dir = os.getcwd()
        # over-ride with system-wide directory
        try:
            curr_dir = self._config.GlobalCurrentDirectory or curr_dir
        except:
            self._config.GlobalCurrentDirectory = curr_dir

        return curr_dir

    def SaveCurrentDirectory(self, curr_dir):
        self._config.GlobalCurrentDirectory = curr_dir

    def GetOrthoCenter(self):
        "Get the origin of the image."
        return component.getUtility(ICurrentOrthoView).GetOrthoPlanes().GetOrthoCenter()

    def RegShowPointVol2(self, x, y, z):
        self.RegLandmarks.UpdateSphere(
            self.RegLandmarks.GetSphereToEdit(), x, y, z)
        return 0

    def RegSetOrthoCenter(self, position):
        """Set the ortho center of the subordinate"""
        self.pane3D.GetOrthoPlanes().SetOrthoCenter(position)
        self.pane3D.GetRenderWindow().Render()

    def RegClosePluginDialog(self):
        self.Registration.CloseDialog()
        return 0

    def GetMaskValue(self):
        """Get the image mask value used for deleting ROIs"""
        if self._maskValue is None:
            self._maskValue = component.getUtility(
                ICurrentImage).GetScalarRange()[0]

        return self._maskValue

    def SetMaskValue(self, val):
        """Set the image mask value used for deleting ROIs"""
        self._maskValue = val
        return 0

    def OnPreferences(self, evt):

        dlg = component.getUtility(IApplicationPreferencesManager)

        dlg.RecreateTree()
        if dlg.ShowModal() == wx.ID_OK:
            value = dlg.GetValue()

    def OnKeyboardShortcuts(self, evt):

        dlg = SE.ShortcutEditor(self)
        dlg.FromMenuBar(self)

        # there's a bug in the shortcut editor - work around it
        # normally we'd just call dlg.ShowModal() but that returns None at the
        # moment

        dlg.PreShow()
        if wx.Dialog.ShowModal(dlg) == wx.ID_OK:

            # save settings here
            manager = dlg.GetShortcutManager()

            # Changes accepted, send back the new shortcuts to the TLW
            # wx.MenuBar
            dlg.ToMenuBar(self)

        dlg.Destroy()

    def OnSetBlankROIBackgroundValue(self):
        """ Callback function for Set Background Value.
        """

        dlg = wx.TextEntryDialog(
            self, 'Background Value:',
            'Background Value for ROI Blanking', 'Blanking')

        dlg.SetValue(str(self.GetMaskValue()))
        try:
            if dlg.ShowModal() == wx.ID_OK:
                value = dlg.GetValue()
            else:
                return
        finally:
            dlg.Destroy()

        try:
            value = float(value)
            self._maskValue = value
        except ValueError:
            logging.warning("Invalid blanking value: %s" % value)

    def onComputeDistanceMap(self, evt):
        """Callback function to compute distance map"""

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():

            image = obj.ComputeDistanceMap(component.getUtility(ICurrentImage))
            self.SetInput(image)
            event.notify(WindowLevelPresetEvent('Auto-adjust (99%)'))

    def OnLabel(self, evt):
        """Label a binarized image."""

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                _n = obj.LabelImageInPlace(image)
                self.UpdateCurrentImage()
            dlg = wx.MessageDialog(self, "{0} discrete image regions found.".format(_n), 'Information',
                                   wx.OK | wx.ICON_INFORMATION)
            dlg.ShowModal()
            dlg.Destroy()

    def OnRegionProps(self, evt):
        """Calculate label region properties"""

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = component.getUtility(ICurrentImage)
            props = obj.CalculateRegionProps(image)

    def OnErodeMorphology(self, evt):
        self.ApplyErodeOrDilate('Erosion', 'BinaryErodeFilter')

    def OnDilateMorphology(self, evt):
        self.ApplyErodeOrDilate('Dilate', 'BinaryDilateFilter')

    def ApplyErodeOrDilate(self, name, op):

        # display a dialog to get user input
        dlg = ErodeDilateMorphologyDialogC.ErodeDilateMorphologyDialogC(self)
        try:
            if dlg.ShowModal() == wx.ID_OK:
                niter = dlg.GetNumberIterations()
            else:
                logging.info("{0} operation cancelled".format(name))
                return
        finally:
            dlg.Destroy()

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                getattr(obj, op)(image, niter)
                self.UpdateCurrentImage()

    def OnThresholdImage(self, evt):
        """ Callback function to binarize an image."""

        _min, _max = component.getUtility(
            ICurrentOrthoView).GetLookupTable().GetTableRange()
        threshold_value = (_min + _max) / 2.0

        if self._threshold_dlg is None:
            self._threshold_dlg = wx.TextEntryDialog(
                self, 'Threshold Value:',
                'Threshold Value for Binarizing Image', '%0.3f' % threshold_value)

        try:
            if self._threshold_dlg.ShowModal() == wx.ID_OK:
                value = self._threshold_dlg.GetValue()
            else:
                return
        finally:
            self._threshold_dlg.Close()

        try:
            threshold_value = float(value)
        except ValueError:
            logging.warning("Invalid threshold value: %s" % threshold_value)
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.ImageThreshold(
                component.getUtility(ICurrentImage), threshold_value)
            self.SetInput(image)
            # window/level values need to be changed
            currentViewerState = component.getUtility(
                ICurrentViewportManager).GetPageState()
            currentViewerState.table_range = 0, 1
            self.PropagateImageLevels()
            self.Refresh()

    def OnAdaptiveThresholdImage(self, evt):

        radius, image3d = self.askForRadius('Adaptive Threshold', 3)

        if radius == -1:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.AdaptiveThresholdFilter(image, radius)
                self.UpdateCurrentImage()

    def OnOtsuThresholdImage(self, evt):

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.OtsuThresholdFilter(image)
                self.UpdateCurrentImage()

    def OnAdaptiveOtsuThresholdImage(self, evt):

        min_level = 1000.0
        chunk_size = 64

        # display a dialog to get user input
        dlg = AdaptiveOtsuThresholdDialogC.AdaptiveOtsuThresholdDialogC(self)
        try:
            if dlg.ShowModal() == wx.ID_OK:
                results = dlg.GetResults()
                if results is None:
                    logging.error("Invalid value")
                    return
                else:
                    min_level, chunk_size = results
            else:
                logging.info("Adaptive threshold operation cancelled")
                return
        finally:
            dlg.Destroy()

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.AdaptiveOtsuThresholdFilter(
                    image, min_level=min_level, chunk_size=chunk_size)
                self.UpdateCurrentImage()

    def ImageFlip(self, axis):
        """Flips image about a given axis, without doubling memory"""
        logging.info("Flipping image along %s-axis..." %
                     {0: 'x', 1: 'y', 2: 'z'}[axis])

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.ImageFlip(component.getUtility(ICurrentImage), axis)
            self.SetInput(image)

    def BlankROI(self, e, invert=False):
        """ Blanks image within currently-selected ROI.
        Set invert=True to blank everything outside ROI.
        """

        # sanity check - is current mask value appropriate?
        mask_value = self.GetMaskValue()
        if (mask_value > component.getUtility(ICurrentImage).GetScalarTypeMax()) or \
                (mask_value < component.getUtility(ICurrentImage).GetScalarTypeMin()):
            dlg = wx.MessageDialog(self, "Current background value of " + str(mask_value) +
                                   " outside of valid range.\n" +
                                   "Please re-enter background value and try again.", 'Warning', wx.OK | wx.ICON_WARNING)
            dlg.ShowModal()
            dlg.Destroy()
            return

        if hasattr(e, 'pane'):
            # we arrived here via a vtkAtamai keyboard event
            num = e.pane.GetImageIndex() - 1
        else:
            # we arrived here via a wx menu event
            num = self.GetCurrentImageIndex()

        with wx.BusyCursor():
            stencil_data = self.GetROIStencilData(num)
            if stencil_data is None:
                return -1

            # VTK-6
            if vtk.vtkVersion().GetVTKMajorVersion() > 5:
                logging.debug(
                    "TODO: figure out how to update stencil data here")
            else:
                stencil_data.Update()

            g = _MicroView.vtkInPlaceImageStencil()

            g.SetMaskValue(mask_value)
            if invert:
                g.ReverseStencilOn()

            # VTK-6
            if vtk.vtkVersion().GetVTKMajorVersion() > 5:
                g.SetInputData(
                    component.getUtility(IImage, name='Image-%d' % (num + 1)).GetRealImage())
            else:
                g.SetInput(
                    component.getUtility(IImage, name='Image-%d' % (num + 1)).GetRealImage())

            g.SetStencil(stencil_data)
            g.SetProgressText("Blanking image...")
            g.AddObserver('ProgressEvent', self.HandleVTKProgressEvent)
            g.Update()
            # g.GetOutput().Update()  # VTK-6

            # since we're doing an IN PLACE operation, we don't need to change
            # the input, so just leave this alone - seems to be causing all sorts
            # of problems with RegionGrow generated ROIs
            # self.SetInput(g.GetOutput())
            # self.ImageData.SetSource(None)
            # the next two lines are a bit kludgy, but are needed to get some other
            # functionality to work correctly
            stencil_data.Modified()

            # component.getUtility(ICurrentOrthoView).GetOrthoPlanes().SetOrthoCenter(component.getUtility(ICurrentOrthoView).GetOrthoPlanes().GetOrthoCenter())
            # this is a HACK fix to get the planes to update and show the result
            # of the blanking - not sure how else to get this to work right now

            orthoView = component.getUtility(
                IOrthoView, name='OrthoView-%d' % (num + 1))

            for i in range(3):
                orthoView.renderPanes[
                    0].GetOrthoPlanes().SetPickedPlaneByNumber(i)
                orthoView.renderPanes[0]._IncrementPush(0)
                orthoView.renderPanes[0]._DecrementPush(0)

            orthoView.GetOrthoPlanes().Modified()
            orthoView.renderPanes[0].Render()
            component.getUtility(ICurrentViewportManager).Render()

            # Now propagate new min/max values in image
            if num == self.GetCurrentImageIndex():
                component.getUtility(
                    ICurrentImage).GetPointData().GetScalars().Modified()
                self.PropagateImageLevels()

            # update plot and histogram windows
            if self.IsPlotWindowVisible():
                self.probe.Modified()
                self.PlotGrayScaleValues()

            if self._histowindow is not None:
                if self._histowindow.IsShown():
                    self._histowindow.ClearPlot()

    def checkForSpatialDimension(self):
        """Returns true of the current image's last (e.g. 3rd) dimension is spacial

        Basically, this checks to see if we have a hyperstack or a 3D image loaded"""

        image = component.getUtility(ICurrentImage)
        dimensions = image.GetDimensionInformation()
        if dimensions[-1].GetTypeName() == 'Distance':
            return True
        else:
            return False

    def askForRadius(self, filter_name, radius, title='Filter radius', image3d=True):

        try:
            dlg = ChooseFilterRadiusDialogC.ChooseFilterRadiusDialogC(
                self, filter_name, '%d' % radius, image3d
            )

            if dlg.ShowModal() == wx.ID_OK:

                try:
                    radius = float(dlg.GetRadius())
                except ValueError:
                    s = "Invalid radius value - please enter a valid number"
                    logging.error(s)
                    dlg_err = wx.MessageDialog(
                        self, s, 'Error', wx.OK | wx.ICON_ERROR)
                    dlg_err.ShowModal()
                    dlg_err.Destroy()
                    return -1, True

                try:
                    image3d = bool(dlg.Get3DFilterCheckboxState())
                except:
                    logging.exception(
                        "on{0}Filter".format(filter_name.capitalize()))
                    return -1, True

            else:
                logging.debug(
                    "{0} filter cancelled".format(filter_name.capitalize()))
                return -1, True
        finally:
            dlg.Destroy()

        return radius, image3d

    def uniformFilter(self, radius=-1):
        """Apply a uniform smoothing filter"""

        image3d = self.checkForSpatialDimension()

        if radius == -1:
            radius, image3d = self.askForRadius('uniform', 3, image3d=image3d)

        if radius == -1:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.UniformFilter(image, radius, image3d=image3d)
                self.UpdateCurrentImage()

    def _gaussSmooth(self, radius=-1):
        """Smooth image with a gaussian filter"""

        image3d = self.checkForSpatialDimension()

        if radius == -1:
            radius, image3d = self.askForRadius(
                'Gaussian', self.gaussian_radius, image3d=image3d)

        if radius == -1:
            return

        self.gaussian_radius = radius

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                image = obj.GaussianFilter(
                    image, self.gaussian_radius, image3d=image3d)
                if image3d:
                    # VTK filter generates a new input
                    self.SetInput(image)
                else:
                    # scipy filter runs in-place
                    self.UpdateCurrentImage()
                event.notify(WindowLevelPresetEvent('Auto-adjust (99%)'))

                # Notify everyone that the current input image has changed
                event.notify(CurrentImageChangeEvent(
                    self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), self.GetCurrentImageTitle()))

    def maximumFilter(self, radius=-1):
        """Perform maximum filter on image"""

        image3d = self.checkForSpatialDimension()

        if radius == -1:
            radius, image3d = self.askForRadius('maximum', 3, image3d=image3d)

        if radius == -1:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.MaximumFilter(image, radius, image3d=image3d)
                self.UpdateCurrentImage()

    def minimumFilter(self, radius=-1):
        """Perform minimum filter on image"""

        image3d = self.checkForSpatialDimension()

        if radius == -1:
            radius, image3d = self.askForRadius('minimum', 3, image3d=image3d)

        if radius == -1:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.WindowDisabler():
            with wx.BusyCursor():
                image = component.getUtility(ICurrentImage)
                obj.MinimumFilter(image, radius, image3d=image3d)
                self.UpdateCurrentImage()

    def _medianFilter(self, radius=-1):
        """Perform median filter on image"""

        image3d = self.checkForSpatialDimension()

        if radius == -1:
            radius, image3d = self.askForRadius(
                'median', self.median_filter_kernel_size, image3d=image3d)

        if radius == -1:
            return

        self.median_filter_kernel_size = int(radius)

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.MedianFilter(component.getUtility(
                ICurrentImage), self.median_filter_kernel_size, image3d=image3d)
            self.SetInput(image)

    def onAnisotropicDialogUpdateButton(self, evt, dlg):
        """manage update button press in anisotropic filter dialog"""

        # get current dialog settings
        res = dlg.GetResults()
        try:

            # get current z-slice image
            slicePlane = component.getUtility(
                ICurrentOrthoPlanes).GetPlanes()[0]
            image = slicePlane.GetOutput()

            with wx.BusyCursor():
                image2 = component.getUtility(IImageProcessor).AnisotropicFilter(image,
                                                                                 int(res[0]), float(
                                                                                     res[1]), float(res[2]),
                                                                                 faces_on=res[3], edges_on=res[4], corners_on=res[5])

                # export to an RGB array
                exporter = vtkImageExportToArray.vtkImageExportToArray()

                colours = vtk.vtkImageMapToColors()
                colours.SetOutputFormatToRGB()
                colours.SetLookupTable(slicePlane.GetLookupTable())
                colours.SetInput(image2.GetRealImage())
                exporter.SetInput(colours.GetOutput())
                arr = numpy.squeeze(exporter.GetArray())
                h, w = arr.shape[0], arr.shape[1]

                # flip image
                arr = arr[::-1, :, :]
                # convert to a wx image
                wx_image = wx.EmptyImage(w, h)
                wx_image.SetData(arr.tostring())

                # rescale
                w, h = dlg.m_bitmapPreview.GetSize()
                wx_image = wx_image.Scale(w, h, wx.IMAGE_QUALITY_HIGH)

                # convert to bitmap
                wxBitmap = wx_image.ConvertToBitmap()
        except:
            w, h = dlg.m_bitmapPreview.GetSize()
            wxBitmap = wx.EmptyBitmap(w, h)
            logging.error("AnisotropicDialog")

        # paste into GUI
        dlg.m_bitmapPreview.SetBitmap(wxBitmap)

    def _anisotropicFilter(self):
        """Perform anisotropic smoothing (edge-preserving) on image"""

        res = None
        try:
            dlg = AnisotropicSmoothingDialogC.AnisotropicSmoothingDialogC(self)
            dlg.set_update_callback(self.onAnisotropicDialogUpdateButton)

            if dlg.ShowModal() == wx.ID_OK:
                res = dlg.GetResults()
        finally:
            dlg.Destroy()

        if res is None:
            return

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        # TODO: we still don't do anything with the 'Gradient' vs. 'Pixel'
        # radio button

        with wx.BusyCursor():
            image = obj.AnisotropicFilter(component.getUtility(
                ICurrentImage), int(res[0]), float(res[1]), float(res[2]),
                faces_on=res[3], edges_on=res[4], corners_on=res[5])
            self.SetInput(image)

    def _gradientMagnitudeFilter(self):
        """Compute the image gradient magnitude"""

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.ImageGradientMagnitudeFilter(component.getUtility(
                ICurrentImage))
            self.SetInput(image)

    def _laplacianFilter(self):
        """Compute the image Laplacian"""

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.ImageLaplacianFilter(component.getUtility(
                ICurrentImage))
            self.SetInput(image)

    def ImageInvert(self):
        """Invert image"""

        logging.info("Inverting image...")

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():
            image = obj.ImageInvert(component.getUtility(ICurrentImage))
            self.SetInput(image)
            event.notify(WindowLevelPresetEvent('Auto-adjust (99%)'))

    def SetInput(self, image):

        # get image all by it's lonesome
        with wx.WindowDisabler():
            with wx.BusyCursor():
                # update image
                image.Update()
                image.GetPointData().GetScalars().AddObserver(
                    'ModifiedEvent', lambda e, o, s=self: s.PropagateImageLevels())
                # disconnect pipeline (this line may change with VTK-6)
                # image.SetSource(None)
                self.UpdateDimensionInfo()

        # disconnect reader
        r = component.getUtility(ICurrentMicroViewInput).GetReader()
        if r:
            logging.warning("Reader still exists!")
            r.SetOutput(None)

        # if image is a bare vtkImageData object, we must wrap it
        if isinstance(image, vtk.vtkImageData):
            logging.warning("bare image passed in - code needs to be updated!")
            image = MVImage.MVImage(image)

        # register the image
        component.getGlobalSiteManager().registerUtility(
            image, IImage, name='Image-%d' % (self.GetCurrentImageIndex() + 1))
        component.getGlobalSiteManager().registerUtility(image, ICurrentImage)

        # wire up to the current orthoview
        self.SetOrthoInput(image)

        # Notify everyone that the current input image has changed
        event.notify(CurrentImageChangeEvent(
            self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), self.GetCurrentImageTitle()))

        # Perform garbage collection
        gc.collect()

    def SetOrthoInput(self, image, orthoView=None, mviewIn=None):

        if not orthoView:
            orthoView = component.getUtility(ICurrentOrthoView)
        if not mviewIn:
            mviewIn = component.getUtility(ICurrentMicroViewInput)

        orthoView.SetInputData(image, mviewIn)

        # adjust toolbar
        self.setModePaletteToolState(image)

        # delete low-level reader
        mviewIn.GetReader().clear_reader()

    def GetInput(self):
        return component.getUtility(ICurrentImage)

    def ConvertToGrayScale(self):
        """Convert Image to grayscale representation"""

        logging.info("Converting image to grayscale...")

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        with wx.BusyCursor():

            image = obj.ConvertToGrayScale(
                component.getUtility(ICurrentImage))
            self.SetInput(image)

    def DownsampleImage(self, nbits=8):
        """Downsample the displayed image to 8-bit."""

        logging.info("Downsampling image to %d bits" % nbits)

        # get image processing object
        obj = component.getUtility(IImageProcessor)

        # first, set up pipeline to downsample the current image
        with wx.BusyCursor():
            image = obj.DownsampleImage(
                component.getUtility(ICurrentImage), nbits)

        # now wire it in as the current image
        self.SetInput(image)

    def PlotHistogramValues(self, evt):

        # If we hit this code via the app menu, we use our notion of what the
        # current image is
        image_index = self.GetCurrentImageIndex()

        if evt:
            # If this event object is a vtkAtamai.Event object, it means we got
            # here via the keyboard
            if hasattr(evt, 'pane'):
                image_index = evt.pane.GetImageIndex() - 1

        # bail out immediately if an image transform has been loaded
        t = component.getUtility(ICurrentOrthoView).GetImageTransform()

        if t is not None:
            dlg = wx.MessageDialog(
                self, "MicroView's plotting routines do not support transformed images!", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return False
        else:
            with wx.BusyCursor():
                # get the two lists for histogram

                roistats = self.GetROIStats(image_index)
                if self._histowindow is None:
                    roistats.SetBinSize(1.0)

                x, y = roistats.GetHistogram()
                if len(x) == 0:
                    return

                # Plot the histogram of the frequency of gray scale values
                self.createHistogramWindow()

                self.ShowHistoWindow()

                self._histowindow.SetHistogramData(x, y)
                spacing = component.getUtility(ICurrentImage).GetSpacing()
                self._histowindow.SetVoxelDimensions(
                    spacing[0], spacing[1], spacing[2])
                self._histowindow.PlotData()
                self._histowindow.SetFileName(self.GetFileName())
                self._histowindow.Reset()
                self._histowindow.Show()

            return 0

    def SetFileName(self, filename):
        component.getUtility(
            ICurrentViewportManager).GetPageState().SetFileName(filename)

    def GetFileName(self):
        return component.getUtility(ICurrentViewportManager).GetPageState().GetFileName()

    def GetPrintableFileName(self):

        filename = self.GetFileName()
        if filename:
            filename = filename.encode(sys.getfilesystemencoding() or 'UTF-8')
        return filename

    def ReadTransformFile(self, evt):
        """Read and apply an .xfm transform to all loaded images"""

        r = vtk.vtkMNITransformReader()
        curr_dir = self.GetCurrentDirectory()
        ft = collections.OrderedDict(
            [("MNI Matrix Transformation", ["*.xfm"]), ("All files", ["*"])])
        filename = EVSFileDialog.askopenfilename(
            filetypes=ft, defaultdir=curr_dir, message='Choose Transform file...')

        # return now if user didn't select a file
        if not filename:
            return -1

        curr_dir = os.path.dirname(os.path.abspath(filename))
        self.SaveCurrentDirectory(curr_dir)

        r.SetFileName(filename)
        t = r.GetTransform()

        # invert transform
        if t is None:
            dlg = wx.MessageDialog(
                self, "Unable to read transform from file", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
        else:
            t = t.GetInverse()

        self.SetImageTransform(t)

        return 0

    def SetImageTransform(self, t):
        """Applies a given vtk transform to all displayed images"""

        component.getUtility(ICurrentOrthoView).SetImageTransform(t)
        self.UpdateMenus()

    def DeleteImageTransform(self):
        """Restores default transform"""

        o = component.getUtility(ICurrentOrthoView)
        if o.GetImageTransform():
            component.getUtility(ICurrentOrthoView).SetImageTransform(None)

        self.UpdateMenus()

    def OnFileOpen(self, filename=None, reset=True, do_spect=True, create_tab=False):
        """Opens and reads a support file. Returns False on error, True otherwise."""

        mviewIn = MicroViewIO.MicroViewInput()

        if not mviewIn.OpenFile(filename=filename):
            return False

        ft = mviewIn.GetFileType()

        if ft == MicroViewIO.FileType.plugin:
            # we've been tossed an egg.  Copy it into the user's plugin
            # directory
            _dir = os.path.join(
                appdirs.user_data_dir("MicroView", "Parallax Innovations"), "Plugins")
            short = os.path.split(filename)[-1]
            try:
                if not os.path.exists(_dir):
                    os.makedirs(_dir)
                shutil.copy(filename, os.path.join(_dir, short))
                logging.info("Plugin '%s' installed in '%s'." % (short, _dir))
            except:
                logging.exception("MicroViewMain")
                dlg = wx.MessageDialog(self, "Unable to install plugin '%s' into '%s'" % (
                    short, _dir), 'Error', wx.OK | wx.ICON_ERROR)
                dlg.ShowModal()
                dlg.Destroy()
                return True
            dlg = wx.MessageDialog(self, "Plugin '%s' installed in user directory.\nPlease restart MicroView to load plugin" % (
                short), 'Plugin Installed', wx.OK | wx.ICON_INFORMATION)
            dlg.ShowModal()
            dlg.Destroy()

        elif ft == MicroViewIO.FileType.license:
            # we've been tossed an license.  Copy it into the user's license
            # directory
            _dir = appdirs.user_data_dir("MicroView", "Parallax Innovations")
            short = 'license.conf'
            try:
                if not os.path.exists(_dir):
                    os.makedirs(_dir)
                shutil.copy(filename, os.path.join(_dir, short))
                logging.info("License '%s' installed in '%s'." % (short, _dir))
            except:
                logging.exception("MicroViewMain")
                dlg = wx.MessageDialog(self, "Unable to install license '%s' into '%s'" % (
                    short, _dir), 'Error', wx.OK | wx.ICON_ERROR)
                dlg.ShowModal()
                dlg.Destroy()
                return True
            dlg = wx.MessageDialog(self, "License '%s' installed in user directory.\nPlease restart MicroView" % (
                short), 'License Installed', wx.OK | wx.ICON_INFORMATION)
            dlg.ShowModal()
            dlg.Destroy()

        elif ft == MicroViewIO.FileType.image:

            image = mviewIn.GetImageOutput()

            # Sanity check
            if not image.IsA('vtkImageData'):
                dlg = wx.MessageDialog(self, "No image was read from this file.  The file was read successfully, but produced a '%s' object, rather than a 'vtkImageData' object." %
                                       image.GetClassName(), 'Error', wx.OK | wx.ICON_ERROR)
                dlg.ShowModal()
                dlg.Destroy()
                return

            self.CreateNewImage(image, mviewIn, filename=mviewIn.GetFileName(), title=mviewIn.GetTitle(),
                                reset=reset, do_spect=do_spect, create_tab=create_tab)
        elif ft == MicroViewIO.FileType.spreadsheet:
            # TODO: move this code into MicroViewIO
            data = mviewIn.GetSpreadsheetOutput()
            self.CreateNewSpreadsheetTab(
                data, mviewIn, filename=mviewIn.GetFileName())
            # Keep a history of recent files
            self._history.AddFileToHistory(mviewIn.GetFileName())
            logging.info("Loaded spreadsheet from file: %s" %
                         mviewIn.GetFileName())

        return True

    def OnFileOpenInNewTab(self, evt):
        self.OnFileOpen(create_tab=True)

    def OnNewTab(self, evt):
        """Create a new tab with a default image"""
        self.CreateImageTab()
        self.DisplayView("All")

    def OnReload(self, page):
        """Reload current image from disk"""
        filename = page.GetPageState().GetFileName()
        if filename:
            self.OnFileOpen(filename)

    def OnRename(self, page):
        """Rename current image"""

        state = page.GetPageState()
        idx = page.GetImageIndex()

        # split off #:
        num, title = state.GetTitle().split(' ', 1)

        dlg = wx.TextEntryDialog(
            self, 'Image {} rename:'.format(num[:-1]),
            'Enter new title for image', 'Rename')

        dlg.SetValue(title)
        try:
            if dlg.ShowModal() == wx.ID_OK:
                value = dlg.GetValue()
            else:
                return
        finally:
            dlg.Destroy()

        new_title = num + " " + value
        state.SetTitle(new_title)
        self._view_notebook.SetPageText(page.GetImageIndex() - 1, new_title)

        # construct a filename that matches title
        filename = state.GetFileName()
        if not filename:
            _dir = self.GetCurrentDirectory()
        else:
            _dir = os.path.dirname(filename)

        new_filename = os.path.join(_dir, value)
        state.SetFileName(new_filename)

        mviewIn = component.getUtility(
            IMicroViewInput, name='MicroViewInput-%d' % idx)
        mviewIn.SetFileName(new_filename)

    def OnDuplicateImage(self, page):
        """Clone image into new tab"""

        # sanity check
        if not IViewerState.providedBy(page.GetPageState()):
            return

        idx = page.GetImageIndex()
        image = component.getUtility(IImage, name='Image-%d' % idx)
        filename = page.GetPageState().GetFileName()

        # increase reference count on image
        image.IncrementReferenceCount()

        self.CreateNewImage(
            image, filename=filename, reset=True, create_tab=True)

    def CreateNewSpreadsheetTab(self, data, mviewIn, filename):
        tabname = os.path.split(filename)[-1]

        sheet_bm = self._stockicons.getMenuBitmap("icon_spreadsheet")

        try:
            # we will set up spreadsheet here
            if filename.lower().endswith('csv'):

                # sniff
                try:
                    with open(filename, 'rb') as _file:
                        dialect = unicodecsv.Sniffer().sniff(_file.read(1024))
                except:
                    # we assume comma delimited if we can't guess
                    delimiter = ','
                    dialect = unicodecsv.excel

                # handle CSV
                f = unicodecsv.reader(open(filename, 'rb'), dialect=dialect)

                # get header
                header = f.next()

                # create panel for spreadsheet - if we got this far, the csv
                # file is probably okay

                # Increase spreadsheet counter
                self._numberOfSpreadsheets += 1
                self._numberOfCurrentSpreadsheetsDisplayed += 1

                spreadsheet = SpreadsheetGUIC.SpreadsheetGUIC(
                    self, index=self._numberOfSpreadsheets)
                self._view_notebook.AddPage(
                    spreadsheet, tabname, select=True, bitmap=sheet_bm)

                component.getGlobalSiteManager().registerUtility(
                    spreadsheet, ISpreadsheet, name='Spreadsheet-%d' % self._numberOfSpreadsheets)

                self.SetupSpreadsheetHeader(spreadsheet, header)
                row = 0
                ncols = spreadsheet.m_grid.GetNumberCols()
                nerrs = 0
                with wx.BusyCursor():
                    for value in f:
                        col = 0
                        # add row if needed
                        if spreadsheet.m_grid.GetNumberRows() <= row:
                            spreadsheet.m_grid.InsertRows(row)
                        for v in value:
                            if col >= ncols:
                                if nerrs > 5:
                                    logging.warning(
                                        "Ignoring spurious input on line {0}".format(row))
                                    nerrs += 1
                            else:
                                spreadsheet.m_grid.SetCellValue(row, col, v)
                            col += 1
                        row += 1

            elif os.path.splitext(filename.lower())[-1] in (".xls", ".xlsx"):
                # handle Excel format

                if filename.lower().endswith('.xlsx'):
                    reader = XLSXReader(filename)
                else:
                    reader = XLSReader(filename)

                # adjust number of columns
                # for now, generate fake header values
                header = map(str, range(reader.ncols))

                # create panel for spreadsheet - if we got this far, the Excel
                # file is probably okay

                # Increase spreadsheet counter
                self._numberOfSpreadsheets += 1
                self._numberOfCurrentSpreadsheetsDisplayed += 1

                spreadsheet = SpreadsheetGUIC.SpreadsheetGUIC(
                    self, index=self._numberOfSpreadsheets)
                self._view_notebook.AddPage(
                    spreadsheet, tabname, select=True, bitmap=sheet_bm)

                component.getGlobalSiteManager().registerUtility(
                    spreadsheet, ISpreadsheet, name='Spreadsheet-%d' % self._numberOfSpreadsheets)

                self.SetupSpreadsheetHeader(spreadsheet, header)
                # adjust number of rows
                spreadsheet.m_grid.AppendRows(reader.nrows - 1)

                for row in range(reader.nrows):
                    for col in range(reader.ncols):

                        # returns value, font_idx, fgcolour, bgcolour
                        val, font, fg, bg = reader.get_cell(row, col)

                        if font:
                            spreadsheet.m_grid.SetCellFont(row, col, font)
                        if fg:
                            spreadsheet.m_grid.SetCellTextColour(row, col, fg)
                        if bg:
                            spreadsheet.m_grid.SetCellBackgroundColour(
                                row, col, bg)

                        if not isinstance(val, (str, unicode)):
                            val = str(val)

                        spreadsheet.m_grid.SetCellValue(row, col, val)

            spreadsheet.m_grid.AutoSizeColumns(True)
            spreadsheet.m_grid.AutoSizeRows(True)

            # set spreadsheet page info
            page_state = spreadsheet.GetPageState()
            page_state.SetFileName(filename)
            page_state.SetTitle(os.path.split(filename)[-1])

            return True
        except:
            logging.exception("CreateNewSpreadsheetTab")
            dlg = wx.MessageDialog(self, "Unable to load '%s'" %
                                   filename, 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return False

    def SetupSpreadsheetHeader(self, spreadsheet, header):
        # add or remove columns
        if len(header) > 26:
            spreadsheet.m_grid.AppendCols(len(header) - 26)
        elif len(header) < 26:
            spreadsheet.m_grid.DeleteCols(len(header), 26 - len(header))

        # adjust the column titles - at the moment we limit things to 26
        col = 0
        for val in header:
            spreadsheet.m_grid.SetColLabelValue(col, val)
            col += 1

    def CreateNewImage(self, image, mviewIn=None, filename='', title=None, header=None, reset=True, do_spect=True,
                       create_tab=False):

        if mviewIn is None:
            mviewIn = MicroViewIO.MicroViewInput()

        # if image is a bare vtkImageData object, we must wrap it
        if isinstance(image, vtk.vtkImageData):
            logging.warning("bare image passed in - code needs to be updated!")
            image = MVImage.MVImage(image)

        if header is not None:
            image.SetHeader(header)
        else:
            header = image.GetHeader()

        # create a new tab if needed

        # if there aren't any windows open, create a tab
        if self._view_notebook.GetPageCount() == 0:
            create_tab = True

        if create_tab:
            self.CreateImageTab(image, mviewIn, filename=filename, title=title)
            viewerState = component.getUtility(
                IViewportManager, name='ViewportManager-%d' % self._numberOfImages).GetPageState()
        else:
            # -----------------------------------------------------------------------
            # this should be refactored along with duplicate code in
            # CreateImageTab(), but we'll test here first

            # release memory from current image (TODO: handle multiple views of
            # same image)
            try:
                current_image = component.getUtility(ICurrentImage)
                current_image.DecrementReferenceCount()
            except:
                pass

            component.getGlobalSiteManager().registerUtility(
                image, IImage, name='Image-%d' % (self.GetCurrentImageIndex() + 1))
            component.getGlobalSiteManager().registerUtility(
                image, ICurrentImage)
            component.getGlobalSiteManager().registerUtility(
                mviewIn, IMicroViewInput, name='MicroViewInput-%d' % (self.GetCurrentImageIndex() + 1))
            component.getGlobalSiteManager().registerUtility(
                mviewIn, ICurrentMicroViewInput)

            viewerState = component.getUtility(IViewportManager, name='ViewportManager-%d' % (
                self.GetCurrentImageIndex() + 1)).GetPageState()
            if filename:
                viewerState.SetFileName(filename)
            orthoView = component.getUtility(
                IOrthoView, name='OrthoView-%d' % (self.GetCurrentImageIndex() + 1))

            with wx.WindowDisabler():
                with wx.BusyCursor():
                    image.Update()
                    self.SetOrthoInput(
                        image, orthoView=orthoView, mviewIn=mviewIn)
                    # TODO: fix next two lines for VTK-6 compatibility
                    # mviewIn.GetReader().Update()
                    #orthoView.SetInputData(image, mviewIn)

            if filename:
                if title is None:
                    title = os.path.split(filename)[-1]
                tabname = '%d: %s' % (self.GetCurrentImageIndex() + 1, title)
            else:
                tabname = '%d: Default View' % (
                    self.GetCurrentImageIndex() + 1)

            viewerState.SetTitle(tabname)
            self.SetWindowTitle(filename, tabname)

            self.PropagateImageLevels()

            # adjust scrollbar enable/disable
            numc = image.GetNumberOfScalarComponents()
            zrange = image.GetDimensions()[2]
            self.SetWindowLevelState(numc, zrange)

            # Notify everyone that the current input image has changed
            event.notify(CurrentImageChangeEvent(
                self.GetCurrentImageIndex(), self.GetNumberOfImagesCurrentlyLoaded(), self.GetCurrentImageTitle()))
            # -----------------------------------------------------------------

        # Delete any previous image transform
        self.DeleteImageTransform()

        if (self._plotwindow):
            pane = self._mgr.GetPane(self._plotwindow)
            pane.Caption(self._plotwindow_title)
        if (self._histowindow):
            pane = self._mgr.GetPane(self._histowindow)
            pane.Caption(self._histowindow_title)

        if image.GetMeasurementUnit() == 'mm':
            self._config.bMeasurementUnitIsMM = True
        else:
            self._config.bMeasurementUnitIsMM = False

        rank = image.GetRank()

        if reset:
            if rank == 2:
                viewerState.DisplayAllViews = False
                self.DisplayView('pane2d_1')
            else:
                # what sort of image stack is this?
                dim_info = image.GetDimensionInformation()

                if isinstance(dim_info[2], MVImage.Distance):
                    viewerState.DisplayAllViews = True
                    self.DisplayView(viewerState.layout)
                else:
                    viewerState.DisplayAllViews = False
                    self.DisplayView('pane2d_1')

                # Make sure all three planes are visible
                # TODO: I think this is wrong...
                component.getUtility(ICurrentOrthoView).GetOrthoPlanes().AllVisible(
                    self.pane3D.GetRenderer())

        # set limit/value of z-slice slider
        zrange = image.GetDimensions()[2]
        self.WindowLevelToolbar.GetSliceSlider().SetRange(0, zrange - 1)

        # set default lookup table
        self.SetLookupTable(
            image, viewerState, component.getUtility(ICurrentOrthoView))

        return self._numberOfImages

    def SetLookupTable(self, image, pageState, orthoView):
        """Set modality-specific lookup table"""

        dicom_header = image.GetDICOMHeader()
        if dicom_header.Modality in ('NM', 'PT', 'ST'):
            # set default colour table
            idx = 9  # GE 256-colour table for PET/SPECT etc.
        else:
            idx = 0  # Standard grey-scale colour table
        event.notify(ChangeImageLUT(orthoView, image, pageState, idx))

    def ShowLineSegment(self, e=None):
        """Shows the line segment"""
        origin = component.getUtility(ICurrentImage).GetOrigin()
        spacing = component.getUtility(ICurrentImage).GetSpacing()
        extent = component.getUtility(ICurrentImage).GetWholeExtent()
        x = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetSagittalPlane().GetSlicePosition()
        y = origin[1] + extent[1] / 2 * spacing[1]
        y0 = y + extent[1] / 3 * spacing[1]
        y1 = y - extent[1] / 3 * spacing[1]
        z = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetAxialPlane().GetSlicePosition()
        component.getUtility(
            ICurrentOrthoView).lineSegment.SetPoint(0, x, y0, z)
        component.getUtility(
            ICurrentOrthoView).lineSegment.SetPoint(1, x, y1, z)
        self.pane3D.Render()

    def ClearLineSegment(self, evt, orthoName=None):
        """Clears the line segment"""

        if self._plotwindow:
            self._plotwindow.ClearPlot()

        try:
            if orthoName is None:
                orthoView = component.getUtility(ICurrentOrthoView)
            else:
                orthoView = component.getUtility(IOrthoView, name=orthoName)
            orthoView.lineSegment.ClearLineSegment()
            orthoView.lineSegmentAnnotate.SetText(" ")
            orthoView.renderPanes[0].Render()
        except interface.interfaces.ComponentLookupError:
            # If there's no current image, do nothing
            pass

    def LineSegmentObserver(self, obj, evt):

        # Schedule a plot window update

        # we'd like to give regular plot updates without bring the system
        # to it's knees - we'll guarantee the user a redraw at least every
        # half second
        dt = time.time() - self._last_delayed_plot_redraw

        # schedule an update in 0.25 second
        if self._delayed_plot_update_cb:
            self._delayed_plot_update_cb.cancel()

        if dt > 1.0:
            # perform redraw immediately
            self.doScheduledPlotRedraw()
        else:
            # schedule for later
            self._delayed_plot_update_cb = reactor.callLater(
                0.25, self.doScheduledPlotRedraw)

    def doScheduledPlotRedraw(self):

        self._delayed_plot_update_cb = None
        self._last_delayed_plot_redraw = time.time()

        # Update plot window, if a plot window exists
        if self.IsPlotWindowVisible():
            self.PlotGrayScaleValues()

    def ClearROI(self, e=None):
        """Clears the ROI."""

        # If we hit this code via the app menu, we use our notion of what the
        # current image is
        image_index = self.GetCurrentImageIndex()

        if e:
            # If this event object is a vtkAtamai.Event object, it means we got
            # here via the keyboard
            if hasattr(e, 'pane'):
                image_index = e.pane.GetImageIndex() - 1

        # get current image
        image = component.getUtility(ICurrentImage)

        # get stencil data and owner
        stencil_data = image.GetStencilData()
        owner = image.GetStencilDataOwner()

        if stencil_data:
            if owner == "histogram":
                logging.debug("Disabling histogram ROI")
                # act as a proxy for the histogram tool
                self._histowindow.HideHighlight(image_index)
                # set stencil data for current image to None
                image.SetStencilData(None)
                # send event
                event.notify(ROIDisabledEvent("histogram", image_index))
            else:
                self.DisableROIPlugin(image_index)

    def DisableROIPlugin(self, image_index):
        image = component.getUtility(
            IImage, name='Image-%d' % (image_index + 1))
        plugin_name = image.GetStencilDataOwner()
        if plugin_name:
            plugin = self.pluginManager.GetPluginReferenceByName(plugin_name)
            if plugin:
                try:
                    plugin().DisableROI(image_index)
                except:
                    logging.exception("MicroViewMain")

    def ConfigMenuItemsForROI(self, state):

        self.menubar.FindItemById(self.menubar.FindMenuItem(
            'File', 'Save Crop region...')).Enable(state)
        self.menubar.FindItemById(self.menubar.FindMenuItem(
            'File', 'Save Crop coordinates')).Enable(state)
        self.menubar.FindItemById(self.menubar.FindMenuItem(
            'File', 'Save Image values...')).Enable(state)
        self.menubar.FindItemById(self.menubar.FindMenuItem(
            'Tools', 'Histogram...')).Enable(state)

    def CheckImageSize(self):
        """Check the size of self.ImageData to see
        if it fits into memory. This seems only an issue for
        windows only.
        """
        if sys.platform != 'win32':
            return True

        # 64 bit check - this should cause x64 to ignore code after this check
        if sys.maxsize > (2 ** 31 - 1):
            return True

        if self._the_3GB_switch:
            max_volume_size = 16040000  # (kilo bytes)
        else:
            max_volume_size = 840000  # (kilo bytes)

        component.getUtility(ICurrentImage).UpdateInformation()
        component.getUtility(ICurrentImage).PropagateUpdateExtent()
        # check the size of volume to see if we can hanle it
        size = component.getUtility(ICurrentImage).GetEstimatedMemorySize()
        if size > max_volume_size:
            message = "The input volume size exceeds the maximum limit\n" + \
                      "of this application (%d MB).\n\n" % (max_volume_size / 1000) + \
                      "Please use SliceView to pick a VOI or save as 8-bit volume\n" + \
                      "and then load the output of SliceView into MicroView."
            dlg = wx.MessageDialog(
                self, message, 'Info', wx.OK | wx.ICON_INFORMATION)
            dlg.ShowModal()
            dlg.Destroy()
            return False

        return True

    def PropagateImageLevels(self, image=None):
        """Update all VTK objects that depend on image min/max and actual image
            data"""

        if image is None:
            image = component.getUtility(ICurrentImage)
        (minval, maxval) = image.GetScalarRange()

        self.UpdateWindowAndLevelValues(minval, maxval)

    def UpdateCurrentImage(self):
        """Image has been modified - force a redraw"""

        component.getUtility(ICurrentImage).GetRealImage().Modified()

        for plane in component.getUtility(ICurrentOrthoPlanes).GetPlanes():
            plane.Modified()
            plane.GetImageReslice().Modified()
            plane.GetImageReslice().Update()

        # Notify everyone that the current input image has changed
        self.PropagateImageLevels()
        event.notify(WindowLevelPresetEvent('Auto-adjust (99%)'))
        event.notify(CurrentImageChangeEvent(
            self.GetCurrentImageIndex(),
            self.GetNumberOfImagesCurrentlyLoaded(),
            self.GetCurrentImageTitle()))

        component.getUtility(ICurrentOrthoView).Render()

    def UpdateDimensionInfo(self):

        self.dims = component.getUtility(ICurrentImage).GetDimensions()
        self.rank = 3
        for i in self.dims:
            if i == 1:
                self.rank -= 1

    def UpdateWindowAndLevelValues(self, minval, maxval):

        currentViewerState = component.getUtility(
            ICurrentViewportManager).GetPageState()

        _min, _max = currentViewerState.table_range

        self.modePalettePane.SetDataRange((minval, maxval))
        self.WindowLevelToolbar.SetImageDataRange((minval, maxval))

        # TODO: refactor this
        if self.bIsMicroView:
            currentOrthoView = component.getUtility(ICurrentOrthoView)
            currentOrthoView.GetLookupTable().SetTableRange(_min, _max)

    def Quit(self, evt):
        """Quits MicroView"""

        # persist application state
        self._persistMgr.SaveAndUnregister()

        # Start by giving plugin manager the lead - let pluginmanager shutdown
        # gracefully
        if not self.pluginManager.Quit():
            logging.info("Quit aborted by plugin manager")
            return

        # persist various objects
        try:
            self._config.save()    # a wx.Config object
            # a python shelf
            MicroViewObjectShelve.MicroViewObjectShelve.getObject(
            ).getRoot().close()
        except:
            logging.warning("Unable to commit change to database")

        # close all notebook pages
        windows = []
        for page_idx in range(self._view_notebook.GetPageCount()):
            windows.append(self._view_notebook.GetPage(page_idx))

        self.Freeze()
        for window in windows:
            try:
                page_idx = self._view_notebook.GetPageIndex(window)
                self.OnNotebookPageClose(None, page_index=page_idx)
                self._view_notebook.DeletePage(page_idx)
            except:
                # log exceptions, but press on with shutdown
                logging.exception("MicroViewMain")
        self.Thaw()

        if reactor.running:
            reactor.stop()

    def showMacAddresses(self, evt):
        "Displays mac address information"

        self.PushStatusText("Checking hardware...")

        with wx.BusyCursor():
            m = ethernet.getMACAddresses()

        self.PopStatusText()

        _msg = "Current Ethernet MAC addresses:\n\n"
        for nic in m:
            if nic != '00:00:00:00:00:00':
                _msg = _msg + "\t" + nic + "\n"

        dlg = wx.MessageDialog(
            self, _msg, 'Network Information', wx.OK | wx.ICON_INFORMATION)
        dlg.ShowModal()
        dlg.Destroy()

    def openPluginDirectory(self, evt):
        """Open file browser in plugin directory"""

        _dir = os.path.join(
            appdirs.user_data_dir("MicroView", "Parallax Innovations"), "Plugins")

        if not os.path.exists(_dir):
            try:
                os.makedirs(_dir)
            except:
                pass

        if sys.platform == 'darwin':
            open_cmd = 'open'
        elif sys.platform == 'linux2':
            open_cmd = 'xdg-open'

        if sys.platform == 'win32':
            os.startfile(_dir)
        else:
            subprocess.call([open_cmd, _dir])

    def showTKeyHint(self, evt):
        s = 'Press the "t" key over the 3D viewport, or any of the 2D viewports to save a snapshot to a file.'

        dlg = wx.MessageDialog(self, s, 'Info', wx.OK | wx.ICON_INFORMATION)
        dlg.ShowModal()
        dlg.Destroy()

    def saveScene(self, evt):
        """Save current 3D viewport to a rendering scene file"""

        # list of exporters that work
        vals = [('out.x3d', vtk.vtkX3DExporter),
                ('out.vrml', vtk.vtkVRMLExporter), ('out.iv', vtk.vtkIVExporter)]

        for (filename, classname) in vals:

            exp = classname()

            # the render window
            #filename = 'out.x3d'
            #exp = vtk.vtkX3DExporter()
            exp.SetFileName(filename)
            exp.SetRenderWindow(self.pane3D.GetRenderWindow())
            with wx.BusyCursor():
                exp.Write()

    def showAbout(self, evt):
        """Displays information about MicroView."""
        if self._aboutDialog is None:

            if sys.platform == 'win32':
                plat = 'Windows'
            elif sys.platform.startswith('linux'):
                plat = 'Linux'
            else:
                plat = 'OS X'

            if sys.maxsize == 2 ** 31 - 1:
                plat += ' (32-bit)'
            else:
                plat += ' (64-bit)'

            self._aboutDialog = wx.AboutDialogInfo()
            self._aboutDialog.Name = "MicroView"
            self._aboutDialog.Version = MicroView.PACKAGE_VERSION
            self._aboutDialog.Copyright = "Copyright (c) 2000-2008 GE Healthcare\nCopyright (c) 2011-2015 Parallax Innovations Inc."

            s = "MicroView is a 2D and 3D image viewer, based on wxPython and VTK.\n\n" + \
                "Platform: {0}\n" + \
                "VTK Version: {1}\n" + \
                "wx Version: {2}\n" + \
                "SHA1: {3}"
            self._aboutDialog.SetDescription(wx.lib.wordwrap.wordwrap(s.format(plat,
                                                                               vtk.vtkVersion().GetVTKVersion(),
                                                                               wx.version(),
                                                                               PACKAGE_SHA1), 350, wx.ClientDC(self)))
            self._aboutDialog.WebSite = (
                "http://www.parallax-innovations.com/microview", "MicroView site")
            self._aboutDialog.Developers = [
                "Jeremy Gill", "Andrew Kennedy", "Hua Qian", "Brad Arthur",
                "Faruq Ladak", "Peter Feiner"]

            licenseText = """
            See license.txt for license details.
            """

            self._aboutDialog.License = wx.lib.wordwrap.wordwrap(
                licenseText, 500, wx.ClientDC(self))

        # Then we call wx.AboutBox giving it that info object
        wx.AboutBox(self._aboutDialog)

        return 0

    def DisplayView(self, PaneName='All', uselayout=1, image_num=None):

        if image_num is None:
            viewport_manager = component.getUtility(ICurrentViewportManager)
            orthoView = component.getUtility(ICurrentOrthoView)
        else:
            viewport_manager = component.getUtility(
                IViewportManager, name='ViewportManager-%d' % image_num)
            orthoView = component.getUtility(
                IOrthoView, name='OrthoView-%d' % image_num)

        viewerState = viewport_manager.GetPageState()

        if viewerState.DisplayAllViews and uselayout:
            PaneName = viewerState.layout

        # get panes
        pane3D = orthoView.renderPanes[0]
        pane2d_1 = orthoView.renderPanes[1]
        pane2d_2 = orthoView.renderPanes[2]
        pane2d_3 = orthoView.renderPanes[3]

        # Arrange the panes in the standard grid
        if PaneName == 'All':
            viewport_manager.SetViewTo1Side3((pane3D,
                                              pane2d_1,
                                              pane2d_2,
                                              pane2d_3), label=PaneName)

        # Arrange the panes in a 2x2 grid
        elif PaneName == 'All_2x2':
            viewport_manager.SetViewTo2By2((pane3D,
                                            pane2d_1,
                                            pane2d_2,
                                            pane2d_3), label=PaneName)

        # Display the 3D pane
        elif PaneName == 'pane3D':
            viewport_manager.SetViewToSingle(pane3D, label=PaneName)

        # Display pane2d_1
        elif PaneName == 'pane2d_1':
            viewport_manager.SetViewToSingle(pane2d_1, label=PaneName)

        # Display pane2d_2
        elif PaneName == 'pane2d_2':
            viewport_manager.SetViewToSingle(pane2d_2, label=PaneName)

        # Display pane2d_3
        elif PaneName == 'pane2d_3':
            viewport_manager.SetViewToSingle(pane2d_3, label=PaneName)

        try:
            if image_num is None:
                image = component.getUtility(ICurrentImage)
            else:
                image = component.getUtility(
                    IImage, name='Image-%d' % image_num)

            if (image.GetDimensions()[2] <= 1):
                viewerState.DisplayAllViews = False
            elif uselayout:
                viewerState.DisplayAllViews = not viewerState.DisplayAllViews
            else:
                if (PaneName == 'All') or (PaneName == 'All_2x2'):
                    viewerState.DisplayAllViews = False
                else:
                    viewerState.DisplayAllViews = True
        except:
            logging.exception('MicroViewMain')

    # Enable or disable the time point navigation menu commands, depending
    # on whether or not there are images for time points prior/after the
    # currently displayed time point.
    def _setTimePointNavigationState(self):
        menu = self.menubar.component('Visualize' + self._menu_append)
        state = 'normal'
        if (self._nImageIdx < 1):
            state = 'disable'
        menu.entryconfig(menu.index('Previous Time Point'), state=state)

        state = 'normal'
        if (self.mviewIn.TimePointCount() - 1 <= self._nImageIdx):
            state = 'disable'
        menu.entryconfig(menu.index('Next Time Point'), state=state)

    # This is a generic handler for a change in the time point to be
    # displayed.  This functionality only applies when there is an
    # imported that has multiple time points (like a perfusion experiment).
    #
    # JDG - this routine shouldn't need to be here -- must follow up with Del
    def _handleTimePointChange(self):
        self._setTimePointNavigationState()
        self.mviewIn.SetCurrentTimePoint(self._nImageIdx)

        if self.mviewIn.GetImportedImage() != (None, None):
            # abort any previous histogram calculation
            image, title = self.mviewIn.GetImportedImage()
            component.getGlobalSiteManager().registerUtility(
                image, IImage, name='Image-%d' % (self.GetCurrentImageIndex() + 1))
            component.getGlobalSiteManager().registerUtility(
                image, ICurrentImage)

            # check image size before load it into memory
            if self.CheckImageSize():
                return False

            wlTable = component.getUtility(ICurrentOrthoView).GetLookupTable()
            _min, _max = wlTable.GetTableRange()
            self.ProcessImage()
            wlTable.SetTableRange(_min, _max)  # why is this needed??

            self.title.SetFileName(self.mviewIn.GetImportedFileName() + '...')
            component.getUtility(ICurrentViewportManager).Render()

    # Navigate to and display the image for the previous time point.
    def _viewPreviousTimePoint(self):
        self._nImageIdx = self._nImageIdx - 1
        self._handleTimePointChange()
        return 0

    # Navigate to and display the image for the next time point.
    def _viewNextTimePoint(self):
        self._nImageIdx += 1
        self._handleTimePointChange()
        return 0

    def OnFileImport(self, evt):
        """Import a stack of 2D images"""

        mviewIn = MicroViewIO.MicroViewInput()
        mviewIn.ShowImportDialog()

        image, title = mviewIn.GetImportedImage()

        if image is not None:

            # create a new tab
            filename = mviewIn.GetImportedFileName()
            self.CreateImageTab(
                image, mviewIn, filename=filename, title=filename + '...')
            viewerState = component.getUtility(
                IViewportManager, name='ViewportManager-%d' % self._numberOfImages).GetPageState()

            # Delete any previous image transform
            self.DeleteImageTransform()

            if self._histowindow:
                self._histowindow.ClearPlot()

            if image.GetMeasurementUnit() == 'mm':
                self._config.bMeasurementUnitIsMM = True
            else:
                self._config.bMeasurementUnitIsMM = False

            if (self.rank == 2):
                viewerState.DisplayAllViews = False
                self.DisplayView('pane3D')
            else:
                # what sort of image stack is this?
                dim_info = image.GetDimensionInformation()
                if isinstance(dim_info[2], MVImage.Distance):
                    viewerState.DisplayAllViews = True
                    self.DisplayView(viewerState.layout)
                else:
                    viewerState.DisplayAllViews = False
                    self.DisplayView('pane2d_1')

                # Make sure all three planes are visible
                # TODO: I think this is wrong...
                component.getUtility(ICurrentOrthoView).GetOrthoPlanes().AllVisible(
                    self.pane3D.GetRenderer())

    def OnFileExport(self, evt):
        """Export a stack of 2D images"""
        self.mviewOut.ExportImage()

    def OnDICOMDIRExport(self, evt):
        """Export current image to a DICOMDIR folder"""
        self.mviewOut.ExportDICOMDIR()

    def OnDICOMDIRRegenerate(self, evt):
        """Generate/regenerate a DICOMDIR index for a disk tree"""
        self.mviewOut.DICOMDIRRegenerate()

    def SetImageReslice(self, imagereslice):
        self.imagereslice = imagereslice

    def HandleVTKProgressEvent(self, obj, evt):
        """
        A VTK object generated a progress event - convert it to a zope-style
        event
        """

        event.notify(ProgressEvent(
            obj.GetProgressText(), obj.GetProgress()))

    @component.adapter(ProgressEvent)
    def ProgressMethod(self, evt):

        if evt._progress_percentage > 1:
            logging.error(
                evt._progress_text + "(progress value is larger than 1.)")
            evt._progress_percentage = 1.0
            self.SetStatusText('Ready')
        elif evt._progress_percentage < 0:
            logging.error(
                evt._progress_text + "(progress value is smaller than 0.)")
            evt._progress_percentage = 0.0
            self.SetStatusText('Ready')

        if math.isnan(evt._progress_percentage):
            logging.error("progress value not a number!!")
        else:
            self.SetStatusText(
                evt._progress_text or 'progress text missing...')
            self.SetStatusText('%d %%' % int(
                evt._progress_percentage * 100.0), 1)

        if evt._progress_percentage == 1:
            self.SetStatusText('Ready')
            self.SetStatusText('', 1)


    def SetResampleImage(self, resampleimage):
        self.ResampleImage = resampleimage

    def showHelp(self, l=''):
        # if no specific link has been chosen, try to determine which plugin
        # page to jump to

        if l == '':
            cur = self.pluginManager.GetActivePluginName()
            if cur:
                plugin = self.pluginManager.GetPluginReferenceByName(cur)
                l = plugin().GetHelpPageLink()
        l = self._helpDict.GetHelpPageLink(l)

        if l == '':
            l = "index.html"

        url = 'http://microview.parallax-innovations.com/docs/' + l

        # fall back is to use a webbrowser
        webbrowser.open(url)

        return 0

    def Reset(self, e=None):
        """Reset the view panes back to their original state"""

        component.getUtility(ICurrentOrthoView).ResetView()
        return 0

    def SaveCropRegion(self, filename=''):
        """Save cropped image to a file"""

        # due to Ubuntu brain-deadness, we can get here without a rotated image
        t = component.getUtility(ICurrentOrthoView).GetImageTransform()
        stencil = self.GetROIStencilData()

        if stencil is None:
            dlg = wx.MessageDialog(
                self, "No ROI selected.  Please define a region of interest before using this function.", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return

        if t is not None:
            self.mviewOut.SetImageTransform(t.GetInverse())

        return self.mviewOut.SaveSubVolume(filename, stencil_data=stencil)
        # TODO: examine VTK-6 following 4 lines
        # if self._ImageTransform:
        #    self.mviewOut.SetImageTransform(self._ImageTransform.GetInverse())
        # self.mviewOut.SetROIStencilData(self.GetROIStencilData())
        # return self.mviewOut.SaveSubVolume(filename)

    def OnLoadCropCoords(self, evt):
        """Load a crop coordinate file and create ROI"""

        curr_dir = self.GetCurrentDirectory()
        ft = collections.OrderedDict(
            [("SubVolumeCoordinates file", ["*.xml"]), ("All files", ["*"])])
        filename = EVSFileDialog.askopenfilename(
            filetypes=ft, defaultdir=curr_dir)

        # return now if user didn't select a file
        if not filename:
            return -1

        curr_dir = os.path.dirname(os.path.abspath(filename))
        self.SaveCurrentDirectory(curr_dir)

        mviewIn = component.getUtility(ICurrentMicroViewInput)
        return mviewIn.LoadVolumeCoordinatesFromDisk(filename)

    def SaveCropRegionCoordinates(self, filename=None):
        """Save the crop coordinates to a file"""
        return self.mviewOut.SaveSubVolumeCoordinatesToDisk(filename, stencil=self.GetROIStencilData())

    def SaveCropRegionCoordinatesSpecifyName(self):
        """Save the crop coordinates to a file"""

        ft = collections.OrderedDict([("XML files", ["*.xml"])])
        filename = EVSFileDialog.asksaveasfilename(
            message='Save Crop Coordinates', filetypes=ft)
        if filename:
            ret = self.mviewOut.SaveSubVolumeCoordinatesToDisk(
                filename, stencil=self.GetROIStencilData())

    def HandleROIKeyEvent(self, evt):
        """Handle Atamai keyboard event - rebroadcast as a zope event"""

        # what window has current focus?
        _win = self.FindFocus()

        # tell plugin manager to activate any plugins that respond to ROI keys
        self.pluginManager.ActivateStandardROIPlugins(bShouldSelect=False)

        # restore focus
        _win.SetFocus()

        event.notify(ROIKeyEvent(evt))

    def SaveAreaAsImage(self, evt=None):
        """Write the scene in a ROI to an image file provided the ROI is rectangular and 2D."""

        try:
            if self.GetROIType() != "box":
                raise Exception(
                    "SaveAreaAsImage: rectangular ROI must be selected!")
            wlTable = component.getUtility(ICurrentOrthoView).GetLookupTable()
            self.mviewOut.SetWLTable(wlTable)
            self.mviewOut.SaveAreaAsImage(stencil=self.GetROIStencilData())
        except Exception, e:
            msg = e.message
            logging.error(msg)
            dlg = wx.MessageDialog(
                self, msg, 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()

    def ImportRaw(self):
        """Import a raw image - ask the user for details of image"""
        pass

    def SaveAs(self, filename=None):
        """Save the entire image to a file in user chosen
        file type.
        """

        self.mviewOut.SetImageTransform(None)

        mviewIn = component.getUtility(ICurrentMicroViewInput)

        defaultfilename = mviewIn.GetFileName()
        if defaultfilename:
            defaultfilename = os.path.split(defaultfilename)[-1]
        ret = self.mviewOut.SaveVolume(
            filename=filename, defaultfilename=defaultfilename)

        if ret != 0:
            return

        viewerState = component.getUtility(ICurrentViewportManager).GetPageState()

        filename = self.mviewOut.GetFileName()
        viewerState.SetFileName(filename)
        viewerState.SetTitle(os.path.split(filename)[-1])

        self.SetFileName(self.mviewOut.GetFileName())
        self.SetWindowTitle(
            self.mviewOut.GetFileName(), viewerState.GetTitle())
        mviewIn.SetFileName(self.mviewOut.GetFileName())

    def SetWindowTitle(self, filename, title):
        """the window title is the file name"""

        self._plotwindow_title = 'MicroView - ' + (filename or '')
        self._histowindow_title = 'MicroView - ' + (filename or '')
        self.SetTitle('MicroView - ' + (filename or ''))
        self.SetRepresentedFilename(filename or '')
        if filename:
            shortname = title
#            shortname = '%d: %s' % (self.GetCurrentImageIndex() + 1, os.path.split(filename)[-1])
            viewport = component.getUtility(ICurrentViewportManager)
            viewport.GetPageState().SetTitle(shortname)
            idx = self._view_notebook.GetPageIndex(viewport)
            self._view_notebook.SetPageText(idx, shortname)

    def GenerateReorientedImage(self):

        # current image etc.
        mv = component.getUtility(IMicroViewMainFrame)

        image = self.GetCurrentImage()
        if image is None:
            return

        # set up transform
        transform = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetTransform()

        # create and setup reslice object
        reslice = vtk.vtkImageReslice()
        reslice.SetInterpolationModeToCubic()
        reslice.SetInput(image.GetRealImage())
        reslice.AutoCropOutputOn()
        reslice.SetResliceAxes(transform.GetMatrix())

        # set background value
        reslice.SetBackgroundLevel(mv.GetMaskValue())
        output_image = MVImage.MVImage(reslice.GetOutputPort(), input=image)
        patient_name = str(output_image.GetDICOMHeader().PatientName)

        if patient_name:
            patient_name += ' - Reoriented image'
        else:
            patient_name += 'Reoriented image'

        output_image.GetDICOMHeader().PatientName = patient_name

        return output_image

    def CreateNewImageFromReorientedImage(self):

        # first, create new image
        new_image = self.GenerateReorientedImage()

        # generate a new tab / filename
        mviewIn = component.getUtility(ICurrentMicroViewInput)
        filename = mviewIn.GetFileName()

        if not filename:
            # the selected image didn't come from a file - grab tab name
            # instead
            viewport = component.getUtility(ICurrentViewportManager)
            filename = viewport.GetPageState().GetTitle()
            # remove leading garbage
            pos = filename.find(': ')
            filename = filename[pos + 2:]

        if filename:
            temp = os.path.splitext(filename)
            filename = temp[0] + ' (Reoriented)' + temp[1]

        # now generate a new tab from this image
        self.CreateNewImage(
            new_image, filename=filename, reset=True, create_tab=True)

    def SaveReorientedImageAs(self, evt):
        """Save the re-oriented image image to a file."""

        # due to Ubuntu brain-deadness, we can get here without a rotated image
        t = component.getUtility(
            ICurrentOrthoView).GetOrthoPlanes().GetTransform()

        if t.GetOrientation() == (0.0, 0.0, 0.0):
            dlg = wx.MessageDialog(
                self, "No transformation applied.  Please change cutplane slice orientation before using this function.", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return

        # first, create new image
        new_image = self.GenerateReorientedImage()

        # create a sensible filename
        mviewIn = component.getUtility(ICurrentMicroViewInput)
        defaultfilename = mviewIn.GetFileName()
        if defaultfilename:
            temp = os.path.splitext(mviewIn.GetFileName())
            defaultfilename = temp[0] + ' (Reoriented)' + temp[1]

        self.mviewOut.SaveReorientedVolume(
            new_image, defaultfilename=defaultfilename)

    def SaveTransformedImageAs(self, filename=None):
        """Save the image transformed by user loaded matrix to a file."""

        try:

            t = component.getUtility(ICurrentOrthoView).GetImageTransform()

            self.mviewOut.SetImageTransform(t.GetInverse())
        except AttributeError:
            dlg = wx.MessageDialog(
                self, "No transformation applied.  Unable to save transformed image.", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return -1

           # create a sensible filename
        mviewIn = component.getUtility(ICurrentMicroViewInput)
        defaultfilename = mviewIn.GetFileName()
        if defaultfilename:
            temp = os.path.splitext(mviewIn.GetFileName())
            defaultfilename = temp[0] + ' (Transformed)' + temp[1]

        self.mviewOut.SaveVolume(
            filename=filename, defaultfilename=defaultfilename)
        return 0

    def SaveLineLengthToFile(self, evt):
        component.getUtility(ICurrentOrthoView).lineSegment.SaveLineLengthToFile(
            os.path.dirname(self.GetFileName()) + os.sep + "LineLength.txt")

    def GetROIStats(self, image_index=None):
        if image_index is None:
            image_index = self.GetCurrentImageIndex()
        self._ROIStats.SetROIStencilData(self.GetROIStencilData(image_index))
        return self._ROIStats

    def ShowMeanAndStdDeviation(self, evt):
        """Displays mean/stdev of selected ROI. Returns False if no ROI is selected."""

        # determine which image number our event is associated with (1 indexed)
        image_index = evt.pane.GetImageIndex() - 1
        image = component.getUtility(ICurrentImage)
        roi_stencil_data = image.GetStencilData()

        # only permit this calculation if some sort of ROI exists
        if roi_stencil_data is None:
            logging.warning("ShowMeanAndStdDeviation: no ROI set - ignoring")
            return False

        # VTK-6
        if vtk.vtkVersion().GetVTKMajorVersion() > 5:
            logging.debug("TODO: figure out how to update stencil data here")
        else:
            roi_stencil_data.Update()

        # bail out immediately if an image transform has been loaded
        t = component.getUtility(ICurrentOrthoView).GetImageTransform()

        if t is not None:
            dlg = wx.MessageDialog(
                self, "Function do not support transformed images!", 'Error', wx.OK | wx.ICON_ERROR)
            dlg.ShowModal()
            dlg.Destroy()
            return False

        results_logger = logging.getLogger('results')

        with wx.BusyCursor():
            stats = self.GetROIStats(image_index)
            Mean = stats.GetMeanValue()
            VoxelCount = stats.GetVoxelCount()
            StdDev = stats.GetStdDeviation()
            Min, Max = stats.GetValueRange()
            Total = stats.GetTotal()

            logging.info("Updating Image Statistics...")

            bounds = self.GetROIBounds(image_index)
            text = (
                "ROI: (" + "%0.4f, " * (len(bounds) - 1) + "%0.4f) ") % bounds

            for i in range(component.getUtility(ICurrentImage).GetNumberOfScalarComponents()):
                text += "Component: %d\tNum. Voxels: %d\tMin: %1.2f\tMax: %1.2f\tMean: %1.2f\tStd. Dev.: %1.2f\tTotal: %1.2f" % \
                    (i, VoxelCount, Min[i], Max[
                     i], Mean[i], StdDev[i], Total[i])
            results_logger.info(text)

            _time = time.ctime()

            filename = self.GetPrintableFileName()
            if not filename:
                filename = component.getUtility(
                    ICurrentViewportManager).GetPageState().GetTitle()

            for i in range(component.getUtility(ICurrentImage).GetNumberOfScalarComponents()):
                args = {'Time-stamp': _time, 'Filename/Title': filename or ''}
                b = self.GetROIBounds(image_index)
                args['X'] = (b[1] + b[0]) / 2.0
                args['Y'] = (b[3] + b[2]) / 2.0
                args['Z'] = (b[4] + b[4]) / 2.0
                args['ROI-Width-X'] = (b[1] - b[0])
                args['ROI-Width-Y'] = (b[3] - b[2])
                args['ROI-Width-Z'] = (b[5] - b[4])
                args['Chan.'] = i
                args['ROI-VoxelCount'] = VoxelCount
                args['ROI-Min'] = Min[i]
                args['ROI-Max'] = Max[i]
                args['ROI-Mean'] = Mean[i]
                args['ROI-Stdev'] = StdDev[i]
                args['ROI-Total'] = Total[i]

                if self._statswindow is None:
                    self.createStatisticsWindow()
                    self.ShowStatisticsWindow()

                self._statswindow.insertRow(args)
                self._statswindow.Layout()

        return [Mean, StdDev, Total]

    @component.adapter(HelpEvent)
    def OnUserHelp(self, evt):

        topic = evt.GetHelpTopic()

        if topic == 'import':
            self.showHelp('ar01s06.html#ImageImport')
        elif topic == 'export':
            self.showHelp('ar01s06.html#ImageExport')

    def GetWriter(self):
        return self.mviewOut._writer

    @component.adapter(MicroViewSettingsModifiedEvent)
    def OnConfigModified(self, evt):

        if evt.GetConfig().bUseMinMaxForWindowLevel:
            self.modePalettePane.SetUseMinMaxForWinLev(True)
            self.WindowLevelToolbar.UseMinMaxOn()
        else:
            self.modePalettePane.SetUseMinMaxForWinLev(False)
            self.WindowLevelToolbar.UseMinMaxOff()

        # adjust plot units
        if self._plotwindow:
            self._plotwindow.setUnitLabel(evt.GetConfig().GetUnitLabel())

    def OnAutoImport(self):
        """
        Import a stack of 2D images, but do so without prompting the user
        for any information.  This is done by using the "external selection"
        connection to acquire the set of files to import.
        """
        try:
            self.mviewIn.AutoImageImport()
        except:
            return False

        if self.mviewIn.GetImportedImage() != (None, None):
            # abort any previous histogram calculation
            image, title = self.mviewIn.GetImportedImage()

            component.getGlobalSiteManager().registerUtility(
                image, IImage, name='Image-%d' % (self.GetCurrentImageIndex() + 1))
            component.getGlobalSiteManager().registerUtility(
                image, ICurrentImage)

            # check image size before load it into memory
            if self.CheckImageSize():
                return False

            with wx.BusyCursor():
                self.ProcessImage()

            self._nImageIdx = 0
            self._setTimePointNavigationState()

            self.title.SetFileName(self.mviewIn.GetImportedFileName() + '...')

    def GetROIType(self, image_index=None):
        """
        ROI types: 'box', 'cylinder', 'serial', 'custom'
        """
        if image_index is None:
            image_index = self.GetCurrentImageIndex()
        if image_index == -1:
            return
        image = component.getUtility(
            IImage, name='Image-%d' % (image_index + 1))
        if not image:
            return
        stencil_owner = image.GetStencilDataOwner()
        if not stencil_owner:
            return
        plugin = self.pluginManager.GetPluginReferenceByName(stencil_owner)
        return plugin().GetROIType(image_index)

    def GetROIStencilData(self, image_index=None):
        """Get the current ROI vtkStencilData object for an image"""
        if image_index is None:
            image_index = self.GetCurrentImageIndex()
        return component.getUtility(IImage, name='Image-%d' % (image_index + 1)).GetStencilData()

    def GetROIExtent(self, image_index=None):
        if image_index is None:
            image_index = self.GetCurrentImageIndex()
        stencil_data = self.GetROIStencilData(image_index)
        if stencil_data:
            stencil_data.Update()
            return stencil_data.GetWholeExtent()
        else:
            return None

    def GetROIBounds(self, image_index=None):
        if image_index is None:
            image_index = self.GetCurrentImageIndex()

        e = self.GetROIExtent(image_index)

        # convert extent -> bounds
        image = component.getUtility(
            IImage, name='Image-%d' % (image_index + 1))
        b = [0, 0, 0, 0, 0, 0]
        s = image.GetSpacing()
        o = image.GetOrigin()

        for i in range(3):
            b[2 * i] = o[i] + s[i] * (e[2 * i] - 0.5)
            b[2 * i + 1] = o[i] + s[i] * (e[2 * i + 1] + 0.5)
        return tuple(b)

    def togglePanelSize(self, evt):

        panel = self._mgr.GetPane(self.leftframe)

        if panel.IsShown():
            panel.Hide()
        else:
            panel.Show()

        self._mgr.Update()

    def menuTogglePanelSize(self, evt):
        self.togglePanelSize(evt)

    def menuToggleShowModePalette(self, evt):

        panel = self._mgr.GetPane(self.modePalettePane)

        if panel.IsShown():
            panel.Hide()
        else:
            panel.Show()

        self._mgr.Update()

    def menuToggleShowWindowLevelPanel(self, evt):

        panel = self._mgr.GetPane(self.WindowLevelToolbar)

        if panel.IsShown():
            panel.Hide()
        else:
            panel.Show()

        self._mgr.Update()

    def ShowBottomWindow(self):

        panel = self._mgr.GetPane(self._bottom_notebook)
        if not panel.IsShown():
            panel.Show()
            self._mgr.Update()

    def menuHideBottomWindow(self, evt):

        panel = self._mgr.GetPane(self._bottom_notebook)
        panel.Hide()
        self._mgr.Update()

    def menuShowLog(self, evt):

        self.ShowBottomWindow()

        # now select log page
        index = self._bottom_notebook.GetPageIndex(self.logPane)
        self._bottom_notebook.SetSelection(index)

    def menuShowResults(self, evt):

        self.ShowBottomWindow()

        # now select log page
        index = self._bottom_notebook.GetPageIndex(self.resultsPane)
        self._bottom_notebook.SetSelection(index)

    def menuToggleShowStatisticsWindow(self, evt):

        if self._statswindow is None:
            self.createStatisticsWindow()

        panel = self._mgr.GetPane(self._bottom_notebook)

        if panel.IsShown():
            panel.Hide()
        else:
            panel.Show()

        self._mgr.Update()

        # now select stats page
        index = self._bottom_notebook.GetPageIndex(self._statswindow)
        self._bottom_notebook.SetSelection(index)

    def ShowStatisticsWindow(self):

        if self._statswindow is None:
            self.createStatisticsWindow()

        panel = self._mgr.GetPane(self._bottom_notebook)
        panel.Show()
        self._mgr.Update()

        # now select stats page
        index = self._bottom_notebook.GetPageIndex(self._statswindow)
        self._bottom_notebook.SetSelection(index)

    def ShowHistoWindow(self):

        panel = self._mgr.GetPane(self._bottom_notebook)
        if not panel.IsShown():
            panel.Show()
            self._mgr.Update()

        # now select histo page
        index = self._bottom_notebook.GetPageIndex(self._histowindow)
        if self._bottom_notebook.GetSelection() != index:
            self._bottom_notebook.SetSelection(index)

    def ShowPlotWindow(self):

        panel = self._mgr.GetPane(self._bottom_notebook)
        if not panel.IsShown():
            panel.Show()
            self._mgr.Update()

        # now select plot page
        index = self._bottom_notebook.GetPageIndex(self._plotwindow)
        if self._bottom_notebook.GetSelection() != index:
            self._bottom_notebook.SetSelection(index)

    def ShowSpectrumWindow(self):

        panel = self._mgr.GetPane(self._bottom_notebook)
        if not panel.IsShown():
            panel.Show()
            self._mgr.Update()

        # now select spectrum page
        index = self._bottom_notebook.GetPageIndex(self._spectrumwindow)
        if self._bottom_notebook.GetSelection() != index:
            self._bottom_notebook.SetSelection(index)

    def ShowResultsWindow(self):

        panel = self._mgr.GetPane(self._bottom_notebook)

        # now select results page
        index = self._bottom_notebook.GetPageIndex(self.resultsPane)
        self._bottom_notebook.SetSelection(index)

        panel.Show()
        self._mgr.Update()

    def IsPlotWindowVisible(self):

        if self._plotwindow is None:
            return False
        else:
            return (self._bottom_notebook.IsShown() and (self._bottom_notebook.GetCurrentPage() == self._plotwindow))

    def menuShowPlot(self, e=None):

        if self._plotwindow is None:
            self.createPlotWindow()

        self.ShowPlotWindow()

    def menuShowHistogram(self, e=None):

        if self._histowindow is None:
            self.createHistogramWindow()

        self.ShowHistoWindow()

    def menuToggleShowScales(self, evt):
        # post a zope event notification indicating that we've requested a change
        # in the visibility of 2D viewport scale actors
        event.notify(ScalesVisibilityChangeEvent(evt.IsChecked()))

    def menuToggleShowColourBar(self, evt):
        """Show/Hide colour bar on side of image"""
        event.notify(ColourBarVisibilityChangeEvent(evt.IsChecked()))

    def SetRegisterMode(self, bIsEnabled):

        self.RegisterToggleSideButton.Show(bIsEnabled)

        if not bIsEnabled:
            if self._registerSide == 'right':
                self.RegisterToggleSide()

    def SetRegisterLookupTable2(self, lut, dataRange):
        self._registerLookupTable2 = lut
        self._registerDataRange2 = dataRange

    def RegisterToggleSide(self):

        if not self._registerLookupTable2:
            return

        if self._registerSide == 'left':
            self.RegisterToggleSideButton.SetBitmap(
                self._register_right_bitmap)
            self.WindowLevelToolbar.SetLookupTable(self._registerLookupTable2)
            self.WindowLevelToolbar.SetDataRange(self._registerDataRange2)
            self._registerSide = 'right'
        else:
            self.RegisterToggleSideButton.SetBitmap(self._register_left_bitmap)
            self.WindowLevelToolbar.SetLookupTable(
                component.getUtility(ICurrentOrthoView).GetLookupTable())
            self.WindowLevelToolbar.SetDataRange(
                component.getUtility(ICurrentImage).GetScalarRange())
            self._registerSide = 'left'

############################################################
# Methods migrated from xmlrpc server
############################################################
    def DoCameraRoll(self, ang):
        camera = self.pane3D._Renderer.GetActiveCamera()
        camera.Roll(ang)
        self.pane3D.Render()
        return 0

    def DoCameraAzimuth(self, ang):
        """Rotate the camera, in the azimuthal direction, by ang degrees"""
        camera = self.pane3D._Renderer.GetActiveCamera()
        camera.Azimuth(ang)
        self.pane3D.Render()
        return 0

    def DoCameraYaw(self, ang):
        camera = self.pane3D._Renderer.GetActiveCamera()
        camera.Yaw(ang)
        self.pane3D.Render()
        return 0

    def DoCameraElevation(self, ang):
        camera = self.pane3D._Renderer.GetActiveCamera()
        camera.Elevation(ang)
        camera.OrthogonalizeViewUp()
        self.pane3D.Render()
        return 0

    def DoCameraPitch(self, ang):
        camera = self.pane3D._Renderer.GetActiveCamera()
        camera.Pitch(ang)
        self.pane3D.Render()
        return 0
